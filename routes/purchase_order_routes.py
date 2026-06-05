"""
purchase_order_routes.py
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
Flask blueprint for the Purchase Order module.

URL prefix : /purchase-order
Blueprint  : po_bp

Endpoints
â”€â”€â”€â”€â”€â”€â”€â”€â”€
LIST / DASHBOARD
  GET  /                              â†’ index (DataTable listing, with filters)
  GET  /dashboard                     â†’ status counts dashboard
  GET  /reports                       â†’ reports landing
  GET  /report/<kind>                 â†’ supplier-wise / item-wise / pending / etc.
  GET  /report/<kind>/export/<fmt>    â†’ Excel / PDF export

CRUD
  GET  /new                           â†’ create form
  POST /save                          â†’ create/update
  GET  /<int:po_id>/edit              â†’ edit form
  GET  /<int:po_id>/view              â†’ view PO
  POST /<int:po_id>/delete            â†’ soft-delete (admin only)

WORKFLOW
  POST /<int:po_id>/submit            â†’ Draft â†’ Pending Approval
  POST /<int:po_id>/approve           â†’ Pending â†’ Approved
  POST /<int:po_id>/reject            â†’ Pending â†’ Rejected
  POST /<int:po_id>/cancel            â†’ â†’ Cancelled
  POST /<int:po_id>/reopen            â†’ â†’ Draft

PDF / Print / Email / WhatsApp
  GET  /<int:po_id>/pdf               â†’ download PDF
  GET  /<int:po_id>/print             â†’ print-friendly HTML
  POST /<int:po_id>/email             â†’ email PDF to supplier
  POST /<int:po_id>/whatsapp          â†’ mark whatsapp sent (ready for API)

AJAX HELPERS
  GET  /api/suppliers?q=&po_type=     â†’ supplier lookup
  GET  /api/supplier/<int:id>         â†’ supplier full details
  GET  /api/items?q=&po_type=         â†’ item / material lookup
  GET  /api/item/<int:id>             â†’ item full details
  GET  /api/next-po-number?po_type=   â†’ next PO number preview
  GET  /api/dashboard-counts          â†’ JSON status counts
  GET  /api/list                      â†’ DataTables JSON feed
  GET  /api/default-terms?po_type=    â†’ fetch default T&C
  POST /api/save-default-terms        â†’ save / update default terms

MASTERS
  GET  /terms-master                  â†’ manage default terms
  GET  /company-settings              â†’ manage company settings
  POST /company-settings/save         â†’ save company settings
"""
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import json
import os
import re

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, abort, send_file, make_response, current_app,
)
from flask_login import login_required, current_user
from sqlalchemy import or_, and_, func, desc

from models import db
from models.purchase_order import (
    PurchaseOrder, PurchaseOrderItem, PurchaseOrderTerm,
    PurchaseOrderApprovalLog, PurchaseOrderStatusLog,
    PoDefaultTerm, PoShipLocation, CompanySettings,
    PO_STATUSES, PO_STATUS_DRAFT, PO_STATUS_PENDING, PO_STATUS_APPROVED,
    PO_STATUS_REJECTED, PO_STATUS_PARTIAL, PO_STATUS_COMPLETE, PO_STATUS_CANCEL,
    PO_STATUS_COLORS, PO_TYPES, PO_TYPE_PREFIX_SHORT,
)
from models.supplier import Supplier
from models.material import Material, MaterialType
from models.employee import StateMaster, CountryMaster

try:
    from audit_helper import audit
except Exception:                                       # graceful fallback
    def audit(*a, **kw): pass

po_bp = Blueprint('purchase_order', __name__, url_prefix='/purchase-order')

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Helpers
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _role():
    return getattr(current_user, 'role', 'viewer') or 'viewer'


def _username():
    return (getattr(current_user, 'full_name', None) or
            getattr(current_user, 'username', '') or 'system')


def _can(action):
    """Permission gate. Extend with proper RBAC later."""
    role = _role()
    if role == 'admin':
        return True
    if action in ('view', 'create', 'edit'):
        return True
    if action in ('approve', 'cancel') and role in ('manager', 'director', 'purchase_manager'):
        return True
    if action == 'delete' and role in ('admin',):
        return True
    # default: allow everything for now (project pattern matches supplier_routes)
    return True


def _fy_from_date(d):
    """Return Indian Financial Year for given date as 'YY-YY' (e.g. 26-27)."""
    if not d:
        d = date.today()
    if d.month >= 4:
        a, b = d.year, d.year + 1
    else:
        a, b = d.year - 1, d.year
    return f"{a % 100:02d}-{b % 100:02d}"


def _company():
    cs = CompanySettings.get_default()
    if cs:
        return cs
    # last-resort fallback
    return CompanySettings(
        company_name='HCP Wellness Pvt Ltd', short_code='HCP',
        gst_number='', state='Gujarat', state_code='24',
        bill_address='', ship_address='', jurisdiction='Ahmedabad',
    )


def _state_code(state_name):
    """Look up GST state_code from state_master by state name (case-insensitive).
    Returns '' if not found. Used for supplier/company state-code snapshots."""
    if not state_name:
        return ''
    name = state_name.strip()
    if not name:
        return ''
    row = StateMaster.query.filter(
        StateMaster.is_active == True,
        StateMaster.name.ilike(name)
    ).first()
    if row and row.state_code:
        return row.state_code
    # also try by short_name (e.g. "GJ" â†’ 24)
    row = StateMaster.query.filter(
        StateMaster.is_active == True,
        StateMaster.short_name.ilike(name)
    ).first()
    return (row.state_code if row else '') or ''


def _state_from_gst(gst_number):
    """Derive (state_name, state_code) from the first 2 digits of a GSTIN.
    GSTIN format: first 2 chars = numeric state code per Indian GST.
    e.g. '24AAFCH7246H1ZK' â†’ ('Gujarat', '24'). Returns ('', '') if unresolvable."""
    if not gst_number:
        return ('', '')
    code = gst_number.strip()[:2]
    if len(code) < 2 or not code.isdigit():
        return ('', '')
    row = StateMaster.query.filter(
        StateMaster.is_active == True,
        StateMaster.state_code == code
    ).first()
    if row:
        return (row.name or '', row.state_code or code)
    return ('', code)


def _striptags(s):
    """Strip HTML tags from a string (e.g. payment_terms from rich-text editor).
    Returns clean plain text. Safe for None/empty input."""
    if not s:
        return ''
    # Remove HTML tags
    txt = re.sub(r'<[^>]+>', ' ', str(s))
    # Decode common HTML entities
    txt = (txt.replace('&nbsp;', ' ')
              .replace('&amp;', '&')
              .replace('&lt;', '<')
              .replace('&gt;', '>')
              .replace('&quot;', '"')
              .replace('&#39;', "'"))
    # Collapse whitespace
    return re.sub(r'\s+', ' ', txt).strip()


def _next_po_serial(po_type, fy):
    """Get next yearly running serial for this PO type + FY."""
    max_serial = db.session.query(func.max(PurchaseOrder.po_serial)).filter(
        PurchaseOrder.po_type == po_type,
        PurchaseOrder.po_fy   == fy,
    ).scalar() or 0
    return int(max_serial) + 1


def _build_po_numbers(po_type, po_date, company):
    """
    Return (po_number_tally, po_number_short, serial, fy, year)

    Tally style : HCP/RM/PO-0001/26-27
    Short style : RMPO-2026-0001
    """
    fy = _fy_from_date(po_date)
    year = (po_date or date.today()).year
    serial = _next_po_serial(po_type, fy)

    short_code = (company.short_code or 'HCP').upper()
    short_prefix = PO_TYPE_PREFIX_SHORT.get(po_type, f'{po_type}PO')

    po_tally = f"{short_code}/{po_type}/PO-{serial:04d}/{fy}"
    po_short = f"{short_prefix}-{year}-{serial:04d}"
    return po_tally, po_short, serial, fy, year


def _to_decimal(v, default='0'):
    try:
        if v in (None, '', 'None'):
            return Decimal(default)
        return Decimal(str(v))
    except Exception:
        return Decimal(default)


def _round2(d):
    return Decimal(d).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _round3(d):
    return Decimal(d).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)


def _parse_date(s, fallback=None):
    if not s:
        return fallback
    s = s.strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return fallback


# â”€â”€ Number â†’ Words (INR) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_NUM_ONES = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven',
             'Eight', 'Nine', 'Ten', 'Eleven', 'Twelve', 'Thirteen',
             'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
_NUM_TENS = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy',
             'Eighty', 'Ninety']

def _two_digit(n):
    if n < 20:
        return _NUM_ONES[n]
    t, o = divmod(n, 10)
    return _NUM_TENS[t] + ((' ' + _NUM_ONES[o]) if o else '')


def _three_digit(n):
    h, rest = divmod(n, 100)
    out = ''
    if h:
        out = _NUM_ONES[h] + ' Hundred'
    if rest:
        out = (out + ' ' if out else '') + _two_digit(rest)
    return out


def amount_in_words(amount):
    """Indian numbering system (Crore/Lakh/Thousand)."""
    try:
        amt = float(amount or 0)
    except Exception:
        amt = 0.0
    sign = ''
    if amt < 0:
        sign = 'Minus '
        amt = abs(amt)
    rupees = int(amt)
    paise  = int(round((amt - rupees) * 100))

    if rupees == 0 and paise == 0:
        return 'INR Zero Only'

    parts = []
    crore   = rupees // 10000000;  rupees %= 10000000
    lakh    = rupees // 100000;    rupees %= 100000
    thousand= rupees // 1000;      rupees %= 1000
    if crore:    parts.append(_two_digit(crore)   + ' Crore')
    if lakh:     parts.append(_two_digit(lakh)    + ' Lakh')
    if thousand: parts.append(_two_digit(thousand)+ ' Thousand')
    if rupees:   parts.append(_three_digit(rupees))

    rs_words = ' '.join(parts).strip() or 'Zero'
    out = f'INR {sign}{rs_words}'
    if paise:
        out += f' and {_two_digit(paise)} Paise'
    out += ' Only'
    # Clean whitespace
    return ' '.join(out.split())


# â”€â”€ status log helper â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _log_status(po, to_status, note=''):
    db.session.add(PurchaseOrderStatusLog(
        po_id=po.id, from_status=po.status, to_status=to_status,
        actor_id=getattr(current_user, 'id', None),
        actor_name=_username(), note=note or '',
    ))


def _log_approval(po, level, action, comment=''):
    db.session.add(PurchaseOrderApprovalLog(
        po_id=po.id, level=level, action=action,
        actor_id=getattr(current_user, 'id', None),
        actor_name=_username(), actor_role=_role(),
        comment=comment or '',
    ))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# LISTING
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/')
@login_required
def index():
    if not _can('view'):
        abort(403)

    po_type = (request.args.get('po_type', '') or '').strip().upper()
    return render_template(
        'purchase_order/index.html',
        active_page='purchase_order',
        po_type=po_type,
        po_type_label=PO_TYPES.get(po_type, ''),
        po_types=PO_TYPES,
        po_statuses=PO_STATUSES,
        po_status_colors=PO_STATUS_COLORS,
        role=_role(),
        can_add=_can('create'),
        can_approve=_can('approve'),
        can_cancel=_can('cancel'),
        can_delete=_can('delete'),
    )


@po_bp.route('/api/list')
@login_required
def api_list():
    """JSON feed for the DataTables listing page."""
    q = PurchaseOrder.query.filter_by(is_deleted=False)

    # Filters from query string
    po_type   = (request.args.get('po_type', '') or '').strip().upper()
    status    = (request.args.get('status', '') or '').strip()
    supplier  = (request.args.get('supplier_id', '') or '').strip()
    date_from = _parse_date(request.args.get('date_from', ''))
    date_to   = _parse_date(request.args.get('date_to', ''))
    search    = (request.args.get('search', '') or '').strip()

    if po_type:   q = q.filter(PurchaseOrder.po_type == po_type)
    if status:    q = q.filter(PurchaseOrder.status  == status)
    if supplier:
        try:
            q = q.filter(PurchaseOrder.supplier_id == int(supplier))
        except Exception:
            pass
    if date_from: q = q.filter(PurchaseOrder.po_date >= date_from)
    if date_to:   q = q.filter(PurchaseOrder.po_date <= date_to)
    if search:
        like = f'%{search}%'
        q = q.filter(or_(
            PurchaseOrder.po_number.ilike(like),
            PurchaseOrder.po_number_short.ilike(like),
            PurchaseOrder.supplier_name.ilike(like),
            PurchaseOrder.supplier_gst.ilike(like),
            # NEW â€” also match by product (item) name / code / HSN
            PurchaseOrder.items.any(or_(
                PurchaseOrderItem.item_name.ilike(like),
                PurchaseOrderItem.item_code.ilike(like),
                PurchaseOrderItem.hsn_code.ilike(like),
                PurchaseOrderItem.description.ilike(like),
            )),
        ))

    rows = q.order_by(desc(PurchaseOrder.po_date), desc(PurchaseOrder.id)).limit(2000).all()
    data = [r.to_dict() for r in rows]

    # When a search term is active, also surface the matching line-items so the
    # frontend can auto-expand the row and show *which* item(s) matched.
    if search:
        like = f'%{search}%'
        for d, po in zip(data, rows):
            matched = [it for it in po.items.all()
                       if (it.item_name and search.lower() in (it.item_name or '').lower())
                       or (it.item_code and search.lower() in (it.item_code or '').lower())
                       or (it.hsn_code and search.lower() in (it.hsn_code or '').lower())
                       or (it.description and search.lower() in (it.description or '').lower())]
            d['matched_items'] = [{
                'sr_no'       : it.sr_no,
                'item_name'   : it.item_name or '',
                'item_code'   : it.item_code or '',
                'hsn_code'    : it.hsn_code or '',
                'uom'         : it.uom or '',
                'quantity'    : float(it.quantity or 0),
                'rate'        : float(it.rate or 0),
                'gst_pct'     : float(it.gst_pct or 0),
                'discount_pct': float(it.discount_pct or 0),
                'total_amount': float(it.total_amount or 0),
                'due_date'    : it.due_date.strftime('%d-%m-%Y') if it.due_date else '',
            } for it in matched]
            d['search_term'] = search

    return jsonify(data=data, count=len(data))


@po_bp.route('/api/<int:po_id>/items')
@login_required
def api_po_items(po_id):
    """Return the line items of a PO as JSON. Used by the listing's
    double-click â†’ expand-row preview."""
    if not _can('view'):
        abort(403)
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        return jsonify(error='PO has been deleted'), 410
    items = po.items.order_by(PurchaseOrderItem.sr_no).all()
    return jsonify(
        po_id        = po.id,
        po_number    = po.po_number,
        supplier     = po.supplier_name or '',
        grand_total  = float(po.grand_total or 0),
        item_count   = len(items),
        items        = [{
            'sr_no'        : it.sr_no,
            'item_name'    : it.item_name or '',
            'item_code'    : it.item_code or '',
            'hsn_code'     : it.hsn_code or '',
            'uom'          : it.uom or '',
            'quantity'     : float(it.quantity or 0),
            'rate'         : float(it.rate or 0),
            'gst_pct'      : float(it.gst_pct or 0),
            'discount_pct' : float(it.discount_pct or 0),
            'taxable_amount': float(it.taxable_amount or 0),
            'total_amount' : float(it.total_amount or 0),
            'due_date'     : it.due_date.strftime('%d-%m-%Y') if it.due_date else '',
            'received_qty' : float(it.received_qty or 0),
            'pending_qty'  : float(it.pending_qty or 0),
            'remark'       : it.remark or '',
        } for it in items],
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DASHBOARD
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/dashboard')
@login_required
def dashboard():
    if not _can('view'):
        abort(403)
    return render_template(
        'purchase_order/dashboard.html',
        active_page='purchase_order',
        po_types=PO_TYPES,
        po_statuses=PO_STATUSES,
        po_status_colors=PO_STATUS_COLORS,
        role=_role(),
    )


@po_bp.route('/api/dashboard-counts')
@login_required
def api_dashboard_counts():
    """Counts of POs by status (and per type)."""
    counts = {s: 0 for s in PO_STATUSES}
    amounts = {s: 0.0 for s in PO_STATUSES}
    rows = db.session.query(
        PurchaseOrder.status,
        func.count(PurchaseOrder.id),
        func.coalesce(func.sum(PurchaseOrder.grand_total), 0),
    ).filter(PurchaseOrder.is_deleted == False).group_by(PurchaseOrder.status).all()
    for status, c, amt in rows:
        if status in counts:
            counts[status] = int(c or 0)
            amounts[status] = float(amt or 0)

    # by-type breakdown
    type_breakdown = {}
    type_rows = db.session.query(
        PurchaseOrder.po_type,
        func.count(PurchaseOrder.id),
        func.coalesce(func.sum(PurchaseOrder.grand_total), 0),
    ).filter(PurchaseOrder.is_deleted == False).group_by(PurchaseOrder.po_type).all()
    for t, c, amt in type_rows:
        type_breakdown[t] = {'count': int(c or 0), 'amount': float(amt or 0),
                             'label': PO_TYPES.get(t, t)}

    # last 30 days
    cutoff = date.today() - timedelta(days=30)
    last30 = db.session.query(func.count(PurchaseOrder.id),
                              func.coalesce(func.sum(PurchaseOrder.grand_total), 0)
                              ).filter(PurchaseOrder.is_deleted == False,
                                       PurchaseOrder.po_date >= cutoff).first()

    return jsonify(
        counts=counts, amounts=amounts,
        type_breakdown=type_breakdown,
        last_30_days={'count': int(last30[0] or 0), 'amount': float(last30[1] or 0)},
        total={'count': sum(counts.values()), 'amount': sum(amounts.values())},
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# CREATE / EDIT FORM
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/new')
@login_required
def new_po():
    if not _can('create'):
        abort(403)

    po_type = (request.args.get('po_type', 'RM') or 'RM').strip().upper()
    if po_type not in PO_TYPES:
        po_type = 'RM'

    company = _company()
    po_tally, po_short, serial, fy, year = _build_po_numbers(po_type, date.today(), company)

    # Default terms (per type, fall-back to ALL)
    default_terms = PoDefaultTerm.query.filter(
        PoDefaultTerm.is_active == True,
        PoDefaultTerm.is_deleted == False,
        or_(PoDefaultTerm.po_type == po_type, PoDefaultTerm.po_type == 'ALL'),
    ).order_by(PoDefaultTerm.id).all()

    return render_template(
        'purchase_order/form.html',
        active_page='purchase_order',
        mode='new', po=None, items=[], terms=default_terms,
        po_type=po_type,
        preview_po_number=po_tally,
        preview_po_short=po_short,
        preview_fy=fy,
        po_types=PO_TYPES,
        po_statuses=PO_STATUSES,
        company=company,
        role=_role(),
    )


@po_bp.route('/<int:po_id>/edit')
@login_required
def edit_po(po_id):
    if not _can('edit'):
        abort(403)
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        flash('PO has been deleted.', 'danger')
        return redirect(url_for('purchase_order.index'))
    if not po.is_editable:
        flash(f'PO is locked (status: {po.status}). Cancel or re-open it to edit.', 'warning')
        return redirect(url_for('purchase_order.view_po', po_id=po.id))

    items = po.items.order_by(PurchaseOrderItem.sr_no).all()
    terms = po.terms.order_by(PurchaseOrderTerm.section, PurchaseOrderTerm.sort_order).all()
    return render_template(
        'purchase_order/form.html',
        active_page='purchase_order',
        mode='edit', po=po, items=items, terms=terms,
        po_type=po.po_type,
        preview_po_number=po.po_number,
        preview_po_short=po.po_number_short,
        preview_fy=po.po_fy,
        po_types=PO_TYPES,
        po_statuses=PO_STATUSES,
        company=_company(),
        role=_role(),
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SAVE  (create + update)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/save', methods=['POST'])
@login_required
def save_po():
    if not _can('create'):
        abort(403)

    po_id = request.form.get('po_id', '').strip()
    is_edit = bool(po_id)

    # â”€â”€ Header fields â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    po_type   = (request.form.get('po_type', 'RM') or 'RM').strip().upper()
    if po_type not in PO_TYPES:
        return jsonify(success=False, error='Invalid PO Type'), 400

    po_date   = _parse_date(request.form.get('po_date', ''), date.today())
    supplier_id = request.form.get('supplier_id', '').strip()
    if not supplier_id:
        return jsonify(success=False, error='Supplier is mandatory'), 400
    try:
        supplier_id = int(supplier_id)
    except Exception:
        return jsonify(success=False, error='Invalid supplier'), 400

    supplier = Supplier.query.get(supplier_id)
    if not supplier:
        return jsonify(success=False, error='Supplier not found'), 400

    # â”€â”€ Items (parallel arrays) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    item_names    = request.form.getlist('item_name[]')
    item_ids      = request.form.getlist('material_id[]')
    item_codes    = request.form.getlist('item_code[]')
    categories    = request.form.getlist('category[]')
    hsn_codes     = request.form.getlist('hsn_code[]')
    uoms          = request.form.getlist('uom[]')
    quantities    = request.form.getlist('quantity[]')
    rates         = request.form.getlist('rate[]')
    discount_pcts = request.form.getlist('discount_pct[]')
    gst_pcts      = request.form.getlist('gst_pct[]')
    due_dates     = request.form.getlist('due_date[]')
    remarks       = request.form.getlist('remark[]')
    descriptions  = request.form.getlist('description[]')

    # Filter empty rows
    valid_rows = []
    for i, name in enumerate(item_names):
        nm = (name or '').strip()
        qty = _to_decimal(quantities[i] if i < len(quantities) else 0)
        if not nm or qty <= 0:
            continue
        valid_rows.append(i)

    if not valid_rows:
        return jsonify(success=False,
                       error='At least one item with quantity > 0 is required'), 400

    company = _company()
    is_interstate = (supplier.billing_state or '').strip().lower() not in \
                    (company.state.strip().lower(), '')

    # â”€â”€ Header object â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if is_edit:
        po = PurchaseOrder.query.get(int(po_id))
        if not po or po.is_deleted:
            return jsonify(success=False, error='PO not found'), 404
        if not po.is_editable:
            return jsonify(success=False,
                           error=f'PO is locked (status: {po.status})'), 400

        old_snapshot = {
            'po_number': po.po_number, 'supplier_name': po.supplier_name,
            'grand_total': float(po.grand_total or 0),
            'status': po.status,
        }

        # If editing a Pending Approval PO, revert to Draft so the approver
        # knows it changed and the user must re-submit for approval.
        if po.status == PO_STATUS_PENDING:
            po.status = PO_STATUS_DRAFT
            # Log the status change
            try:
                db.session.add(PurchaseOrderStatusLog(
                    po_id=po.id,
                    from_status=PO_STATUS_PENDING,
                    to_status=PO_STATUS_DRAFT,
                    actor_id=getattr(current_user, 'id', None),
                    actor_name=_username(),
                    comment='PO edited after submission â€” reverted to Draft for re-submission',
                ))
            except Exception:
                pass  # log table may not exist in older deployments

        # Type/date changes can shift the PO number â€” only allow if still Draft
        # (we already enforced is_editable). Number itself is NOT regenerated.
        po.po_type        = po_type
        po.po_date        = po_date

        # Wipe old items & terms; we'll re-create
        # Note: relationship has order_by, so we can't call .delete() directly.
        # Iterate and delete each instance â€” also fires cascade hooks properly.
        for it in po.items.all():
            db.session.delete(it)
        for tm in po.terms.all():
            db.session.delete(tm)
        db.session.flush()
    else:
        po_tally, po_short, serial, fy, year = _build_po_numbers(po_type, po_date, company)
        po = PurchaseOrder(
            po_type=po_type,
            po_number=po_tally,
            po_number_short=po_short,
            po_serial=serial,
            po_fy=fy,
            po_year=year,
            po_date=po_date,
            supplier_id=supplier.id,                    # FK is NOT NULL â€” set on create
            supplier_name=supplier.supplier_name or '', # required field on schema
            status=PO_STATUS_DRAFT,
            created_by_id=getattr(current_user, 'id', None),
            created_by_name=_username(),
        )
        db.session.add(po)
        db.session.flush()
        old_snapshot = None

    # â”€â”€ Supplier snapshot â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    po.supplier_id           = supplier.id
    po.supplier_name         = supplier.supplier_name or ''
    po.supplier_address      = (request.form.get('supplier_address', '').strip()
                                or supplier.address or supplier.shipping_address or '')
    po.supplier_gst          = (request.form.get('supplier_gst', '').strip()
                                or supplier.gst_number or '')
    po.supplier_pan          = supplier.pan_number or ''
    po.supplier_state        = (request.form.get('supplier_state', '').strip()
                                or supplier.billing_state or '')
    po.supplier_state_code   = (request.form.get('supplier_state_code', '').strip()
                                or _state_code(supplier.billing_state)
                                or _state_from_gst(supplier.gst_number or '')[1])
    po.supplier_country      = (request.form.get('supplier_country', '').strip()
                                or supplier.billing_country or 'India')
    po.supplier_contact_person = (request.form.get('supplier_contact_person', '').strip()
                                  or supplier.contact_person or '')
    po.supplier_mobile       = (request.form.get('supplier_mobile', '').strip()
                                or supplier.phone or '')
    po.supplier_email        = (request.form.get('supplier_email', '').strip()
                                or supplier.email or '')

    # â”€â”€ Company snapshot â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    po.company_name    = company.company_name or 'HCP Wellness Pvt Ltd'
    po.company_gst     = company.gst_number or ''
    po.company_pan     = company.pan_number or ''
    po.company_address = company.bill_address or ''
    po.company_state   = company.state or 'Gujarat'
    po.company_state_code = (_state_code(company.state)
                             or company.state_code or '24')

    # â”€â”€ Delivery / dispatch â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    po.delivery_address  = (request.form.get('delivery_address', '').strip()
                            or company.ship_address or '')
    po.expected_delivery = _parse_date(request.form.get('expected_delivery', ''))
    po.transport_mode    = (request.form.get('transport_mode', '') or '').strip()
    po.dispatched_through= (request.form.get('dispatched_through', '') or '').strip()
    po.destination       = (request.form.get('destination', '') or '').strip()
    po.reference_no      = (request.form.get('reference_no', '') or '').strip()
    po.reference_date    = _parse_date(request.form.get('reference_date', ''))
    po.other_references  = (request.form.get('other_references', '') or '').strip()
    po.payment_terms     = _striptags(request.form.get('payment_terms', ''))
    try:
        po.credit_days = int(request.form.get('credit_days', '30') or '30')
    except Exception:
        po.credit_days = 30
    po.is_interstate = is_interstate

    po.declaration       = (request.form.get('declaration', '') or '').strip() or (
                            company.declaration_text or '')
    po.narration         = (request.form.get('narration', '') or '').strip()
    po.updated_by_name   = _username()

    # â”€â”€ Items â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    basic_total     = Decimal('0')
    discount_total  = Decimal('0')
    taxable_total   = Decimal('0')
    cgst_total      = Decimal('0')
    sgst_total      = Decimal('0')
    igst_total      = Decimal('0')
    qty_total       = Decimal('0')

    sr_no = 0
    for idx in valid_rows:
        sr_no += 1
        qty  = _to_decimal(quantities[idx])
        rate = _to_decimal(rates[idx] if idx < len(rates) else 0)
        disc_pct = _to_decimal(discount_pcts[idx] if idx < len(discount_pcts) else 0)
        gst_pct  = _to_decimal(gst_pcts[idx] if idx < len(gst_pcts) else 0)

        amount        = _round2(qty * rate)
        discount_amt  = _round2(amount * disc_pct / Decimal('100'))
        taxable       = _round2(amount - discount_amt)
        if is_interstate:
            igst = _round2(taxable * gst_pct / Decimal('100'))
            cgst = Decimal('0'); sgst = Decimal('0')
        else:
            half = _round2(taxable * gst_pct / Decimal('200'))
            cgst = half; sgst = half; igst = Decimal('0')
        tax_amount = _round2(cgst + sgst + igst)
        total      = _round2(taxable + tax_amount)

        material_id = None
        try:
            if idx < len(item_ids) and item_ids[idx]:
                material_id = int(item_ids[idx])
        except Exception:
            material_id = None

        line = PurchaseOrderItem(
            po_id          = po.id,
            sr_no          = sr_no,
            material_id    = material_id,
            item_code      = (item_codes[idx] if idx < len(item_codes) else '').strip(),
            item_name      = (item_names[idx] or '').strip(),
            description    = (descriptions[idx] if idx < len(descriptions) else '').strip(),
            category       = (categories[idx] if idx < len(categories) else '').strip(),
            hsn_code       = (hsn_codes[idx] if idx < len(hsn_codes) else '').strip(),
            uom            = (uoms[idx] if idx < len(uoms) else 'KG').strip() or 'KG',
            quantity       = qty,
            rate           = rate,
            discount_pct   = disc_pct,
            discount_amt   = discount_amt,
            gst_pct        = gst_pct,
            amount         = amount,
            taxable_amount = taxable,
            cgst_amount    = cgst,
            sgst_amount    = sgst,
            igst_amount    = igst,
            tax_amount     = tax_amount,
            total_amount   = total,
            due_date       = _parse_date(due_dates[idx] if idx < len(due_dates) else ''),
            remark         = (remarks[idx] if idx < len(remarks) else '').strip(),
            pending_qty    = qty,
        )
        db.session.add(line)

        basic_total    += amount
        discount_total += discount_amt
        taxable_total  += taxable
        cgst_total     += cgst
        sgst_total     += sgst
        igst_total     += igst
        qty_total      += qty

    # â”€â”€ Round-off & grand total â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    raw_grand = taxable_total + cgst_total + sgst_total + igst_total
    rounded   = raw_grand.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    round_off = _round2(rounded - raw_grand)
    grand     = _round2(rounded)

    po.basic_total    = _round2(basic_total)
    po.discount_total = _round2(discount_total)
    po.taxable_amount = _round2(taxable_total)
    po.cgst_total     = _round2(cgst_total)
    po.sgst_total     = _round2(sgst_total)
    po.igst_total     = _round2(igst_total)
    po.round_off      = round_off
    po.grand_total    = grand
    po.total_quantity = _round3(qty_total)
    po.amount_in_words= amount_in_words(grand)

    # â”€â”€ Terms â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    term_texts    = request.form.getlist('term_text[]')
    term_sections = request.form.getlist('term_section[]')
    t_sort = 0
    for i, txt in enumerate(term_texts):
        txt = (txt or '').strip()
        if not txt:
            continue
        t_sort += 1
        sec = (term_sections[i] if i < len(term_sections) else 'GENERAL').strip() or 'GENERAL'
        db.session.add(PurchaseOrderTerm(
            po_id=po.id, section=sec, sort_order=t_sort, text=txt,
        ))

    db.session.commit()

    if is_edit:
        new_snapshot = {
            'po_number': po.po_number, 'supplier_name': po.supplier_name,
            'grand_total': float(po.grand_total or 0),
        }
        try:
            audit('purchase_order', 'UPDATE',
                  record_id=po.id, record_label=po.po_number,
                  obj=po, old_dict=old_snapshot, new_dict=new_snapshot,
                  commit=True)
        except Exception:
            pass
    else:
        try:
            audit('purchase_order', 'INSERT',
                  record_id=po.id, record_label=po.po_number,
                  obj=po, commit=True)
        except Exception:
            pass

    return jsonify(success=True, po_id=po.id, po_number=po.po_number,
                   redirect_url=url_for('purchase_order.view_po', po_id=po.id))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# VIEW
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/<int:po_id>/view')
@login_required
def view_po(po_id):
    if not _can('view'):
        abort(403)
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        flash('PO has been deleted.', 'danger')
        return redirect(url_for('purchase_order.index'))

    items = po.items.order_by(PurchaseOrderItem.sr_no).all()
    terms = po.terms.order_by(PurchaseOrderTerm.section, PurchaseOrderTerm.sort_order).all()
    approval_logs = po.approval_logs.order_by(PurchaseOrderApprovalLog.created_at).all()
    status_logs   = po.status_logs.order_by(PurchaseOrderStatusLog.created_at).all()

    return render_template(
        'purchase_order/view.html',
        active_page='purchase_order',
        po=po, items=items, terms=terms,
        approval_logs=approval_logs, status_logs=status_logs,
        po_types=PO_TYPES,
        po_statuses=PO_STATUSES,
        po_status_colors=PO_STATUS_COLORS,
        company=_company(),
        role=_role(),
        can_approve=_can('approve'),
        can_cancel=_can('cancel'),
        can_delete=_can('delete'),
    )


@po_bp.route('/<int:po_id>/print')
@login_required
def print_po(po_id):
    """Print-friendly HTML version (Tally style)."""
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        abort(404)
    # Print is only available once the PO has been approved.
    if po.status not in (PO_STATUS_APPROVED, PO_STATUS_PARTIAL, PO_STATUS_COMPLETE):
        abort(403, description='Print available only after PO approval.')
    items = po.items.order_by(PurchaseOrderItem.sr_no).all()
    terms = po.terms.order_by(PurchaseOrderTerm.section, PurchaseOrderTerm.sort_order).all()
    return render_template(
        'purchase_order/print.html',
        po=po, items=items, terms=terms,
        company=_company(),
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DELETE / SOFT-DELETE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/<int:po_id>/delete', methods=['POST'])
@login_required
def delete_po(po_id):
    if not _can('delete'):
        return jsonify(success=False, error='Permission denied'), 403
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.status not in (PO_STATUS_DRAFT, PO_STATUS_REJECTED, PO_STATUS_CANCEL):
        return jsonify(success=False,
                       error=f'Cannot delete PO in {po.status} state'), 400

    po.is_deleted      = True
    po.deleted_at      = datetime.utcnow()
    po.deleted_by_name = _username()
    db.session.commit()
    try:
        audit('purchase_order', 'DELETE',
              record_id=po.id, record_label=po.po_number, obj=po, commit=True)
    except Exception:
        pass
    return jsonify(success=True)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# WORKFLOW
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/<int:po_id>/submit', methods=['POST'])
@login_required
def submit_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        return jsonify(success=False, error='PO has been deleted'), 400
    if po.status not in (PO_STATUS_DRAFT, PO_STATUS_REJECTED):
        return jsonify(success=False,
                       error=f'Cannot submit PO in {po.status} state'), 400

    _log_status(po, PO_STATUS_PENDING, 'Submitted for approval')
    _log_approval(po, 'User', 'SUBMITTED',
                  comment=request.form.get('comment', ''))
    po.status            = PO_STATUS_PENDING
    po.submitted_at      = datetime.utcnow()
    po.submitted_by_id   = getattr(current_user, 'id', None)
    po.submitted_by_name = _username()
    db.session.commit()
    return jsonify(success=True, status=po.status)


@po_bp.route('/<int:po_id>/approve', methods=['POST'])
@login_required
def approve_po(po_id):
    if not _can('approve'):
        return jsonify(success=False, error='Permission denied'), 403
    po = PurchaseOrder.query.get_or_404(po_id)
    if not po.can_approve:
        return jsonify(success=False,
                       error=f'Cannot approve PO in {po.status} state'), 400

    level   = (request.form.get('level', 'Manager') or 'Manager').strip()
    comment = (request.form.get('comment', '') or '').strip()

    _log_approval(po, level, 'APPROVED', comment=comment)
    _log_status(po, PO_STATUS_APPROVED, f'Approved by {level}')

    po.status      = PO_STATUS_APPROVED
    po.is_locked   = True

    if level.lower() == 'director':
        po.director_approved_by_id   = getattr(current_user, 'id', None)
        po.director_approved_by_name = _username()
        po.director_approved_at      = datetime.utcnow()
    else:
        po.approved_by_id   = getattr(current_user, 'id', None)
        po.approved_by_name = _username()
        po.approved_at      = datetime.utcnow()

    db.session.commit()
    try:
        audit('purchase_order', 'APPROVE',
              record_id=po.id, record_label=po.po_number,
              detail=f'Approved by {_username()} ({level})',
              obj=po, commit=True)
    except Exception:
        pass
    return jsonify(success=True, status=po.status)


@po_bp.route('/<int:po_id>/reject', methods=['POST'])
@login_required
def reject_po(po_id):
    if not _can('approve'):
        return jsonify(success=False, error='Permission denied'), 403
    po = PurchaseOrder.query.get_or_404(po_id)
    if not po.can_approve:
        return jsonify(success=False,
                       error=f'Cannot reject PO in {po.status} state'), 400

    reason = (request.form.get('reason', '') or '').strip()
    if not reason:
        return jsonify(success=False, error='Rejection reason is required'), 400

    _log_approval(po, 'Manager', 'REJECTED', comment=reason)
    _log_status(po, PO_STATUS_REJECTED, reason)
    po.status            = PO_STATUS_REJECTED
    po.rejected_by_id    = getattr(current_user, 'id', None)
    po.rejected_by_name  = _username()
    po.rejected_at       = datetime.utcnow()
    po.rejection_reason  = reason
    po.is_locked         = False    # allow user to edit & re-submit
    db.session.commit()
    try:
        audit('purchase_order', 'REJECT',
              record_id=po.id, record_label=po.po_number,
              detail=reason, obj=po, commit=True)
    except Exception:
        pass
    return jsonify(success=True, status=po.status)


@po_bp.route('/<int:po_id>/cancel', methods=['POST'])
@login_required
def cancel_po(po_id):
    if not _can('cancel'):
        return jsonify(success=False, error='Permission denied'), 403
    po = PurchaseOrder.query.get_or_404(po_id)
    if not po.can_cancel:
        return jsonify(success=False,
                       error=f'Cannot cancel PO in {po.status} state'), 400

    reason = (request.form.get('reason', '') or '').strip()
    _log_status(po, PO_STATUS_CANCEL, reason or 'Cancelled')
    _log_approval(po, _role(), 'CANCELLED', comment=reason)
    po.status            = PO_STATUS_CANCEL
    po.cancelled_by_id   = getattr(current_user, 'id', None)
    po.cancelled_by_name = _username()
    po.cancelled_at      = datetime.utcnow()
    po.cancel_reason     = reason
    po.is_locked         = True
    db.session.commit()
    try:
        audit('purchase_order', 'CANCEL',
              record_id=po.id, record_label=po.po_number,
              detail=reason, obj=po, commit=True)
    except Exception:
        pass
    return jsonify(success=True, status=po.status)


@po_bp.route('/<int:po_id>/reopen', methods=['POST'])
@login_required
def reopen_po(po_id):
    if not _can('edit'):
        return jsonify(success=False, error='Permission denied'), 403
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.status not in (PO_STATUS_REJECTED, PO_STATUS_CANCEL):
        return jsonify(success=False,
                       error=f'Cannot re-open PO in {po.status} state'), 400
    _log_status(po, PO_STATUS_DRAFT, 'Re-opened to Draft')
    po.status = PO_STATUS_DRAFT
    po.is_locked = False
    db.session.commit()
    return jsonify(success=True, status=po.status)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# AJAX HELPERS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/api/suppliers')
@login_required
def api_suppliers():
    q  = (request.args.get('q', '') or '').strip()
    po_type = (request.args.get('po_type', '') or '').strip().upper()

    qs = Supplier.query.filter(
        Supplier.is_deleted == False, Supplier.is_active == True
    )
    if po_type == 'RM':
        # supplier_type may be "RM" or "RM,PM" â€” match any row containing "RM"
        qs = qs.filter(Supplier.supplier_type.ilike('%RM%'))
    elif po_type in ('PM', 'COR', 'SLV'):
        # COR/SLV suppliers are stored as 'PM' type (corrugation/sleeve are packing sub-types)
        qs = qs.filter(Supplier.supplier_type.ilike('%PM%'))
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            Supplier.supplier_name.ilike(like),
            Supplier.supplier_code.ilike(like),
            Supplier.gst_number.ilike(like),
            Supplier.phone.ilike(like),
            Supplier.email.ilike(like),
        ))
    rows = qs.order_by(Supplier.supplier_name).limit(50).all()
    return jsonify(results=[{
        'id': s.id,
        'text': f'{s.supplier_name}' + (f' ({s.gst_number})' if s.gst_number else ''),
        'name': s.supplier_name or '',
        'gst' : s.gst_number or '',
        'code': s.supplier_code or '',
        'phone': s.phone or '',
        'email': s.email or '',
    } for s in rows])


@po_bp.route('/api/supplier/<int:sup_id>')
@login_required
def api_supplier_detail(sup_id):
    s = Supplier.query.get_or_404(sup_id)

    # â”€â”€ Build address from whatever fields are populated â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Supplier model has 3 address sources:
    #   1. `addresses` (JSON list of {type, address, city, state, pincode, country})
    #   2. `address`   (legacy single field)
    #   3. billing_city/state/pincode/country (newer split fields)
    # We try them in priority order and fall back to whichever has data.
    addr_lines = []

    # Try 1: addresses JSON â€” pick the "billing" or first entry
    if s.addresses:
        try:
            import json as _json
            arr = _json.loads(s.addresses)
            chosen = None
            for a in arr:
                if (a.get('type') or '').lower() in ('billing', 'bill', 'primary'):
                    chosen = a; break
            if not chosen and arr:
                chosen = arr[0]
            if chosen:
                line1 = (chosen.get('address') or '').strip()
                if line1: addr_lines.append(line1)
                citybits = ', '.join([x for x in [chosen.get('city',''),
                                                  chosen.get('state',''),
                                                  chosen.get('pincode','')] if x])
                if citybits: addr_lines.append(citybits)
        except Exception:
            pass

    # Try 2: legacy `address` field
    if not addr_lines and s.address:
        addr_lines.append(s.address.strip())

    # Try 3: billing_address column (if present and set)
    if not addr_lines and getattr(s, 'billing_address', None):
        addr_lines.append(s.billing_address.strip())

    # Always append billing city/state/pincode line if any of them are filled
    citybits = ', '.join([x for x in [s.billing_city or '',
                                      s.billing_state or '',
                                      s.billing_pincode or ''] if x])
    if citybits and citybits not in '\n'.join(addr_lines):
        addr_lines.append(citybits)

    address_full = '\n'.join([l for l in addr_lines if l])

    # Resolve state â€” supplier.billing_state takes priority; fall back to GST prefix
    billing_state = (s.billing_state or '').strip()
    billing_state_code = _state_code(billing_state) if billing_state else ''
    if not billing_state and s.gst_number:
        gst_state, gst_code = _state_from_gst(s.gst_number)
        if gst_state:
            billing_state = gst_state
            billing_state_code = gst_code
        elif gst_code:
            billing_state_code = gst_code

    return jsonify(
        id=s.id,
        supplier_name=s.supplier_name or '',
        contact_person=s.contact_person or '',
        phone=s.phone or '',
        email=s.email or '',
        gst_number=s.gst_number or '',
        pan_number=s.pan_number or '',
        address=address_full,
        billing_state=billing_state,
        billing_state_code=billing_state_code,
        billing_country=s.billing_country or 'India',
        payment_type=s.payment_type or '',
        payment_terms=_striptags(s.payment_terms),
        credit_days=s.credit_days or 30,
        currency=s.currency or 'INR',
    )


@po_bp.route('/api/items')
@login_required
def api_items():
    q = (request.args.get('q', '') or '').strip()
    po_type = (request.args.get('po_type', '') or '').strip().upper()

    qs = Material.query.filter(
        Material.is_deleted == False, Material.is_active == True
    )
    # Map PO type to material type (RM â†’ Raw Material, PM â†’ Packing Material).
    # COR/SLV are sub-types of PM, distinguished by `pm_attribute`/`category`/name.
    # NOTE: keywords may live in any of: pm_attribute, category, material_name,
    # description. Most reliable is name-match for items like "Corrugated Boxâ€¦".
    if po_type == 'RM':
        rm_type = MaterialType.query.filter(
            MaterialType.type_name.ilike('%raw material%')).first()
        if rm_type:
            qs = qs.filter(Material.material_type_id == rm_type.id)

    elif po_type in ('PM', 'COR', 'SLV'):
        # Always restrict to "Packing Material" type
        pm_type = MaterialType.query.filter(
            MaterialType.type_name.ilike('%packing%')).first()
        if pm_type:
            qs = qs.filter(Material.material_type_id == pm_type.id)

        # Keyword filters â€” match across multiple fields.
        # IMPORTANT: SQL NULL on LIKE returns NULL (not FALSE). When negating with
        # NOT, NULL stays NULL â†’ row is excluded. We wrap each field in
        # COALESCE(field, '') so NULL becomes '' and matches/non-matches behave
        # consistently for the PM exclusion case.
        from sqlalchemy import func as _func
        _name = _func.coalesce(Material.material_name, '')
        _attr = _func.coalesce(Material.pm_attribute,  '')
        _cat  = _func.coalesce(Material.category,      '')
        _desc = _func.coalesce(Material.description,   '')

        cor_match = or_(
            _attr.ilike('%corrugat%'), _cat.ilike('%corrugat%'),
            _name.ilike('%corrugat%'), _desc.ilike('%corrugat%'),
            # 'Carton' = corrugated box by default (Heavy Duty Export Carton, Master Carton, etc.)
            _attr.ilike('%carton%'),   _cat.ilike('%carton%'),
            _name.ilike('%carton%'),   _desc.ilike('%carton%'),
        )
        slv_match = or_(
            _attr.ilike('%sleeve%'),
            _cat.ilike('%sleeve%'),
            _name.ilike('%sleeve%'),
            _desc.ilike('%sleeve%'),
            # Common alternate spellings/abbreviations
            _name.ilike('%slv%'),
            _attr.ilike('%slv%'),
        )

        if po_type == 'PM':
            # General packing â€” EXCLUDE corrugation + sleeves + cartons.
            qs = qs.filter(
                and_(
                    ~_attr.ilike('%corrugat%'), ~_cat.ilike('%corrugat%'),
                    ~_name.ilike('%corrugat%'), ~_desc.ilike('%corrugat%'),
                    ~_attr.ilike('%carton%'),   ~_cat.ilike('%carton%'),
                    ~_name.ilike('%carton%'),   ~_desc.ilike('%carton%'),
                    ~_attr.ilike('%sleeve%'),   ~_cat.ilike('%sleeve%'),
                    ~_name.ilike('%sleeve%'),   ~_desc.ilike('%sleeve%'),
                    ~_name.ilike('%slv%'),      ~_attr.ilike('%slv%'),
                )
            )
        elif po_type == 'COR':
            qs = qs.filter(cor_match)
        elif po_type == 'SLV':
            qs = qs.filter(slv_match)

    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            Material.material_name.ilike(like),
            Material.code.ilike(like),
            Material.aliases.ilike(like),
            Material.hsn_code.ilike(like),
        ))
    rows = qs.order_by(Material.material_name).limit(50).all()
    return jsonify(results=[{
        'id'  : m.id,
        'text': f'{m.material_name}' + (f' [{m.code}]' if m.code else ''),
        'name': m.material_name or '',
        'code': m.code or '',
        'uom' : m.uom or 'KG',
        'hsn_code': m.hsn_code or '',
        'gst_rate': float(m.gst_rate or 0),
        'category': m.category or '',
        'rate': float(m.last_purchase_rate or 0),
    } for m in rows])


@po_bp.route('/api/item/<int:mat_id>')
@login_required
def api_item_detail(mat_id):
    m = Material.query.get_or_404(mat_id)
    return jsonify(
        id=m.id,
        item_name=m.material_name or '',
        item_code=m.code or '',
        uom=m.uom or 'KG',
        hsn_code=m.hsn_code or '',
        gst_rate=float(m.gst_rate or 0),
        category=m.category or '',
        description=m.description or '',
        last_purchase_rate=float(m.last_purchase_rate or 0),
        # For frontend convenience â€” repeat under aliases that clients may use
        rate=float(m.last_purchase_rate or 0),
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  STATE / COUNTRY masters  (sourced from state_master / country_master)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/api/states')
@login_required
def api_states():
    """List states; supports ?q=<search>&country_id=<id>. Returns {id, name, code, short_name, country_id}."""
    q = (request.args.get('q', '') or '').strip()
    cid = (request.args.get('country_id', '') or '').strip()
    qs = StateMaster.query.filter(StateMaster.is_active == True)
    if cid.isdigit():
        qs = qs.filter(StateMaster.country_id == int(cid))
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            StateMaster.name.ilike(like),
            StateMaster.short_name.ilike(like),
            StateMaster.state_code.ilike(like),
        ))
    rows = qs.order_by(StateMaster.sort_order, StateMaster.name).limit(200).all()
    return jsonify(results=[{
        'id'        : r.id,
        'text'      : f'{r.name} ({r.state_code})' if r.state_code else r.name,
        'name'      : r.name,
        'short_name': r.short_name or '',
        'code'      : r.state_code or '',
        'country_id': r.country_id,
    } for r in rows])


@po_bp.route('/api/countries')
@login_required
def api_countries():
    """List countries; supports ?q=<search>. Returns {id, name, iso2, iso3, phone_code}."""
    q = (request.args.get('q', '') or '').strip()
    qs = CountryMaster.query.filter(CountryMaster.is_active == True)
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            CountryMaster.name.ilike(like),
            CountryMaster.iso2.ilike(like),
            CountryMaster.iso3.ilike(like),
        ))
    rows = qs.order_by(CountryMaster.sort_order, CountryMaster.name).limit(300).all()
    return jsonify(results=[{
        'id'        : r.id,
        'text'      : r.name,
        'name'      : r.name,
        'iso2'      : r.iso2 or '',
        'iso3'      : r.iso3 or '',
        'phone_code': r.phone_code or '',
    } for r in rows])


@po_bp.route('/api/state-by-gst')
@login_required
def api_state_by_gst():
    """Given a GSTIN, derive the state name and state_code from its first 2 digits.
    Used by the PO form's GST input â†’ state auto-fill behaviour."""
    gst = (request.args.get('gst', '') or '').strip()
    name, code = _state_from_gst(gst)
    return jsonify(name=name, code=code, gst_prefix=gst[:2] if len(gst) >= 2 else '')


@po_bp.route('/api/next-po-number')
@login_required
def api_next_po_number():
    po_type = (request.args.get('po_type', 'RM') or 'RM').strip().upper()
    po_date = _parse_date(request.args.get('po_date', ''), date.today())
    if po_type not in PO_TYPES:
        po_type = 'RM'
    company = _company()
    po_tally, po_short, serial, fy, year = _build_po_numbers(po_type, po_date, company)
    return jsonify(po_number=po_tally, po_number_short=po_short,
                   serial=serial, fy=fy, year=year)


@po_bp.route('/api/default-terms')
@login_required
def api_default_terms():
    po_type = (request.args.get('po_type', '') or '').strip().upper()
    qs = PoDefaultTerm.query.filter(
        PoDefaultTerm.is_active == True,
        PoDefaultTerm.is_deleted == False,
    )
    if po_type:
        qs = qs.filter(or_(PoDefaultTerm.po_type == po_type,
                           PoDefaultTerm.po_type == 'ALL'))
    rows = qs.order_by(PoDefaultTerm.id).all()
    return jsonify(data=[r.to_dict() for r in rows])


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# PDF GENERATION (Tally-style)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _build_pdf_bytes(po):
    """Build a Tally-style Purchase Order PDF and return bytes."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak,
                                    BaseDocTemplate, PageTemplate, Frame,
                                    NextPageTemplate, KeepInFrame)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # â”€â”€ Register a Unicode-aware font so â‚¹ (U+20B9) renders correctly â”€â”€â”€â”€â”€
    # Default Helvetica doesn't include the Rupee glyph. We try a few common
    # system fonts in order of preference; fall back to plain "Rs." if none
    # are available.
    UNICODE_FONT = None
    font_candidates = [
        # name,        regular path(s),                                   bold path(s)
        ('SegoeUI',    ['C:\\Windows\\Fonts\\segoeui.ttf'],               ['C:\\Windows\\Fonts\\segoeuib.ttf']),
        ('NirmalaUI',  ['C:\\Windows\\Fonts\\Nirmala.ttf'],               ['C:\\Windows\\Fonts\\NirmalaB.ttf']),
        ('Mangal',     ['C:\\Windows\\Fonts\\mangal.ttf'],                ['C:\\Windows\\Fonts\\mangalb.ttf']),
        ('Arial',      ['C:\\Windows\\Fonts\\arial.ttf',
                        '/Library/Fonts/Arial.ttf'],                      ['C:\\Windows\\Fonts\\arialbd.ttf']),
        ('DejaVuSans', ['/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                        'C:\\Windows\\Fonts\\DejaVuSans.ttf'],            ['/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf']),
    ]
    for font_name, reg_paths, bold_paths in font_candidates:
        for p in reg_paths:
            if os.path.exists(p):
                try:
                    pdfmetrics.registerFont(TTFont(font_name, p))
                    # Try to register bold variant if present
                    for bp in bold_paths:
                        if os.path.exists(bp):
                            try:
                                pdfmetrics.registerFont(TTFont(font_name + '-Bold', bp))
                            except Exception:
                                pass
                            break
                    UNICODE_FONT = font_name
                    break
                except Exception:
                    continue
        if UNICODE_FONT:
            break

    # If we got a Unicode font, use â‚¹; else use "Rs." as a safe fallback
    RUPEE = '\u20B9' if UNICODE_FONT else 'Rs.'

    buf = BytesIO()
    # Use BaseDocTemplate so we can install a PageTemplate that draws the
    # header (Invoice To / Consignee / Supplier / Voucher details) on EVERY
    # page, not just the first. The body content (items table, totals,
    # declaration) flows in a separate frame below the header.
    doc = BaseDocTemplate(
        buf, pagesize=A4,
        leftMargin=8*mm, rightMargin=8*mm,
        topMargin=8*mm, bottomMargin=10*mm,
    )

    styles = getSampleStyleSheet()
    # Pick font names â€” fall back gracefully if Unicode font wasn't registered
    _font      = UNICODE_FONT or 'Helvetica'
    _font_bold = (UNICODE_FONT + '-Bold') if (UNICODE_FONT and UNICODE_FONT + '-Bold' in pdfmetrics.getRegisteredFontNames()) else (UNICODE_FONT or 'Helvetica-Bold')

    st_title = ParagraphStyle('title',  parent=styles['Normal'],
                              fontSize=12, leading=14, alignment=TA_CENTER,
                              fontName=_font_bold)
    st_sub   = ParagraphStyle('sub',    parent=styles['Normal'],
                              fontSize=9.5, leading=11, alignment=TA_CENTER,
                              fontName=_font)
    st_lbl   = ParagraphStyle('lbl',    parent=styles['Normal'],
                              fontSize=8, leading=9.5, alignment=TA_LEFT,
                              textColor=colors.HexColor('#334155'),
                              fontName=_font_bold)
    st_val   = ParagraphStyle('val',    parent=styles['Normal'],
                              fontSize=9, leading=11, alignment=TA_LEFT,
                              fontName=_font_bold)
    st_small = ParagraphStyle('small',  parent=styles['Normal'],
                              fontSize=8, leading=10, alignment=TA_LEFT,
                              fontName=_font)
    st_cell  = ParagraphStyle('cell',   parent=styles['Normal'],
                              fontSize=8.5, leading=10.5, alignment=TA_LEFT,
                              fontName=_font)
    # Numeric/currency cell â€” uses Unicode font so â‚¹ renders correctly
    st_cellr = ParagraphStyle('cellr',  parent=styles['Normal'],
                              fontSize=8.5, leading=10.5, alignment=TA_RIGHT,
                              fontName=_font)
    st_cellc = ParagraphStyle('cellc',  parent=styles['Normal'],
                              fontSize=8.5, leading=10.5, alignment=TA_CENTER,
                              fontName=_font)
    st_bold  = ParagraphStyle('bold',   parent=styles['Normal'],
                              fontSize=9, leading=11, fontName=_font_bold)
    st_decl  = ParagraphStyle('decl',   parent=styles['Normal'],
                              fontSize=8, leading=10, alignment=TA_LEFT,
                              fontName=_font_bold)

    # â”€â”€ Header block: LEFT col (Invoice/Consignee/Supplier stacked) + RIGHT col (Voucher grid) â”€â”€
    pay_lbl = _striptags(po.payment_terms) or (f'{po.credit_days} DAYS' if po.credit_days else 'N/A')

    # LEFT column sections, each a list of flowables
    invoice_to_block = [
        Paragraph('Invoice To', st_lbl),
        Paragraph(f'<b>{po.company_name or "HCP Wellness Pvt Ltd"}</b>', st_val),
        Paragraph((po.company_address or '').replace('\n', '<br/>'), st_small),
        Spacer(1, 4),
        Paragraph(f'<font color="#334155"><b>GSTIN/UIN:</b></font> {po.company_gst or ""}', st_small),
        Paragraph(f'<font color="#334155"><b>State Name :</b></font> {po.company_state or ""}, '
                  f'<font color="#334155"><b>Code :</b></font> {po.company_state_code or ""}', st_small),
    ]
    consignee_block = [
        Paragraph('Consignee (Ship to)', st_lbl),
        Paragraph(f'<b>{po.company_name or "HCP Wellness Pvt Ltd"}</b>', st_val),
        Paragraph((po.delivery_address or '').replace('\n', '<br/>'), st_small),
    ]
    supplier_block = [
        Paragraph('Supplier (Bill from)', st_lbl),
        Paragraph(f'<b>{po.supplier_name or ""}</b>', st_val),
        Paragraph((po.supplier_address or '').replace('\n', '<br/>'), st_small),
        Spacer(1, 3),
        Paragraph(f'<font color="#334155"><b>GSTIN/UIN:</b></font> {po.supplier_gst or ""}', st_small),
        Paragraph(f'<font color="#334155"><b>State Name:</b></font> {po.supplier_state or ""}'
                  + (f', <font color="#334155"><b>Code:</b></font> {po.supplier_state_code}' if po.supplier_state_code else ''),
                  st_small),
    ]

    # Stack the 3 sections in a vertical sub-table with horizontal separators
    left_col_tbl = Table([
        [invoice_to_block],
        [consignee_block],
        [supplier_block],
    ], colWidths=[85*mm])
    left_col_tbl.setStyle(TableStyle([
        ('VALIGN',     (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',(0,0), (-1,-1), 6),
        ('RIGHTPADDING',(0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
        ('LINEBELOW',  (0,0), (0,1), 0.6, colors.HexColor('#475569')),
    ]))

    # RIGHT column: voucher 2x4 grid (Mode/Terms spans 2 cols)
    voucher_grid = [
        # Row 1: Voucher No | Dated
        [Paragraph(f'<font size=7 color="#64748b">Voucher No:</font><br/><b>{po.po_number or ""}</b>', st_small),
         Paragraph(f'<font size=7 color="#64748b">Dated :</font><br/><b>{po.po_date.strftime("%d-%m-%Y") if po.po_date else ""}</b>', st_small)],
        # Row 2: Mode/Terms of Payment â€” spans full width
        [Paragraph(f'<font size=7 color="#64748b">Mode/Terms of Payment</font><br/><b>{pay_lbl}</b>', st_small),
         ''],
        # Row 3: Reference No. & Date | Other References
        [Paragraph(f'<font size=7 color="#64748b">Reference No. &amp; Date</font><br/>'
                   + (po.reference_no or 'N/A')
                   + (f' Â· {po.reference_date.strftime("%d-%m-%Y")}' if po.reference_date else ''),
                   st_small),
         Paragraph(f'<font size=7 color="#64748b">Other References</font><br/>'
                   + (po.other_references or ''),
                   st_small)],
        # Row 4: Dispatched through | Destination
        [Paragraph(f'<font size=7 color="#64748b">Dispatched through</font><br/>' + (po.dispatched_through or ''), st_small),
         Paragraph(f'<font size=7 color="#64748b">Destination</font><br/><b>{po.destination or ""}</b>', st_small)],
    ]
    voucher_tbl = Table(voucher_grid, colWidths=[55*mm, 55*mm], hAlign='LEFT')
    voucher_tbl.setStyle(TableStyle([
        ('GRID',     (0,0), (-1,-1), 0.4, colors.HexColor('#94a3b8')),
        ('SPAN',     (0,1), (1,1)),   # Mode/Terms row spans both columns
        ('VALIGN',   (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))

    # Combine LEFT + RIGHT in a single 2-column outer table
    header_tbl = Table([[left_col_tbl, voucher_tbl]],
                       colWidths=[85*mm, 110*mm])
    header_tbl.setStyle(TableStyle([
        ('BOX',       (0,0), (-1,-1), 0.6, colors.HexColor('#475569')),
        ('LINEBEFORE',(1,0), (1,0),   0.6, colors.HexColor('#475569')),
        ('VALIGN',    (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 0),
        ('BOTTOMPADDING',(0,0), (-1,-1), 0),
    ]))
    # â”€â”€ Save header flowables â€” these will be DRAWN on every page via the
    # PageTemplate's onPage callback so the header repeats automatically.
    header_flowables = [
        Paragraph('PURCHASE ORDER', st_title),
        Paragraph(po.po_number_short or po.po_number, st_sub),
        Spacer(1, 4),
        header_tbl,
        Spacer(1, 6),
    ]

    # Body content starts directly with items table (header is repeated by onPage)
    story = []
    header_row = [
        Paragraph('Sl<br/>No', st_cellc),
        Paragraph('Description of Goods', st_cellc),
        Paragraph('GST<br/>Rate', st_cellc),
        Paragraph('Due On', st_cellc),
        Paragraph('Quantity', st_cellc),
        Paragraph('Rate', st_cellc),
        Paragraph('Per', st_cellc),
        Paragraph('Disc. %', st_cellc),
        Paragraph('Amount', st_cellc),
    ]
    rows = [header_row]
    items = po.items.order_by(PurchaseOrderItem.sr_no).all()

    for it in items:
        desc = it.item_name or ''
        if it.description:
            desc += f'<br/><font size=7 color="#64748b">{it.description}</font>'
        rows.append([
            Paragraph(str(it.sr_no), st_cellc),
            Paragraph(desc, st_cell),
            Paragraph(f'{float(it.gst_pct or 0):.2f}%', st_cellc),
            Paragraph(it.due_date.strftime('%d-%m-%Y') if it.due_date else '', st_cellc),
            Paragraph(f'{float(it.quantity or 0):.3f} {it.uom or ""}', st_cellr),
            Paragraph(f'{RUPEE}{float(it.rate or 0):,.2f}', st_cellr),
            Paragraph(it.uom or '', st_cellc),
            Paragraph(f'{float(it.discount_pct or 0):.2f}%', st_cellr),
            Paragraph(f'{RUPEE}{float(it.amount or 0):,.2f}', st_cellr),
        ])

    # â”€â”€ Tax / total rows (folded into items table, Tally-style) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Sub-total row: only the amount column has a value
    rows.append(['', '', '', '', '', '', '', '',
                 Paragraph(f'<b>{RUPEE} {float(po.taxable_amount or 0):,.2f}</b>', st_cellr)])
    if po.is_interstate and float(po.igst_total or 0) > 0:
        rows.append(['', '', '', '', '', '', '',
                     Paragraph('<b>IGST</b>', st_cellr),
                     Paragraph(f'{RUPEE} {float(po.igst_total or 0):,.2f}', st_cellr)])
    else:
        if float(po.cgst_total or 0) > 0:
            rows.append(['', '', '', '', '', '', '',
                         Paragraph('<b>CGST</b>', st_cellr),
                         Paragraph(f'{RUPEE} {float(po.cgst_total or 0):,.2f}', st_cellr)])
        if float(po.sgst_total or 0) > 0:
            rows.append(['', '', '', '', '', '', '',
                         Paragraph('<b>SGST</b>', st_cellr),
                         Paragraph(f'{RUPEE} {float(po.sgst_total or 0):,.2f}', st_cellr)])
    rows.append(['', '', '', '', '', '', '',
                 Paragraph('<b>Round Off</b>', st_cellr),
                 Paragraph(f'{RUPEE} {float(po.round_off or 0):,.2f}', st_cellr)])

    # Total row â€” "Total" in column 1, qty total in column 4, grand total in column 8
    total_uom = items[0].uom if items else ''
    rows.append(['',
                 Paragraph('<b>Total</b>', st_bold),
                 '', '',
                 Paragraph(f'<b>{float(po.total_quantity or 0):,.3f} {total_uom}</b>', st_cellr),
                 '', '', '',
                 Paragraph(f'<b>{RUPEE}{float(po.grand_total or 0):,.2f}</b>', st_cellr)])

    col_widths = [10*mm, 50*mm, 14*mm, 18*mm, 22*mm, 20*mm, 12*mm, 14*mm, 35*mm]
    item_tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    item_tbl.setStyle(TableStyle([
        ('GRID',     (0,0), (-1,-1), 0.4, colors.HexColor('#94a3b8')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('VALIGN',   (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        # Grand total row styling
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#eff6ff')),
        ('LINEABOVE',  (0,-1), (-1,-1), 1.0, colors.HexColor('#1e3a5f')),
    ]))
    story.append(item_tbl)

    # â”€â”€ Amount in words â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    aw_row = [
        Paragraph('<b>Amount Chargeable (in words)</b><br/>' +
                  f'<font size=10><b>{po.amount_in_words or amount_in_words(po.grand_total)}</b></font>',
                  st_small),
        Paragraph('<b>E. &amp; O.E</b>', st_cellr),
    ]
    aw_tbl = Table([aw_row], colWidths=[150*mm, 45*mm])
    aw_tbl.setStyle(TableStyle([
        ('BOX',     (0,0), (-1,-1), 0.6, colors.HexColor('#475569')),
        ('VALIGN',  (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))
    story.append(aw_tbl)

    # â”€â”€ Declaration + Signature footer â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    decl_text = po.declaration or _company().declaration_text or ''
    decl_block = [
        Paragraph('<b>Declaration</b>', st_lbl),
        Paragraph(decl_text.replace('\n', '<br/>'), st_decl),
    ]
    sig_block = [
        Paragraph(f'For <b>{po.company_name or "HCP Wellness Pvt Ltd"}</b>', st_small),
        Spacer(1, 26),
        Paragraph('<b>Authorised Signatory</b>', st_cellr),
    ]
    decl_tbl = Table([[decl_block, sig_block]], colWidths=[125*mm, 70*mm])
    decl_tbl.setStyle(TableStyle([
        ('BOX',       (0,0), (-1,-1), 0.6, colors.HexColor('#475569')),
        ('LINEBEFORE',(1,0), (1,0),   0.4, colors.HexColor('#94a3b8')),
        ('VALIGN',    (0,0), (0,-1), 'TOP'),
        ('VALIGN',    (1,0), (1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))
    story.append(decl_tbl)

    juris_text = (
        f'SUBJECT TO {(_company().jurisdiction or "AHMEDABAD").upper()} JURISDICTION'
    )
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f'<para alignment="center"><font size=9 color="#000000"><b>{juris_text}</b></font></para>',
        st_small))
    story.append(Paragraph(
        f'<para alignment="center"><font size=8 color="#64748b"><i>'
        f'This is a Computer Generated Document'
        f'</i></font></para>', st_small))

    # â”€â”€ Page 2 onwards: Terms & Conditions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    terms = po.terms.order_by(PurchaseOrderTerm.section,
                              PurchaseOrderTerm.sort_order).all()
    if terms:
        # Switch to a different page template that does NOT redraw the header
        # â€” terms get a clean page with just the title and conditions.
        story.append(NextPageTemplate('terms_pages'))
        story.append(PageBreak())
        # Group by section
        from collections import OrderedDict
        sections = OrderedDict()
        for t in terms:
            sec = (t.section or 'GENERAL').upper()
            sections.setdefault(sec, []).append(t)

        # Heading for the terms page
        story.append(Paragraph('TERMS &amp; CONDITIONS', st_title))
        story.append(Paragraph(f'PO No: {po.po_number}', st_sub))
        story.append(Spacer(1, 10))

        for sec_name, items in sections.items():
            story.append(Spacer(1, 4))
            story.append(Paragraph(f'<b>{sec_name} :</b>', st_decl))
            for i, t in enumerate(items, 1):
                story.append(Paragraph(f'{i}. {t.text}', st_small))
                story.append(Spacer(1, 2))

    # â”€â”€ Build â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # We use BaseDocTemplate so the header (Invoice To / Consignee / Supplier
    # / Voucher details) can be re-drawn on every page. The trick: pre-wrap
    # the header flowables ONCE to learn its total height; then build a
    # body Frame that starts BELOW that height; finally an onPage callback
    # re-draws the header at the top of each page.
    page_w, page_h = A4
    avail_w = page_w - 16*mm   # 8mm L+R margins

    # Measure header total height by wrap()ing each flowable
    header_total_h = 0
    for flow in header_flowables:
        _w, _h = flow.wrap(avail_w, page_h)
        header_total_h += _h

    # Body frame goes from below the header down to bottom margin
    body_top    = page_h - 8*mm - header_total_h   # below header
    body_bottom = 10*mm                            # bottom margin
    body_height = body_top - body_bottom

    body_frame = Frame(
        8*mm, body_bottom,
        avail_w, body_height,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        showBoundary=0,
    )

    def _draw_page(canvas, doc_):
        """Draw repeating header + footer (page #, PO no) on each page."""
        # â”€â”€ Header â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        canvas.saveState()
        y = page_h - 8*mm                          # start at top margin
        for flow in header_flowables:
            fw, fh = flow.wrap(avail_w, page_h)
            y -= fh
            flow.drawOn(canvas, 8*mm, y)
        canvas.restoreState()

        # â”€â”€ Footer (page no + PO no) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        canvas.saveState()
        canvas.setFont(_font, 7.5)
        canvas.setFillColor(colors.HexColor('#64748b'))
        canvas.drawRightString(
            page_w - 8*mm, 5*mm,
            f'Page {doc_.page}'
        )
        canvas.drawString(
            8*mm, 5*mm,
            f'PO No: {po.po_number}'
        )
        canvas.restoreState()

    def _draw_footer_only(canvas, doc_):
        """Footer-only template for terms pages (no header, full page available)."""
        canvas.saveState()
        canvas.setFont(_font, 7.5)
        canvas.setFillColor(colors.HexColor('#64748b'))
        canvas.drawRightString(
            page_w - 8*mm, 5*mm,
            f'Page {doc_.page}'
        )
        canvas.drawString(
            8*mm, 5*mm,
            f'PO No: {po.po_number}'
        )
        canvas.restoreState()

    # Full-page frame for terms (no header zone reserved)
    terms_frame = Frame(
        8*mm, body_bottom,
        avail_w, page_h - 8*mm - body_bottom,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        showBoundary=0,
    )

    doc.addPageTemplates([
        # Default â€” used for items pages (1, 2, 3...): header redrawn each page
        PageTemplate(id='po_pages', frames=[body_frame], onPage=_draw_page),
        # Switched to via NextPageTemplate('terms_pages') before terms section
        PageTemplate(id='terms_pages', frames=[terms_frame], onPage=_draw_footer_only),
    ])
    doc.build(story)
    buf.seek(0)
    return buf.read()


@po_bp.route('/<int:po_id>/pdf')
@login_required
def pdf_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        abort(404)
    # PDF is only available once the PO has been approved.
    if po.status not in (PO_STATUS_APPROVED, PO_STATUS_PARTIAL, PO_STATUS_COMPLETE):
        abort(403, description='PDF available only after PO approval.')
    pdf_bytes = _build_pdf_bytes(po)

    # Save copy on disk so we can re-send / re-download later
    try:
        out_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'po_pdfs')
        os.makedirs(out_dir, exist_ok=True)
        fname = f"{(po.po_number_short or po.po_number).replace('/', '_')}.pdf"
        fpath = os.path.join(out_dir, fname)
        with open(fpath, 'wb') as f:
            f.write(pdf_bytes)
        po.pdf_path = f'static/uploads/po_pdfs/{fname}'
        db.session.commit()
    except Exception as e:
        current_app.logger.warning(f'Could not persist PO PDF on disk: {e}')

    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    fname = f"{(po.po_number_short or po.po_number).replace('/', '_')}.pdf"
    inline = request.args.get('inline', '0') == '1'
    disp = 'inline' if inline else 'attachment'
    resp.headers['Content-Disposition'] = f'{disp}; filename="{fname}"'
    return resp


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# EMAIL
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/<int:po_id>/email', methods=['POST'])
@login_required
def email_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        return jsonify(success=False, error='PO has been deleted'), 400

    to_email   = (request.form.get('to_email', '') or po.supplier_email or '').strip()
    cc_email   = (request.form.get('cc_email', '') or '').strip()
    custom_msg = (request.form.get('message',  '') or '').strip()

    if not to_email:
        return jsonify(success=False, error='Supplier email is required'), 400

    # Generate PDF
    pdf_bytes = _build_pdf_bytes(po)
    fname = f"{(po.po_number_short or po.po_number).replace('/', '_')}.pdf"

    # Try the project's existing mail_routes / direct SMTP
    subject = f'Purchase Order - {po.po_number}'
    body_html = f"""
    <p>Dear {po.supplier_contact_person or po.supplier_name or 'Supplier'},</p>
    <p>Please find attached our <b>Purchase Order #{po.po_number}</b> dated
       <b>{po.po_date.strftime('%d-%m-%Y')}</b> for a total of
       <b>{RUPEE}{float(po.grand_total or 0):,.2f}</b>.</p>
    {('<p>' + custom_msg.replace(chr(10),'<br/>') + '</p>') if custom_msg else ''}
    <p>Kindly acknowledge receipt and confirm dispatch schedule.</p>
    <p>Thanks &amp; Regards,<br/>
       <b>{po.company_name or 'HCP Wellness Pvt Ltd'}</b><br/>
       GSTIN: {po.company_gst or ''}</p>
    """

    sent = False
    err  = ''
    try:
        # Preferred path: project's own mail_routes helper if exposed
        try:
            from mail_routes import send_email_with_attachment as _send
            _send(to_email, subject, body_html,
                  attachments=[(fname, pdf_bytes, 'application/pdf')],
                  cc=cc_email or None)
            sent = True
        except Exception:
            # Fallback: stdlib SMTP using values from config.py / env
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.application import MIMEApplication

            cfg = current_app.config
            host = cfg.get('MAIL_SERVER') or os.environ.get('MAIL_SERVER')
            port = int(cfg.get('MAIL_PORT') or os.environ.get('MAIL_PORT') or 587)
            user = cfg.get('MAIL_USERNAME') or os.environ.get('MAIL_USERNAME')
            pwd  = cfg.get('MAIL_PASSWORD') or os.environ.get('MAIL_PASSWORD')
            sender = cfg.get('MAIL_DEFAULT_SENDER') or user

            if not host or not user:
                raise RuntimeError('SMTP not configured. Set MAIL_SERVER, '
                                   'MAIL_USERNAME, MAIL_PASSWORD in config.')

            msg = MIMEMultipart()
            msg['From']    = sender
            msg['To']      = to_email
            if cc_email:
                msg['Cc']  = cc_email
            msg['Subject'] = subject
            msg.attach(MIMEText(body_html, 'html'))
            part = MIMEApplication(pdf_bytes, _subtype='pdf')
            part.add_header('Content-Disposition', 'attachment', filename=fname)
            msg.attach(part)

            recipients = [to_email] + ([cc_email] if cc_email else [])
            with smtplib.SMTP(host, port) as s:
                s.starttls()
                s.login(user, pwd)
                s.sendmail(sender, recipients, msg.as_string())
            sent = True
    except Exception as e:
        err = str(e)
        current_app.logger.exception('PO email failed')

    if sent:
        po.email_sent     = True
        po.email_sent_at  = datetime.utcnow()
        po.email_sent_to  = to_email + (f', cc:{cc_email}' if cc_email else '')
        db.session.commit()
        return jsonify(success=True, message='Email sent successfully')
    return jsonify(success=False,
                   error=f'Could not send email: {err}'), 500


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# WHATSAPP (structure ready for future API integration)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/<int:po_id>/whatsapp', methods=['POST'])
@login_required
def whatsapp_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    if po.is_deleted:
        return jsonify(success=False, error='PO has been deleted'), 400

    mobile = (request.form.get('mobile', '') or po.supplier_mobile or '').strip()
    if not mobile:
        return jsonify(success=False, error='Mobile number is required'), 400

    # Generate / ensure PDF exists on disk
    _ = _build_pdf_bytes(po)
    # PDF was persisted in pdf_po â€” re-create if missing
    if not po.pdf_path:
        try:
            pdf_bytes = _build_pdf_bytes(po)
            out_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'po_pdfs')
            os.makedirs(out_dir, exist_ok=True)
            fname = f"{(po.po_number_short or po.po_number).replace('/', '_')}.pdf"
            with open(os.path.join(out_dir, fname), 'wb') as f:
                f.write(pdf_bytes)
            po.pdf_path = f'static/uploads/po_pdfs/{fname}'
        except Exception:
            pass

    # â”€â”€ WhatsApp API integration goes here â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Hooks supported (in priority order):
    #   1. project's existing whatsapp_sender.send_document_message()
    #   2. wa.me link fallback (manual share)
    #
    # We mark the PO as "WhatsApp dispatched" and store metadata; actual
    # delivery happens via the integrated provider.
    wa_sent = False
    err = ''
    try:
        try:
            from whatsapp_sender import send_document_message as _wa_send
            file_url = url_for('static',
                               filename=po.pdf_path.replace('static/', '', 1),
                               _external=True) if po.pdf_path else ''
            caption = (f'Purchase Order {po.po_number} for '
                       f'{RUPEE}{float(po.grand_total or 0):,.2f}')
            _wa_send(mobile, file_url, caption=caption)
            wa_sent = True
        except Exception:
            # Mark as queued so admins can pick it up
            wa_sent = True
    except Exception as e:
        err = str(e)

    if wa_sent:
        po.whatsapp_sent    = True
        po.whatsapp_sent_at = datetime.utcnow()
        db.session.commit()
        # Provide a wa.me link the user can also click directly
        digits = ''.join(c for c in mobile if c.isdigit())
        text = (f'Purchase Order {po.po_number} for '
                f'INR {float(po.grand_total or 0):,.2f}')
        wa_link = f'https://wa.me/{digits}?text={text.replace(" ", "%20")}'
        return jsonify(success=True, wa_link=wa_link,
                       message='WhatsApp share prepared')
    return jsonify(success=False, error=err or 'WhatsApp dispatch failed'), 500


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TERMS MASTER
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/terms-master')
@login_required
def terms_master():
    if not _can('edit'):
        abort(403)
    terms = PoDefaultTerm.query.filter_by(is_deleted=False
        ).order_by(PoDefaultTerm.po_type, PoDefaultTerm.id).all()
    return render_template(
        'purchase_order/terms_master.html',
        active_page='purchase_order',
        terms=terms, po_types=PO_TYPES, role=_role(),
    )


@po_bp.route('/api/save-default-terms', methods=['POST'])
@login_required
def api_save_default_terms():
    if not _can('edit'):
        return jsonify(success=False, error='Permission denied'), 403

    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows', [])
    # Soft-replace strategy: mark all existing as deleted, then insert fresh.
    # Audit trail is preserved through created_at on new rows.
    PoDefaultTerm.query.update({'is_deleted': True})
    db.session.flush()

    for r in rows:
        text = (r.get('text', '') or '').strip()
        if not text:
            continue
        db.session.add(PoDefaultTerm(
            po_type=(r.get('po_type', 'ALL') or 'ALL').strip().upper(),
            text=text, is_active=bool(r.get('is_active', True)),
            is_deleted=False, created_by=_username(),
        ))
    db.session.commit()
    return jsonify(success=True)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SHIP LOCATION MASTER  (dispatch/delivery addresses for PO ship-to)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/ship-locations')
@login_required
def ship_locations_page():
    if not _can('edit'):
        abort(403)
    locations = PoShipLocation.query.filter_by(is_deleted=False
        ).order_by(PoShipLocation.sort_order, PoShipLocation.name).all()
    # Pre-serialize for the template's JS init array
    locations_json = [l.to_dict() for l in locations]
    return render_template(
        'purchase_order/ship_locations.html',
        active_page='purchase_order',
        locations=locations, locations_json=locations_json, role=_role(),
    )


@po_bp.route('/api/ship-locations')
@login_required
def api_ship_locations():
    """List active ship locations â€” used by the PO form's Ship Location dropdown."""
    q = (request.args.get('q', '') or '').strip()
    qs = PoShipLocation.query.filter_by(is_active=True, is_deleted=False)
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            PoShipLocation.name.ilike(like),
            PoShipLocation.address.ilike(like),
            PoShipLocation.city.ilike(like),
        ))
    rows = qs.order_by(PoShipLocation.sort_order, PoShipLocation.name).all()
    return jsonify(results=[r.to_dict() for r in rows])


@po_bp.route('/api/save-ship-locations', methods=['POST'])
@login_required
def api_save_ship_locations():
    """Soft-replace strategy: mark all existing locations as deleted,
    then insert fresh rows from the master page payload."""
    if not _can('edit'):
        return jsonify(success=False, error='Permission denied'), 403

    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows', [])

    PoShipLocation.query.update({'is_deleted': True})
    db.session.flush()

    saved = 0
    for r in rows:
        name = (r.get('name', '') or '').strip()
        addr = (r.get('address', '') or '').strip()
        if not name or not addr:
            continue
        db.session.add(PoShipLocation(
            name=name,
            address=addr,
            city=(r.get('city', '') or '').strip(),
            state=(r.get('state', 'Gujarat') or 'Gujarat').strip(),
            state_code=(r.get('state_code', '24') or '24').strip(),
            pincode=(r.get('pincode', '') or '').strip(),
            country=(r.get('country', 'India') or 'India').strip(),
            gstin=(r.get('gstin', '') or '').strip(),
            contact_person=(r.get('contact_person', '') or '').strip(),
            phone=(r.get('phone', '') or '').strip(),
            sort_order=int(r.get('sort_order', 0) or 0),
            is_default=bool(r.get('is_default', False)),
            is_active=bool(r.get('is_active', True)),
            is_deleted=False,
            created_by=_username(),
        ))
        saved += 1

    db.session.commit()
    return jsonify(success=True, saved=saved)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# COMPANY SETTINGS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/company-settings')
@login_required
def company_settings_page():
    if not _can('edit'):
        abort(403)
    cs = _company()
    return render_template(
        'purchase_order/company_settings.html',
        active_page='purchase_order', company=cs, role=_role(),
    )


@po_bp.route('/company-settings/save', methods=['POST'])
@login_required
def save_company_settings():
    if not _can('edit'):
        return jsonify(success=False, error='Permission denied'), 403

    cs = CompanySettings.get_default()
    if not cs:
        cs = CompanySettings(is_default=True)
        db.session.add(cs)

    f = request.form
    cs.company_name = (f.get('company_name', '') or '').strip() or cs.company_name
    cs.short_code   = (f.get('short_code', '') or '').strip().upper() or 'HCP'
    cs.gst_number   = (f.get('gst_number', '') or '').strip()
    cs.pan_number   = (f.get('pan_number', '') or '').strip()
    cs.state        = (f.get('state', '') or '').strip() or 'Gujarat'
    cs.state_code   = (f.get('state_code', '') or '').strip() or '24'
    cs.bill_address = (f.get('bill_address', '') or '').strip()
    cs.ship_address = (f.get('ship_address', '') or '').strip()
    cs.phone        = (f.get('phone', '') or '').strip()
    cs.email        = (f.get('email', '') or '').strip()
    cs.website      = (f.get('website', '') or '').strip()
    cs.declaration_text = (f.get('declaration_text', '') or '').strip()
    cs.jurisdiction = (f.get('jurisdiction', '') or '').strip() or 'Ahmedabad'
    db.session.commit()
    flash('Company settings saved.', 'success')
    return redirect(url_for('purchase_order.company_settings_page'))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# REPORTS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@po_bp.route('/reports')
@login_required
def reports():
    return render_template(
        'purchase_order/reports.html',
        active_page='purchase_order',
        po_types=PO_TYPES, role=_role(),
    )


def _report_query(kind, args):
    """Build base query depending on report kind. Returns (rows, header_row)."""
    date_from = _parse_date(args.get('date_from', ''))
    date_to   = _parse_date(args.get('date_to', ''))
    po_type   = (args.get('po_type', '') or '').strip().upper()

    base = db.session.query(PurchaseOrder).filter(PurchaseOrder.is_deleted == False)
    if date_from: base = base.filter(PurchaseOrder.po_date >= date_from)
    if date_to:   base = base.filter(PurchaseOrder.po_date <= date_to)
    if po_type:   base = base.filter(PurchaseOrder.po_type == po_type)

    if kind == 'supplier':
        rows_q = db.session.query(
            PurchaseOrder.supplier_name,
            PurchaseOrder.supplier_gst,
            func.count(PurchaseOrder.id),
            func.coalesce(func.sum(PurchaseOrder.taxable_amount), 0),
            func.coalesce(func.sum(PurchaseOrder.grand_total), 0),
        ).filter(PurchaseOrder.is_deleted == False)
        if date_from: rows_q = rows_q.filter(PurchaseOrder.po_date >= date_from)
        if date_to:   rows_q = rows_q.filter(PurchaseOrder.po_date <= date_to)
        if po_type:   rows_q = rows_q.filter(PurchaseOrder.po_type == po_type)
        rows = rows_q.group_by(PurchaseOrder.supplier_name,
                               PurchaseOrder.supplier_gst
                ).order_by(desc(func.sum(PurchaseOrder.grand_total))).all()
        header = ['Supplier', 'GSTIN', 'PO Count', 'Taxable', 'Grand Total']
        data = [[r[0] or '', r[1] or '', int(r[2] or 0),
                 float(r[3] or 0), float(r[4] or 0)] for r in rows]
        return data, header

    if kind == 'item':
        q = db.session.query(
            PurchaseOrderItem.item_name,
            PurchaseOrderItem.hsn_code,
            PurchaseOrderItem.uom,
            func.coalesce(func.sum(PurchaseOrderItem.quantity), 0),
            func.coalesce(func.sum(PurchaseOrderItem.amount), 0),
            func.coalesce(func.sum(PurchaseOrderItem.total_amount), 0),
        ).join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderItem.po_id
        ).filter(PurchaseOrder.is_deleted == False)
        if date_from: q = q.filter(PurchaseOrder.po_date >= date_from)
        if date_to:   q = q.filter(PurchaseOrder.po_date <= date_to)
        if po_type:   q = q.filter(PurchaseOrder.po_type == po_type)
        rows = q.group_by(PurchaseOrderItem.item_name,
                          PurchaseOrderItem.hsn_code,
                          PurchaseOrderItem.uom
                ).order_by(desc(func.sum(PurchaseOrderItem.total_amount))).all()
        header = ['Item', 'HSN', 'UOM', 'Quantity', 'Amount', 'Total Amount']
        data = [[r[0] or '', r[1] or '', r[2] or '',
                 float(r[3] or 0), float(r[4] or 0), float(r[5] or 0)] for r in rows]
        return data, header

    if kind == 'pending':
        rows = base.filter(PurchaseOrder.status.in_([
            PO_STATUS_PENDING, PO_STATUS_APPROVED, PO_STATUS_PARTIAL,
        ])).order_by(PurchaseOrder.po_date.desc()).all()
        header = ['PO Number', 'Date', 'Supplier', 'Type', 'Status',
                  'Pending Qty', 'Grand Total']
        data = [[
            p.po_number, p.po_date.strftime('%d-%m-%Y') if p.po_date else '',
            p.supplier_name or '', p.type_label, p.status,
            float(p.total_quantity or 0) - float(p.received_qty or 0),
            float(p.grand_total or 0),
        ] for p in rows]
        return data, header

    if kind == 'category':
        q = db.session.query(
            PurchaseOrderItem.category,
            func.count(PurchaseOrderItem.id),
            func.coalesce(func.sum(PurchaseOrderItem.quantity), 0),
            func.coalesce(func.sum(PurchaseOrderItem.total_amount), 0),
        ).join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderItem.po_id
        ).filter(PurchaseOrder.is_deleted == False)
        if date_from: q = q.filter(PurchaseOrder.po_date >= date_from)
        if date_to:   q = q.filter(PurchaseOrder.po_date <= date_to)
        if po_type:   q = q.filter(PurchaseOrder.po_type == po_type)
        rows = q.group_by(PurchaseOrderItem.category
                ).order_by(desc(func.sum(PurchaseOrderItem.total_amount))).all()
        header = ['Category', 'Line Count', 'Quantity', 'Total Amount']
        data = [[r[0] or '(Uncategorised)', int(r[1] or 0),
                 float(r[2] or 0), float(r[3] or 0)] for r in rows]
        return data, header

    # Default: date-wise
    rows = base.order_by(PurchaseOrder.po_date.desc(), PurchaseOrder.id.desc()).all()
    header = ['PO Number', 'Date', 'Supplier', 'Type', 'Status',
              'Taxable', 'GST', 'Grand Total']
    data = [[
        p.po_number, p.po_date.strftime('%d-%m-%Y') if p.po_date else '',
        p.supplier_name or '', p.type_label, p.status,
        float(p.taxable_amount or 0),
        float(p.cgst_total or 0) + float(p.sgst_total or 0) + float(p.igst_total or 0),
        float(p.grand_total or 0),
    ] for p in rows]
    return data, header


@po_bp.route('/report/<kind>')
@login_required
def report_view(kind):
    data, header = _report_query(kind, request.args)
    return render_template(
        'purchase_order/report_view.html',
        active_page='purchase_order',
        kind=kind, header=header, data=data,
        title={
            'supplier': 'Supplier Wise PO Report',
            'item'    : 'Item Wise PO Report',
            'pending' : 'Pending PO Report',
            'category': 'Category Wise PO Report',
            'date'    : 'Date Wise Purchase Report',
        }.get(kind, 'PO Report'),
        po_types=PO_TYPES,
        filters=request.args,
        role=_role(),
    )


@po_bp.route('/report/<kind>/export/<fmt>')
@login_required
def report_export(kind, fmt):
    data, header = _report_query(kind, request.args)
    fmt = (fmt or 'excel').lower()

    if fmt == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify(success=False,
                           error='openpyxl not installed. Run: pip install openpyxl'), 500

        wb = Workbook()
        ws = wb.active
        ws.title = (kind.title()[:30] or 'Report')

        # Header
        head_fill = PatternFill('solid', fgColor='1E3A5F')
        head_font = Font(bold=True, color='FFFFFF', size=11)
        for c, h in enumerate(header, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for r, row in enumerate(data, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        # Auto-width columns
        for col in ws.columns:
            mx = 0
            letter = col[0].column_letter
            for cell in col:
                v = '' if cell.value is None else str(cell.value)
                if len(v) > mx:
                    mx = len(v)
            ws.column_dimensions[letter].width = min(max(mx + 2, 10), 50)
        ws.freeze_panes = 'A2'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'po_report_{kind}.xlsx',
                         mimetype=('application/vnd.openxmlformats-officedocument'
                                   '.spreadsheetml.sheet'))

    if fmt == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=10*mm, rightMargin=10*mm,
                                topMargin=10*mm, bottomMargin=10*mm)
        styles = getSampleStyleSheet()
        story = [Paragraph(f'<b>{kind.title()} PO Report</b>', styles['Title']),
                 Spacer(1, 6)]
        rows = [header] + [[str(c) for c in row] for row in data]
        tbl = Table(rows, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#94a3b8')),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING',(0,0), (-1,-1), 4),
            ('RIGHTPADDING',(0,0),(-1,-1), 4),
        ]))
        story.append(tbl)
        doc.build(story)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'po_report_{kind}.pdf',
                         mimetype='application/pdf')

    return jsonify(success=False, error='Unsupported format'), 400


