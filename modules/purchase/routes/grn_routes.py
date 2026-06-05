"""
GRN (Goods Receipt Note) Routes â€” Phase 1
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
Endpoints:
  GET  /grn                          â†’ listing
  GET  /grn/new                      â†’ create form
  GET  /grn/<id>/edit                â†’ edit form
  POST /grn/save                     â†’ save draft / submit
  GET  /grn/<id>/view                â†’ view page
  POST /grn/<id>/submit              â†’ submit (Draft â†’ Completed; locks GRN)
  POST /grn/<id>/cancel              â†’ cancel
  GET  /grn/api/list                 â†’ JSON listing for DataTable
  GET  /grn/api/pos-for-supplier     â†’ POs available for receiving
  GET  /grn/api/po-pending-items/<id>â†’ pending items of a PO
  GET  /grn/api/material/<id>        â†’ material details
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
"""
import os
import re
from datetime import datetime, date, time
from flask import (Blueprint, render_template, request, jsonify,
                   redirect, url_for, flash, abort, current_app)
from flask_login import login_required, current_user
from sqlalchemy import or_, and_, func, desc
from werkzeug.utils import secure_filename

from models import db
from models.grn import (
    GrnMaster, GrnItem, GrnStatusLog,
    GrnStockLedger, GrnBatchStock, GrnScanLog,
    GRN_STATUSES, GRN_STATUS_DRAFT, GRN_STATUS_COMPLETED,
    GRN_STATUS_CANCEL, GRN_TYPES, GRN_STATUS_COLORS,
    SCAN_STATUS_QUARANTINE, SCAN_STATUS_STOCKED_IN, SCAN_STATUS_COLORS,
    DELIVERY_TYPES,
)
from models.purchase_order import (PurchaseOrder, PurchaseOrderItem,
                                   PurchaseOrderStatusLog,
                                   PoShipLocation,
                                   PO_STATUS_DRAFT, PO_STATUS_PENDING,
                                   PO_STATUS_APPROVED, PO_STATUS_REJECTED,
                                   PO_STATUS_PARTIAL, PO_STATUS_COMPLETE,
                                   PO_STATUS_CANCEL)
from models.supplier import Supplier
from models.material import Material, MaterialType
from models.depreciation_note import (DepreciationNote, DepreciationNoteItem,
                                       DN_STATUS_OPEN)

grn_bp = Blueprint('grn', __name__, url_prefix='/grn')


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Helpers
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _username():
    return getattr(current_user, 'full_name', None) or getattr(current_user, 'username', '') or 'system'


def _role():
    return (getattr(current_user, 'role', '') or '').lower()


def _can(action):
    """Lightweight permission gate â€” extend with proper RBAC later."""
    role = _role()
    if role in ('admin', 'director'):
        return True
    if action in ('view', 'create', 'edit', 'submit'):
        return True   # store/warehouse staff can do these
    if action in ('approve', 'reject', 'cancel') and role in ('manager', 'purchase_manager'):
        return True
    return False


def _striptags(s):
    """Strip HTML tags from a string (for plain-text storage)."""
    if not s:
        return ''
    s = re.sub(r'<[^>]+>', '', str(s))
    return (s.replace('&nbsp;', ' ').replace('&amp;', '&')
              .replace('&lt;', '<').replace('&gt;', '>').strip())


def _parse_date(v):
    if not v:
        return None
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(v.strip(), '%Y-%m-%d').date()
    except (ValueError, AttributeError):
        return None


def _parse_time(v):
    if not v:
        return None
    if isinstance(v, time):
        return v
    try:
        return datetime.strptime(v.strip(), '%H:%M').time()
    except (ValueError, AttributeError):
        return None


def _to_float(v, default=0.0):
    try:
        return float(v) if v not in (None, '') else default
    except (ValueError, TypeError):
        return default


def _to_int(v, default=0):
    try:
        return int(float(v)) if v not in (None, '') else default
    except (ValueError, TypeError):
        return default


def _supplier_full_info(sup):
    """Returns a dict with name, address, gst, state, state_code, phone, email
       from a Supplier object â€” handling both legacy `address` and the newer
       `addresses` JSON field, plus billing_state fallback.
    """
    if not sup:
        return {'name':'', 'address':'', 'gst':'', 'state':'', 'state_code':'',
                'phone':'', 'email':''}
    # Address: legacy field first, then primary entry of addresses JSON
    addr = (sup.address or '').strip()
    if not addr:
        try:
            raw = sup.addresses
            if isinstance(raw, str):
                import json as _json
                raw = _json.loads(raw or '[]')
            if raw:
                primary = next((a for a in raw if a.get('is_primary')), raw[0])
                parts = [primary.get('line1',''), primary.get('line2',''),
                         primary.get('city',''),
                         primary.get('state','') +
                         (('-' + primary.get('pincode','')) if primary.get('pincode') else '')]
                addr = ', '.join(p.strip() for p in parts if p and p.strip())
        except Exception:
            pass
    if not addr:
        # Last resort â€” billing city + state
        bits = []
        if getattr(sup, 'billing_city', ''): bits.append(sup.billing_city)
        if getattr(sup, 'billing_state', ''): bits.append(sup.billing_state)
        addr = ', '.join(bits)
    return {
        'name':       sup.supplier_name or '',
        'address':    addr,
        'gst':        sup.gst_number or '',
        'state':      getattr(sup, 'billing_state', '') or '',
        'state_code': getattr(sup, 'billing_state_code', '') or '',
        'phone':      getattr(sup, 'phone', '') or '',
        'email':      getattr(sup, 'email', '') or '',
    }


def _financial_year(d):
    """India financial year: Apr 1 â€“ Mar 31. Returns ('26-27', 2026)."""
    if d.month >= 4:
        return f'{str(d.year)[-2:]}-{str(d.year + 1)[-2:]}', d.year
    return f'{str(d.year - 1)[-2:]}-{str(d.year)[-2:]}', d.year - 1


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STOCK IMPACT ENGINE â€” Phase 3
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _apply_stock_impact(grn):
    """When a GRN moves to Completed, this fires:
       1. For each item with po_item_id: increase PurchaseOrderItem.received_qty
       2. Recalculate parent PO status (Approved â†’ Partial â†’ Completed)
       3. Insert immutable stock_ledger entries (one per item)   [skipped for RM]
       4. Upsert batch_stock table                                [skipped for RM]

       For RM the actual stock-in does NOT happen here â€” it happens via a
       separate QR-sticker scan step (each GRN line gets a printable sticker;
       scanning the sticker is what moves the material into on-hand stock).
    """
    # ALL GRN types now use the QR-scan flow for stock-in.
    # Submit just updates the PO line received_qty + PO status. The actual
    # stock_ledger / batch_stock entries happen later via /grn/api/scan.
    skip_stock = True
    affected_po_ids = set()
    for it in grn.items.all():
        recv = float(it.received_qty or 0)
        if recv <= 0:
            continue

        # â”€â”€ 1. Update PO line item received_qty â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if it.po_item_id:
            po_it = PurchaseOrderItem.query.get(it.po_item_id)
            if po_it:
                prev = float(po_it.received_qty or 0)
                po_it.received_qty = prev + recv
                affected_po_ids.add(po_it.po_id)

        if skip_stock:
            continue   # Skip stock ledger + batch upsert for RM

        # â”€â”€ 3. Stock Ledger entry (immutable) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        db.session.add(GrnStockLedger(
            txn_type     = 'GRN_IN',
            txn_ref_type = 'GRN',
            txn_ref_id   = grn.id,
            txn_ref_no   = grn.grn_number,
            material_id  = it.material_id or 0,
            item_code    = it.item_code or '',
            item_name    = it.item_name or '',
            batch_no     = it.batch_no or '',
            location_id  = it.storage_location_id,
            location_name= it.storage_location_name or grn.receive_location_name or '',
            qty_in       = recv,
            qty_out      = 0,
            uom          = it.uom or 'KG',
            rate         = float(it.rate or 0),
            amount       = float(it.amount or 0),
            remarks      = f'GRN receipt â€” {grn.grn_number}',
            actor_name   = _username(),
        ))

        # â”€â”€ 4. Batch Stock upsert â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if it.material_id:
            batch = GrnBatchStock.query.filter_by(
                material_id=it.material_id,
                batch_no=it.batch_no or '',
                location_id=it.storage_location_id,
            ).first()
            if batch:
                cur_qty = float(batch.qty_on_hand or 0)
                cur_avg = float(batch.avg_rate or 0)
                new_qty = cur_qty + recv
                # Weighted average rate
                if new_qty > 0 and recv > 0 and float(it.rate or 0) > 0:
                    batch.avg_rate = ((cur_avg * cur_qty) + (float(it.rate) * recv)) / new_qty
                batch.qty_on_hand = new_qty
                batch.qty_available = new_qty - float(batch.qty_reserved or 0)
                batch.last_inward_at = datetime.utcnow()
            else:
                db.session.add(GrnBatchStock(
                    material_id   = it.material_id,
                    item_code     = it.item_code or '',
                    item_name     = it.item_name or '',
                    batch_no      = it.batch_no or '',
                    location_id   = it.storage_location_id,
                    location_name = it.storage_location_name or grn.receive_location_name or '',
                    mfg_date      = it.mfg_date,
                    expiry_date   = it.expiry_date,
                    qty_on_hand   = recv,
                    qty_available = recv,
                    avg_rate      = float(it.rate or 0),
                    uom           = it.uom or 'KG',
                    last_inward_at= datetime.utcnow(),
                ))

    # â”€â”€ 2. Recalculate PO statuses â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for po_id in affected_po_ids:
        _recalc_po_status(po_id, note=f'Updated by GRN {grn.grn_number}')

    db.session.flush()


def _next_dn_number():
    """Generate next DN number in DPN-XXXX-FY format."""
    today = date.today()
    # Indian FY: April-March
    if today.month >= 4:
        fy = f"{today.year % 100:02d}-{(today.year + 1) % 100:02d}"
        fy_year = today.year
    else:
        fy = f"{(today.year - 1) % 100:02d}-{today.year % 100:02d}"
        fy_year = today.year - 1

    # Find max serial for this FY
    last = (DepreciationNote.query
            .filter_by(dn_fy=fy)
            .order_by(DepreciationNote.dn_serial.desc())
            .first())
    serial = (last.dn_serial + 1) if last else 1
    dn_number = f"DPN-{serial:04d}-{fy}"
    return dn_number, serial, fy, fy_year


def _reverse_stock_impact(grn):
    """Inverse of _apply_stock_impact() â€” fired when a Completed GRN is Cancelled."""
    # ALL GRN types use the QR-scan flow; ledger/batch entries are managed
    # by per-scan delete (not by GRN cancel). Cancel only reverses the
    # PO received_qty + recalculates PO status.
    skip_stock = True
    affected_po_ids = set()
    for it in grn.items.all():
        recv = float(it.received_qty or 0)
        if recv <= 0:
            continue

        # â”€â”€ 1. Decrement PO line received_qty â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if it.po_item_id:
            po_it = PurchaseOrderItem.query.get(it.po_item_id)
            if po_it:
                po_it.received_qty = max(float(po_it.received_qty or 0) - recv, 0)
                affected_po_ids.add(po_it.po_id)

        if skip_stock:
            continue   # No ledger reverse / batch decrement for RM

        # â”€â”€ 3. Reverse Stock Ledger entry â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        db.session.add(GrnStockLedger(
            txn_type     = 'GRN_REVERSE',
            txn_ref_type = 'GRN',
            txn_ref_id   = grn.id,
            txn_ref_no   = grn.grn_number,
            material_id  = it.material_id or 0,
            item_code    = it.item_code or '',
            item_name    = it.item_name or '',
            batch_no     = it.batch_no or '',
            location_id  = it.storage_location_id,
            location_name= it.storage_location_name or grn.receive_location_name or '',
            qty_in       = 0,
            qty_out      = recv,
            uom          = it.uom or 'KG',
            rate         = float(it.rate or 0),
            amount       = float(it.amount or 0),
            remarks      = f'GRN cancelled â€” {grn.grn_number}',
            actor_name   = _username(),
        ))

        # â”€â”€ 4. Batch Stock decrement â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if it.material_id:
            batch = GrnBatchStock.query.filter_by(
                material_id=it.material_id,
                batch_no=it.batch_no or '',
                location_id=it.storage_location_id,
            ).first()
            if batch:
                batch.qty_on_hand   = max(float(batch.qty_on_hand or 0) - recv, 0)
                batch.qty_available = max(batch.qty_on_hand - float(batch.qty_reserved or 0), 0)
                batch.last_outward_at = datetime.utcnow()

    # â”€â”€ 2. Recalculate PO statuses â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for po_id in affected_po_ids:
        _recalc_po_status(po_id, note=f'Reversed by GRN cancel â€” {grn.grn_number}')

    db.session.flush()


def _recalc_po_status(po_id, note=''):
    """Recalculate a PO's status based on cumulative received qty across items.
       - All items fully received â†’ 'Completed'
       - Some items received      â†’ 'Partial Received'
       - None received            â†’ keep current status (was 'Approved')
    """
    po = PurchaseOrder.query.get(po_id)
    if not po or po.is_deleted:
        return
    # Skip if PO is already Cancelled or in editable states
    if po.status in (PO_STATUS_DRAFT, PO_STATUS_PENDING, PO_STATUS_CANCEL, PO_STATUS_REJECTED):
        return

    items = po.items.all()
    if not items:
        return

    total_ordered = 0.0
    total_received = 0.0
    all_full = True
    any_partial = False
    for it in items:
        ord_q  = float(it.quantity or 0)
        recv_q = float(it.received_qty or 0)
        total_ordered  += ord_q
        total_received += recv_q
        if recv_q < ord_q - 0.001:
            all_full = False
        if recv_q > 0:
            any_partial = True

    new_status = po.status
    if all_full and total_ordered > 0:
        new_status = PO_STATUS_COMPLETE
    elif any_partial:
        new_status = PO_STATUS_PARTIAL
    else:
        new_status = PO_STATUS_APPROVED   # nothing received yet

    if new_status != po.status:
        # Insert status log on the PO side too
        from_status = po.status
        po.status = new_status
        # Lock when Completed (received in full)
        if new_status == PO_STATUS_COMPLETE:
            po.is_locked = True
        try:
            db.session.add(PurchaseOrderStatusLog(
                po_id=po.id,
                from_status=from_status,
                to_status=new_status,
                actor_id=getattr(current_user, 'id', None),
                actor_name=_username(),
                note=note or 'GRN updated received qty',
            ))
        except Exception:
            pass   # status log is best-effort

    # Update PO header received_qty total
    po.received_qty = total_received


def _save_grn_file(file_storage, grn_number, field_label):
    """Save uploaded file under static/uploads/grn/<grn>/<field>_<name>.
       Returns the path relative to /static (for url_for use), or '' on failure.
    """
    if not file_storage or not file_storage.filename:
        return ''
    try:
        safe_grn   = (grn_number or 'unknown').replace('/', '_').replace('\\', '_')
        safe_field = re.sub(r'[^A-Za-z0-9_-]', '', field_label or 'file')[:30]
        safe_name  = secure_filename(file_storage.filename)
        if not safe_name:
            return ''
        # Place under app's static folder so files are served directly
        static_root = os.path.join(current_app.root_path, 'static')
        rel_dir     = os.path.join('uploads', 'grn', safe_grn)
        abs_dir     = os.path.join(static_root, rel_dir)
        os.makedirs(abs_dir, exist_ok=True)
        # Prefix field name + timestamp to avoid collisions
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        fname = f'{safe_field}_{ts}_{safe_name}'
        file_storage.save(os.path.join(abs_dir, fname))
        return os.path.join(rel_dir, fname).replace('\\', '/')
    except Exception as e:
        print(f'[GRN _save_grn_file] error: {e}')
        return ''


def _next_grn_number(grn_type, grn_date):
    """Generate next GRN number â€” separate sequence per type per FY.
       Format: RM-0408-26-27   (also used for short version)
    """
    fy, fy_year = _financial_year(grn_date)
    # Find max serial for this type + fy
    max_serial = db.session.query(func.max(GrnMaster.grn_serial)).filter(
        GrnMaster.grn_type == grn_type,
        GrnMaster.grn_fy   == fy,
        GrnMaster.is_deleted == False,
    ).scalar() or 0
    next_serial = max_serial + 1
    grn_no       = f'{grn_type}-{next_serial:04d}-{fy}'
    grn_no_short = f'{grn_type}-{next_serial:04d}-{fy}'   # same format
    return grn_no, grn_no_short, next_serial, fy, fy_year


def _log_status(grn, to_status, note=''):
    """Add a status-transition entry to the audit log."""
    from_status = grn.status
    db.session.add(GrnStatusLog(
        grn_id      = grn.id,
        from_status = from_status,
        to_status   = to_status,
        actor_id    = getattr(current_user, 'id', None),
        actor_name  = _username(),
        note        = note,
    ))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# LISTING
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/')
@login_required
def index():
    if not _can('view'):
        abort(403)

    grn_type = (request.args.get('grn_type', '') or '').upper().strip()
    return render_template(
        'grn/index.html',
        active_page  = 'grn',
        grn_type     = grn_type,
        grn_types    = GRN_TYPES,
        grn_statuses = GRN_STATUSES,
        grn_type_label = GRN_TYPES.get(grn_type, ''),
        status_colors = GRN_STATUS_COLORS,
        role         = _role(),
    )


@grn_bp.route('/api/list')
@login_required
def api_list():
    """JSON endpoint that powers the index table (search + filter + paginate)."""
    if not _can('view'):
        return jsonify(results=[], total=0), 403

    page   = max(int(request.args.get('page', 1)), 1)
    limit  = max(int(request.args.get('limit', 25)), 1)
    q      = (request.args.get('q', '') or '').strip()
    g_type = (request.args.get('grn_type', '') or '').upper().strip()
    status = (request.args.get('status', '') or '').strip()
    date_from = _parse_date(request.args.get('date_from'))
    date_to   = _parse_date(request.args.get('date_to'))

    qs = GrnMaster.query.filter_by(is_deleted=False)
    if g_type:
        qs = qs.filter(GrnMaster.grn_type == g_type)
    if status:
        qs = qs.filter(GrnMaster.status == status)
    if date_from:
        qs = qs.filter(GrnMaster.grn_date >= date_from)
    if date_to:
        qs = qs.filter(GrnMaster.grn_date <= date_to)
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            GrnMaster.grn_number.ilike(like),
            GrnMaster.po_number.ilike(like),
            GrnMaster.supplier_name.ilike(like),
            GrnMaster.invoice_no.ilike(like),
        ))

    total = qs.count()
    rows = qs.order_by(desc(GrnMaster.id)).offset((page - 1) * limit).limit(limit).all()
    return jsonify(results=[r.to_dict() for r in rows], total=total,
                   page=page, limit=limit)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# CREATE / EDIT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/new')
@login_required
def new_grn():
    if not _can('create'):
        abort(403)
    grn_type = (request.args.get('grn_type', 'RM') or 'RM').upper()
    if grn_type not in GRN_TYPES:
        grn_type = 'RM'

    # Preview the GRN number so the form can show it
    preview_grn_no, preview_short, _, fy, _ = _next_grn_number(grn_type, date.today())

    return render_template(
        'grn/form.html',
        active_page    = 'grn',
        mode           = 'new',
        grn            = None,
        items          = [],
        grn_type       = grn_type,
        grn_types      = GRN_TYPES,
        delivery_types = DELIVERY_TYPES,
        preview_grn_no = preview_grn_no,
        preview_short  = preview_short,
        role           = _role(),
    )


@grn_bp.route('/<int:grn_id>/edit')
@login_required
def edit_grn(grn_id):
    if not _can('edit'):
        abort(403)
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        flash('GRN has been deleted.', 'danger')
        return redirect(url_for('grn.index'))
    if not grn.is_editable:
        flash(f'GRN is locked (status: {grn.status}).', 'warning')
        return redirect(url_for('grn.view_grn', grn_id=grn.id))

    items = grn.items.order_by(GrnItem.sr_no).all()
    items_data = [it.to_dict() for it in items]
    return render_template(
        'grn/form.html',
        active_page    = 'grn',
        mode           = 'edit',
        grn            = grn,
        items          = items,
        items_data     = items_data,
        grn_type       = grn.grn_type,
        grn_types      = GRN_TYPES,
        delivery_types = DELIVERY_TYPES,
        preview_grn_no = grn.grn_number,
        preview_short  = grn.grn_number_short,
        role           = _role(),
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SAVE (handles both create + update)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/save', methods=['POST'])
@login_required
def save_grn():
    if not _can('create'):
        return jsonify(success=False, error='Permission denied'), 403

    grn_id     = request.form.get('grn_id', '').strip()
    is_edit    = bool(grn_id)
    grn_type   = (request.form.get('grn_type', 'RM') or 'RM').upper()
    grn_date   = _parse_date(request.form.get('grn_date')) or date.today()
    action     = request.form.get('action', 'draft').lower()   # 'draft' or 'submit'

    try:
        if is_edit:
            grn = GrnMaster.query.get(int(grn_id))
            if not grn or grn.is_deleted:
                return jsonify(success=False, error='GRN not found'), 404
            if not grn.is_editable:
                return jsonify(success=False,
                               error=f'GRN locked (status: {grn.status})'), 400

            grn.grn_type = grn_type
            grn.grn_date = grn_date

            # Wipe old items; re-create from form payload
            for it in grn.items.all():
                db.session.delete(it)
            db.session.flush()
        else:
            grn_no, grn_short, serial, fy, year = _next_grn_number(grn_type, grn_date)
            grn = GrnMaster(
                grn_type=grn_type,
                grn_number=grn_no,
                grn_number_short=grn_short,
                grn_serial=serial,
                grn_fy=fy,
                grn_year=year,
                grn_date=grn_date,
                created_by_id=getattr(current_user, 'id', None),
                created_by_name=_username(),
            )
            db.session.add(grn)
            db.session.flush()

        # â”€â”€ Header fields â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # GRN-Without-PO flow: when the user picks "NA" in the primary supplier
        # dropdown and chooses from the manual (all-suppliers) one, the form
        # posts is_without_po=1 and the supplier_id comes from the manual
        # dropdown. The form already consolidates this into the hidden
        # `supplier_id` field, but we also accept `manual_supplier_id` as a
        # fallback for clients that post both fields separately.
        is_without_po = (request.form.get('is_without_po', '0') or '0').strip() in ('1', 'true', 'True', 'yes', 'on')
        grn.is_without_po = is_without_po

        if is_without_po:
            # PO references must be cleared â€” this GRN has no parent PO.
            grn.po_id     = None
            grn.po_number = ''
            grn.po_date   = None
            # Prefer manual_supplier_id when provided; fall back to supplier_id.
            sup_id = (_to_int(request.form.get('manual_supplier_id'))
                      or _to_int(request.form.get('supplier_id'))
                      or None)
            grn.supplier_id = sup_id
        else:
            grn.po_id        = _to_int(request.form.get('po_id')) or None
            grn.po_number    = request.form.get('po_number', '').strip()
            grn.po_date      = _parse_date(request.form.get('po_date'))
            grn.supplier_id  = _to_int(request.form.get('supplier_id')) or None

        grn.supplier_name = request.form.get('supplier_name', '').strip()
        grn.supplier_address = request.form.get('supplier_address', '').strip()

        # Always look up supplier master and backfill snapshot fields whenever
        # they came empty in the form. This is the source of truth for supplier
        # info â€” the JS hidden fields can fail to populate (Select2 race etc.),
        # so we trust the FK and fetch fresh.
        if grn.supplier_id:
            _sup = Supplier.query.filter_by(id=grn.supplier_id, is_deleted=False).first()
            if _sup:
                _info = _supplier_full_info(_sup)
                if not grn.supplier_name:    grn.supplier_name    = _info['name']
                if not grn.supplier_address: grn.supplier_address = _info['address']

        # Server-side validation: supplier is required in BOTH flows.
        if not grn.supplier_id:
            db.session.rollback()
            msg = ('Manual supplier is required when creating a GRN without PO.'
                   if is_without_po else 'Supplier is required.')
            return jsonify(success=False, error=msg), 400

        grn.invoice_no   = request.form.get('invoice_no', '').strip()
        grn.invoice_date = _parse_date(request.form.get('invoice_date'))

        # â”€â”€ Invoice file upload â”€â”€
        inv_file = request.files.get('invoice_file')
        if inv_file and inv_file.filename:
            saved_path = _save_grn_file(inv_file, grn.grn_number, 'invoice')
            if saved_path:
                grn.invoice_file = saved_path
        # else: keep existing invoice_file (edit mode w/o re-upload)

        grn.receive_location_id   = _to_int(request.form.get('receive_location_id')) or None
        grn.receive_location_name = request.form.get('receive_location_name', '').strip()
        grn.receive_location_address = request.form.get('receive_location_address', '').strip()

        grn.gate_inward_no   = request.form.get('gate_inward_no', '').strip()
        grn.gate_inward_date = _parse_date(request.form.get('gate_inward_date'))
        grn.gate_inward_time = _parse_time(request.form.get('gate_inward_time'))
        grn.unloading_time   = _parse_time(request.form.get('unloading_time'))

        grn.lr_no            = request.form.get('lr_no', '').strip()
        grn.lr_date          = _parse_date(request.form.get('lr_date'))
        grn.logistics_name   = request.form.get('logistics_name', '').strip()
        grn.delivery_type    = request.form.get('delivery_type', '').strip()
        grn.driver_name      = request.form.get('driver_name', '').strip()
        grn.driver_contact   = request.form.get('driver_contact', '').strip()
        grn.vehicle_no       = request.form.get('vehicle_no', '').strip()
        grn.supervisor_name  = request.form.get('supervisor_name', '').strip()

        # Quality checklist
        grn.qc_test_certificate   = bool(request.form.get('qc_test_certificate'))
        grn.qc_batch_on_product   = bool(request.form.get('qc_batch_on_product'))
        grn.qc_physical_condition = bool(request.form.get('qc_physical_condition'))
        grn.qc_expiry_date        = bool(request.form.get('qc_expiry_date'))
        grn.qc_label_checked      = bool(request.form.get('qc_label_checked'))
        grn.rejection_remarks     = request.form.get('rejection_remarks', '').strip()

        grn.supplier_remarks = request.form.get('supplier_remarks', '').strip()
        grn.internal_remarks = request.form.get('internal_remarks', '').strip()

        grn.updated_by_name = _username()

        # â”€â”€ Item rows â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        item_names         = request.form.getlist('item_name[]')
        material_ids       = request.form.getlist('material_id[]')
        po_item_ids        = request.form.getlist('po_item_id[]')
        po_numbers_row     = request.form.getlist('po_number_row[]')
        item_codes         = request.form.getlist('item_code[]')
        manufacturers      = request.form.getlist('manufacturer[]')
        categories         = request.form.getlist('category[]')
        hsn_codes          = request.form.getlist('hsn_code[]')
        uoms               = request.form.getlist('uom[]')
        batch_nos          = request.form.getlist('batch_no[]')
        mfg_dates          = request.form.getlist('mfg_date[]')
        expiry_dates       = request.form.getlist('expiry_date[]')
        boxes              = request.form.getlist('no_of_boxes[]')
        per_box_qtys       = request.form.getlist('per_box_qty[]')
        ordered_qtys       = request.form.getlist('ordered_qty[]')
        already_recv_qtys  = request.form.getlist('already_received_qty[]')
        remaining_qtys     = request.form.getlist('remaining_qty[]')
        received_qtys      = request.form.getlist('received_qty[]')
        accepted_qtys      = request.form.getlist('accepted_qty[]')
        rejected_qtys      = request.form.getlist('rejected_qty[]')
        rates              = request.form.getlist('rate[]')
        gst_pcts           = request.form.getlist('gst_pct[]')
        storage_loc_ids    = request.form.getlist('storage_location_id[]')
        storage_loc_names  = request.form.getlist('storage_location_name[]')
        rejection_reasons  = request.form.getlist('rejection_reason[]')
        item_remarks       = request.form.getlist('remarks[]')

        sr = 0
        total_ord = total_recv = total_acc = total_rej = 0
        total_box = 0
        total_amt = 0
        validation_errors = []
        for i, name in enumerate(item_names):
            name = (name or '').strip()
            if not name:
                continue
            sr += 1
            recv = _to_float(received_qtys[i] if i < len(received_qtys) else 0)
            acc  = _to_float(accepted_qtys[i] if i < len(accepted_qtys) else 0)
            rej  = _to_float(rejected_qtys[i] if i < len(rejected_qtys) else 0)
            rate = _to_float(rates[i]         if i < len(rates) else 0)
            ordered = _to_float(ordered_qtys[i]      if i < len(ordered_qtys) else 0)
            already = _to_float(already_recv_qtys[i] if i < len(already_recv_qtys) else 0)
            amount = recv * rate

            # â”€â”€ Server-side validation â”€â”€
            # Over-receipt (received > ordered - already): allowed. Frontend
            # asks the user to confirm before submitting. Server accepts it.
            # 1. Received cannot be negative
            if recv < 0:
                validation_errors.append(f'Row {sr} "{name}": Received qty cannot be negative')

            total_ord  += ordered
            total_recv += recv
            total_acc  += acc
            total_rej  += rej
            total_box  += _to_int(boxes[i] if i < len(boxes) else 0)
            total_amt  += amount

            item = GrnItem(
                grn_id      = grn.id,
                sr_no       = sr,
                po_item_id  = None if grn.is_without_po else (_to_int(po_item_ids[i] if i < len(po_item_ids) else None) or None),
                po_number   = '' if grn.is_without_po else (po_numbers_row[i] if i < len(po_numbers_row) else '').strip(),
                material_id = _to_int(material_ids[i] if i < len(material_ids) else None) or None,
                item_code   = (item_codes[i] if i < len(item_codes) else '').strip(),
                item_name   = name,
                category    = (categories[i] if i < len(categories) else '').strip(),
                manufacturer= (manufacturers[i] if i < len(manufacturers) else '').strip(),
                hsn_code    = (hsn_codes[i] if i < len(hsn_codes) else '').strip(),
                uom         = (uoms[i] if i < len(uoms) else 'KG').strip() or 'KG',
                batch_no    = (batch_nos[i] if i < len(batch_nos) else '').strip(),
                mfg_date    = _parse_date(mfg_dates[i]    if i < len(mfg_dates) else ''),
                expiry_date = _parse_date(expiry_dates[i] if i < len(expiry_dates) else ''),
                no_of_boxes = _to_int(boxes[i] if i < len(boxes) else 0),
                per_box_qty = _to_float(per_box_qtys[i] if i < len(per_box_qtys) else 0),
                ordered_qty = ordered,
                already_received_qty = already,
                remaining_qty = max(ordered - already - recv, 0),
                received_qty = recv,
                accepted_qty = acc,
                rejected_qty = rej,
                rate         = rate,
                gst_pct      = _to_float(gst_pcts[i] if i < len(gst_pcts) else 0),
                amount       = amount,
                storage_location_id   = _to_int(storage_loc_ids[i] if i < len(storage_loc_ids) else None) or None,
                storage_location_name = (storage_loc_names[i] if i < len(storage_loc_names) else '').strip(),
                qc_passed    = (rej == 0),
                rejection_reason = (rejection_reasons[i] if i < len(rejection_reasons) else '').strip(),
                remarks      = (item_remarks[i] if i < len(item_remarks) else '').strip(),
            )
            # Per-item COA upload â€” files named coa_file_<i>
            coa = request.files.get(f'coa_file_{i}')
            if coa and coa.filename:
                saved = _save_grn_file(coa, grn.grn_number, f'coa_sr{sr}')
                if saved:
                    item.coa_file = saved
            db.session.add(item)

        # â”€â”€ Validation gate before commit â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if sr == 0:
            db.session.rollback()
            return jsonify(success=False,
                           error='At least one item is required.'), 400
        if validation_errors:
            db.session.rollback()
            return jsonify(success=False,
                           error='Validation failed:\nâ€¢ ' + '\nâ€¢ '.join(validation_errors)), 400

        grn.total_ordered_qty  = total_ord
        grn.total_received_qty = total_recv
        grn.total_accepted_qty = total_acc
        grn.total_rejected_qty = total_rej
        grn.total_box_qty      = total_box
        grn.total_amount       = total_amt

        # Status transition based on action
        if action == 'submit':
            if grn.status == GRN_STATUS_DRAFT:
                _log_status(grn, GRN_STATUS_COMPLETED,
                            'GRN submitted â€” stock receipt recorded')
                grn.status = GRN_STATUS_COMPLETED
                grn.is_locked = True
                grn.submitted_at = datetime.utcnow()
                grn.submitted_by_id = getattr(current_user, 'id', None)
                grn.submitted_by_name = _username()
                db.session.flush()
                # Phase 3 stock impact: update PO received_qty + ledger + batch
                _apply_stock_impact(grn)
        elif not is_edit:
            _log_status(grn, GRN_STATUS_DRAFT, 'GRN created as draft')

        db.session.commit()
        return jsonify(success=True, grn_id=grn.id, grn_number=grn.grn_number,
                       redirect_url=url_for('grn.view_grn', grn_id=grn.id))
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# VIEW
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/<int:grn_id>/view')
@login_required
def view_grn(grn_id):
    if not _can('view'):
        abort(403)
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        flash('GRN has been deleted.', 'danger')
        return redirect(url_for('grn.index'))

    items = grn.items.order_by(GrnItem.sr_no).all()
    status_logs = grn.status_logs.order_by(GrnStatusLog.created_at).all()

    # Load attached Depreciation Note (latest, if any)
    depreciation_note = None
    if grn.has_depreciation_note:
        depreciation_note = (DepreciationNote.query
                             .filter_by(grn_id=grn.id, is_deleted=False)
                             .order_by(DepreciationNote.created_at.desc())
                             .first())

    # Live supplier info (for GST, address, state) â€” falls back to snapshot
    supplier_info = None
    if grn.supplier_id:
        _sup = Supplier.query.filter_by(id=grn.supplier_id, is_deleted=False).first()
        if _sup:
            supplier_info = _supplier_full_info(_sup)

    return render_template(
        'grn/view.html',
        active_page   = 'grn',
        grn           = grn,
        items         = items,
        supplier_info = supplier_info,
        status_logs   = status_logs,
        depreciation_note = depreciation_note,
        grn_types     = GRN_TYPES,
        role          = _role(),
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STATUS ACTIONS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/<int:grn_id>/delete', methods=['POST'])
@login_required
def delete_grn(grn_id):
    """Soft-delete a Draft GRN. Completed GRNs cannot be deleted (must be cancelled)."""
    if not _can('cancel'):
        return jsonify(success=False, error='Permission denied'), 403
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.status != GRN_STATUS_DRAFT:
        return jsonify(success=False,
                       error='Only Draft GRNs can be deleted. Cancel a completed GRN instead.'), 400
    grn.is_deleted = True
    grn.deleted_at = datetime.utcnow()
    grn.deleted_by_name = _username()
    db.session.commit()
    return jsonify(success=True)


@grn_bp.route('/<int:grn_id>/submit', methods=['POST'])
@login_required
def submit_grn(grn_id):
    if not _can('submit'):
        return jsonify(success=False, error='Permission denied'), 403
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.status != GRN_STATUS_DRAFT:
        return jsonify(success=False,
                       error=f'Cannot submit (status: {grn.status})'), 400
    try:
        _log_status(grn, GRN_STATUS_COMPLETED, 'GRN submitted â€” stock receipt recorded')
        grn.status = GRN_STATUS_COMPLETED
        grn.is_locked = True
        grn.submitted_at = datetime.utcnow()
        grn.submitted_by_id = getattr(current_user, 'id', None)
        grn.submitted_by_name = _username()

        # Stock impact: update PO received_qty + Stock Ledger + Batch Stock
        _apply_stock_impact(grn)

        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=f'Submit failed: {e}'), 500


@grn_bp.route('/<int:grn_id>/cancel', methods=['POST'])
@login_required
def cancel_grn(grn_id):
    if not _can('cancel'):
        return jsonify(success=False, error='Permission denied'), 403
    grn = GrnMaster.query.get_or_404(grn_id)
    if not grn.can_cancel:
        return jsonify(success=False,
                       error=f'Cannot cancel (status: {grn.status})'), 400

    reason = (request.form.get('reason') or '').strip()
    was_completed = (grn.status == GRN_STATUS_COMPLETED)

    try:
        _log_status(grn, GRN_STATUS_CANCEL, f'Cancelled: {reason}')
        grn.status = GRN_STATUS_CANCEL
        grn.cancelled_by_id = getattr(current_user, 'id', None)
        grn.cancelled_by_name = _username()
        grn.cancelled_at = datetime.utcnow()
        grn.cancel_reason = reason

        # If GRN was already Completed, reverse the stock impact
        if was_completed:
            _reverse_stock_impact(grn)

        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify(success=False, error=f'Cancel failed: {e}'), 500


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SUPPORTING APIs (for form dropdowns + auto-fetch)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/api/suppliers')
@login_required
def api_suppliers():
    """Suppliers for the GRN form's PRIMARY supplier picker.

    Returns ONLY suppliers who have at least one PO that is:
        â€¢ status = Approved (PO open, nothing received yet), OR
        â€¢ status = Partial Received (PO partially completed, still pending)
    i.e. suppliers whose PO is still pending for GRN.

    Optionally filters PO type to match GRN type (RM â†” RM, PM/COR/SLV â†” PM-family).

    Always appends a sentinel option with id='__NA__' at the end so the user can
    choose to create a GRN without a PO. Selecting it opens a secondary dropdown
    on the client that lists ALL suppliers from master (api_all_suppliers).
    """
    grn_type = (request.args.get('grn_type', '') or '').upper()
    q = (request.args.get('q', '') or '').strip()

    # Sub-query: supplier IDs that have at least one open/partial PO
    po_sup_q = db.session.query(PurchaseOrder.supplier_id).filter(
        PurchaseOrder.is_deleted == False,
        PurchaseOrder.supplier_id.isnot(None),
        PurchaseOrder.status.in_([PO_STATUS_APPROVED, PO_STATUS_PARTIAL]),
    )
    # Match PO type to GRN type so we don't show suppliers whose only
    # pending PO is of an unrelated material family.
    if grn_type == 'RM':
        po_sup_q = po_sup_q.filter(PurchaseOrder.po_type == 'RM')
    elif grn_type in ('PM', 'COR', 'SLV'):
        po_sup_q = po_sup_q.filter(PurchaseOrder.po_type.in_(['PM', 'COR', 'SLV']))
    po_sup_ids = po_sup_q.distinct().subquery()

    qs = (Supplier.query
          .filter_by(is_deleted=False)
          .filter(Supplier.id.in_(po_sup_ids)))
    # Defence-in-depth: also keep the legacy supplier_type sanity filter so a
    # mismatched master row never sneaks in.
    if grn_type == 'RM':
        qs = qs.filter(Supplier.supplier_type.ilike('%RM%'))
    elif grn_type in ('PM', 'COR', 'SLV'):
        qs = qs.filter(Supplier.supplier_type.ilike('%PM%'))
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            Supplier.supplier_name.ilike(like),
            Supplier.gst_number.ilike(like),
        ))
    rows = qs.order_by(Supplier.supplier_name).limit(50).all()
    results = []
    for s in rows:
        info = _supplier_full_info(s)
        results.append({
            'id':      s.id,
            'text':    f'{s.supplier_name}' + (f' ({s.gst_number})' if s.gst_number else ''),
            'name':    info['name'],
            'address': info['address'],
            'gst':     info['gst'],
            'state':   info['state'],
            'state_code': info['state_code'],
            'phone':   info['phone'],
            'email':   info['email'],
        })

    # Always append the "NA" sentinel as the final option. When the user picks
    # it, the form reveals the secondary supplier dropdown (all suppliers).
    # Only include when the search query is empty or matches "NA" loosely, so
    # Select2 search still feels natural.
    if not q or 'na' in q.lower():
        results.append({
            'id':      '__NA__',
            'text':    'NA â€” Create GRN without PO',
            'name':    '',
            'address': '',
            'gst':     '',
            'state':   '',
            'state_code': '',
            'phone':   '',
            'email':   '',
            'is_na':   True,
        })
    return jsonify(results=results)


@grn_bp.route('/api/all-suppliers')
@login_required
def api_all_suppliers():
    """SECONDARY (manual) supplier dropdown â€” returns ALL suppliers from the
    master, without PO filtering. Used when the user picks "NA" in the primary
    supplier dropdown to create a GRN without a PO.

    Still respects grn_type when given so RM GRNs see RM suppliers, etc.,
    purely as a quality-of-life filter; pass grn_type='' to get every supplier.
    """
    grn_type = (request.args.get('grn_type', '') or '').upper()
    q = (request.args.get('q', '') or '').strip()

    qs = Supplier.query.filter_by(is_deleted=False)
    if grn_type == 'RM':
        qs = qs.filter(Supplier.supplier_type.ilike('%RM%'))
    elif grn_type in ('PM', 'COR', 'SLV'):
        qs = qs.filter(Supplier.supplier_type.ilike('%PM%'))
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            Supplier.supplier_name.ilike(like),
            Supplier.gst_number.ilike(like),
        ))
    rows = qs.order_by(Supplier.supplier_name).limit(100).all()
    results = []
    for s in rows:
        info = _supplier_full_info(s)
        results.append({
            'id':      s.id,
            'text':    f'{s.supplier_name}' + (f' ({s.gst_number})' if s.gst_number else ''),
            'name':    info['name'],
            'address': info['address'],
            'gst':     info['gst'],
            'state':   info['state'],
            'state_code': info['state_code'],
            'phone':   info['phone'],
            'email':   info['email'],
        })
    return jsonify(results=results)


@grn_bp.route('/api/pos-for-supplier/<int:sup_id>')
@login_required
def api_pos_for_supplier(sup_id):
    """POs available to receive from this supplier (Approved or Partial)."""
    grn_type = (request.args.get('grn_type', '') or '').upper()

    qs = PurchaseOrder.query.filter(
        PurchaseOrder.supplier_id == sup_id,
        PurchaseOrder.is_deleted == False,
        PurchaseOrder.status.in_([PO_STATUS_APPROVED, PO_STATUS_PARTIAL]),
    )
    if grn_type:
        # For COR/SLV types we still pick PM POs (sub-types share material_type)
        po_type_filter = grn_type if grn_type in ('RM',) else 'PM'
        if grn_type in ('COR', 'SLV', 'PM'):
            qs = qs.filter(PurchaseOrder.po_type.in_(['PM', 'COR', 'SLV']))
        else:
            qs = qs.filter(PurchaseOrder.po_type == po_type_filter)
    rows = qs.order_by(desc(PurchaseOrder.po_date)).limit(50).all()
    return jsonify(results=[{
        'id': p.id,
        'po_number': p.po_number,
        'po_date':   p.po_date.strftime('%Y-%m-%d') if p.po_date else '',
        'po_date_disp': p.po_date.strftime('%d-%m-%Y') if p.po_date else '',
        'text': f'{p.po_number} â€” {p.po_date.strftime("%d-%m-%Y") if p.po_date else ""}',
        'grand_total': float(p.grand_total or 0),
        'po_type': p.po_type,
    } for p in rows])


@grn_bp.route('/api/po-pending-items/<int:po_id>')
@login_required
def api_po_pending_items(po_id):
    """Fetch PO line items with pending qty (ordered - already_received).
       For items already fully received, exclude from results.
    """
    po = PurchaseOrder.query.get_or_404(po_id)
    rows = po.items.order_by(PurchaseOrderItem.sr_no).all()

    items = []
    for it in rows:
        ordered = float(it.quantity or 0)
        already_recv = float(getattr(it, 'received_qty', 0) or 0)
        pending = max(ordered - already_recv, 0)
        if pending <= 0:
            continue  # fully received, skip
        items.append({
            'po_item_id':  it.id,
            'material_id': it.material_id,
            'item_code':   it.item_code or '',
            'item_name':   it.item_name or '',
            'category':    it.category or '',
            'hsn_code':    it.hsn_code or '',
            'uom':         it.uom or 'KG',
            'ordered_qty': ordered,
            'already_received_qty': already_recv,
            'remaining_qty': pending,
            'rate':        float(it.rate or 0),
            'gst_pct':     float(it.gst_pct or 0),
        })
    return jsonify(
        po_id=po.id, po_number=po.po_number,
        po_date=po.po_date.strftime('%Y-%m-%d') if po.po_date else '',
        supplier_id=po.supplier_id,
        supplier_name=po.supplier_name,
        supplier_gst=po.supplier_gst,
        supplier_address=po.supplier_address,
        items=items,
    )


@grn_bp.route('/api/ship-locations')
@login_required
def api_ship_locations():
    """Re-uses the PO module's ship locations master."""
    qs = PoShipLocation.query.filter_by(is_active=True, is_deleted=False
        ).order_by(PoShipLocation.sort_order, PoShipLocation.name).all()
    return jsonify(results=[l.to_dict() for l in qs])


@grn_bp.route('/<int:grn_id>/pdf')
@login_required
def pdf_grn(grn_id):
    """Generate a Tally-style printable GRN PDF."""
    from flask import send_file, make_response
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        abort(404)
    items = grn.items.order_by(GrnItem.sr_no).all()

    supplier_info = None
    if grn.supplier_id:
        _sup = Supplier.query.filter_by(id=grn.supplier_id, is_deleted=False).first()
        if _sup:
            supplier_info = _supplier_full_info(_sup)

    pdf_bytes = _generate_grn_pdf(grn, items, supplier_info)
    safe = grn.grn_number.replace('/', '_').replace('\\', '_')
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename="GRN_{safe}.pdf"'
    return response


def _generate_grn_pdf(grn, items, supplier_info=None):
    """Generate the GRN PDF using reportlab â€” Tally-style layout."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, BaseDocTemplate,
                                    PageTemplate, Frame)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # Try registering a Unicode font for â‚¹ glyph
    UNICODE_FONT = None
    for cand_name, cand_paths in [
        ('SegoeUI',   [r'C:\Windows\Fonts\segoeui.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf']),
        ('NirmalaUI', [r'C:\Windows\Fonts\Nirmala.ttf']),
        ('Arial',     [r'C:\Windows\Fonts\arial.ttf']),
        ('DejaVuSans',['/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf']),
    ]:
        for p in cand_paths:
            if os.path.exists(p):
                try:
                    pdfmetrics.registerFont(TTFont(cand_name, p))
                    UNICODE_FONT = cand_name
                    bold_p = p.replace('.ttf', 'b.ttf').replace('.TTF', 'b.TTF')
                    if os.path.exists(bold_p):
                        try:
                            pdfmetrics.registerFont(TTFont(cand_name + '-Bold', bold_p))
                        except Exception:
                            pass
                    break
                except Exception:
                    continue
        if UNICODE_FONT:
            break

    _font      = UNICODE_FONT or 'Helvetica'
    _font_bold = (UNICODE_FONT + '-Bold') if (UNICODE_FONT and UNICODE_FONT + '-Bold' in pdfmetrics.getRegisteredFontNames()) else (UNICODE_FONT or 'Helvetica-Bold')

    # Department label per type
    dept_map = {'RM':'RM Store', 'PM':'PM Store', 'COR':'Corrugation Store',
                'SLV':'Sleeves Store', 'FG':'FG Store'}
    dept = dept_map.get(grn.grn_type, grn.grn_type + ' Store')

    buf = BytesIO()
    doc = BaseDocTemplate(buf, pagesize=A4,
                          leftMargin=8*mm, rightMargin=8*mm,
                          topMargin=8*mm, bottomMargin=10*mm)
    page_w, page_h = A4
    avail_w = page_w - 16*mm

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle('title',  parent=styles['Normal'],
                              fontSize=13, leading=15, alignment=TA_CENTER,
                              fontName=_font_bold, textColor=colors.HexColor('#0f172a'))
    st_sub   = ParagraphStyle('sub',    parent=styles['Normal'],
                              fontSize=10, leading=12, alignment=TA_CENTER,
                              fontName=_font_bold, textColor=colors.HexColor('#475569'))
    st_lbl   = ParagraphStyle('lbl',    parent=styles['Normal'],
                              fontSize=7.5, leading=9, alignment=TA_LEFT,
                              fontName=_font, textColor=colors.HexColor('#64748b'))
    st_val   = ParagraphStyle('val',    parent=styles['Normal'],
                              fontSize=10, leading=12, alignment=TA_LEFT,
                              fontName=_font_bold)
    st_name  = ParagraphStyle('name',   parent=styles['Normal'],
                              fontSize=10, leading=12.5, alignment=TA_LEFT,
                              fontName=_font_bold)
    st_addr  = ParagraphStyle('addr',   parent=styles['Normal'],
                              fontSize=8.5, leading=10.5, alignment=TA_LEFT,
                              fontName=_font)
    st_gst   = ParagraphStyle('gst',    parent=styles['Normal'],
                              fontSize=8.5, leading=11, alignment=TA_LEFT,
                              fontName=_font)
    st_cellc = ParagraphStyle('cellc',  parent=styles['Normal'],
                              fontSize=8, leading=10, alignment=TA_CENTER,
                              fontName=_font_bold)
    st_cell  = ParagraphStyle('cell',   parent=styles['Normal'],
                              fontSize=8.5, leading=11, alignment=TA_LEFT,
                              fontName=_font)
    st_desc_name = ParagraphStyle('descname', parent=styles['Normal'],
                              fontSize=9.5, leading=11, alignment=TA_LEFT,
                              fontName=_font_bold)
    st_meta  = ParagraphStyle('meta',   parent=styles['Normal'],
                              fontSize=7.5, leading=10, alignment=TA_LEFT,
                              fontName=_font, textColor=colors.HexColor('#475569'))
    st_num   = ParagraphStyle('num',    parent=styles['Normal'],
                              fontSize=9, leading=11, alignment=TA_RIGHT,
                              fontName=_font)
    st_num_b = ParagraphStyle('numb',   parent=styles['Normal'],
                              fontSize=9.5, leading=11, alignment=TA_RIGHT,
                              fontName=_font_bold)
    st_footer = ParagraphStyle('foot',  parent=styles['Normal'],
                              fontSize=9, leading=12, alignment=TA_CENTER,
                              fontName=_font)
    st_footer_b = ParagraphStyle('footb',parent=styles['Normal'],
                              fontSize=9.5, leading=12, alignment=TA_CENTER,
                              fontName=_font_bold)

    # â”€â”€ Build header table (LEFT 3 stacked / RIGHT 2-col grid) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    def L_block(title, name, addr_lines, gst='', state=''):
        flow = []
        if title:
            flow.append(Paragraph(title, st_lbl))
        if name:
            flow.append(Paragraph(name, st_name))
        for line in (addr_lines or []):
            if line:
                flow.append(Paragraph(line, st_addr))
        if gst:
            flow.append(Spacer(1, 2))
            flow.append(Paragraph(f'<font color="#334155"><b>GSTIN/UIN:</b></font> {gst}', st_gst))
        if state:
            flow.append(Paragraph(state, st_gst))
        return flow

    bill_to = L_block('', 'HCP Wellness Pvt Ltd',
                      ['403 Maruti Vertex Elanza,',
                       'Opp. Global Hospital, Nr. GTPL House,',
                       'Sindhubhavan Road, Bodakdev,',
                       'Ahmedabad-380054'],
                      gst='24AAFCH7246H1ZK',
                      state='<font color="#334155"><b>State Name :</b></font> Gujarat, <font color="#334155"><b>Code :</b></font> 24')

    ship_addr = (grn.receive_location_address or
                 'Plot No. 8, Ozone Industrial Estate, Beside Kerala GIDC, Bavla-Bagodara Road, Bhayla, Bavla-382220, Gujarat, India.')
    ship_to = L_block('Consignee (Ship to)',
                      grn.receive_location_name or 'HCP Wellness Private Limited',
                      [ship_addr])

    sup_addr_lines = (grn.supplier_address or '').split('\n') or ['']
    sup_gst   = (supplier_info or {}).get('gst', '')
    sup_state = (supplier_info or {}).get('state', '')
    sup_code  = (supplier_info or {}).get('state_code', '')
    sup_state_line = ''
    if sup_state:
        sup_state_line = f'<font color="#334155"><b>State Name:</b></font> {sup_state}'
        if sup_code:
            sup_state_line += f', <font color="#334155"><b>Code:</b></font> {sup_code}'
    supplier = L_block('Supplier (Bill from)',
                       grn.supplier_name or 'â€”',
                       sup_addr_lines,
                       gst=sup_gst,
                       state=sup_state_line)

    left_block = [
        [bill_to],
        [ship_to],
        [supplier],
    ]
    left_tbl = Table(left_block, colWidths=[avail_w * 0.55])
    left_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor('#1e293b')),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
    ]))

    # â”€â”€ Right grid (2 columns, multiple rows) â”€â”€
    def R_cell(label, value):
        return [Paragraph(label, st_lbl), Paragraph(value or 'â€”', st_val)]

    dept_row = [
        Paragraph(f'<font color="#64748b">Department :</font> <b>{dept}</b>', st_val)
    ]

    grn_date_s = grn.grn_date.strftime('%d-%b-%Y') if grn.grn_date else 'â€”'
    gate_dt = ''
    if grn.gate_inward_date:
        gate_dt = grn.gate_inward_date.strftime('%d-%m-%Y')
        if grn.gate_inward_time:
            gate_dt += ' & ' + grn.gate_inward_time.strftime('%H:%M:%S')

    inv_date_s = grn.invoice_date.strftime('%d-%m-%Y') if grn.invoice_date else 'â€”'

    right_rows = [
        [dept_row, ''],  # full-width department; second col empty (merged via SPAN)
        [R_cell('Receipt Note No.', grn.grn_number_short or grn.grn_number),
         R_cell('Dated :', grn_date_s)],
        [R_cell('Logistic Name', grn.logistics_name or 'â€”'),
         R_cell('Unloading Time', grn.unloading_time.strftime('%H:%M:%S') if grn.unloading_time else 'â€”')],
        [R_cell('Driver Name', grn.driver_name or 'â€”'),
         R_cell('Driver Contact No.', grn.driver_contact or 'â€”')],
        [R_cell('LR No', grn.lr_no or 'â€”'),
         R_cell('LR Date', grn.lr_date.strftime('%d-%m-%Y') if grn.lr_date else 'â€”')],
        [R_cell('GATE INWARD NO', grn.gate_inward_no or 'â€”'),
         R_cell('DATE & TIME', gate_dt or 'â€”')],
        [R_cell('Vehicle No', grn.vehicle_no or 'â€”'),
         R_cell('Delivery Type', grn.delivery_type or 'â€”')],
        [R_cell('Invoice No', grn.invoice_no or 'â€”'),
         R_cell('Invoice Date', inv_date_s)],
        [[Paragraph('Supervisor Name', st_lbl),
          Paragraph(grn.supervisor_name or 'â€”', st_val)], ''],
    ]
    right_tbl = Table(right_rows, colWidths=[avail_w * 0.225, avail_w * 0.225])
    right_tbl.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#1e293b')),
        ('SPAN',         (0,0), (1,0)),    # Department full-width
        ('SPAN',         (0,8), (1,8)),    # Supervisor full-width (last row, idx 8)
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))

    header_tbl = Table([[left_tbl, right_tbl]], colWidths=[avail_w * 0.55, avail_w * 0.45])
    header_tbl.setStyle(TableStyle([
        ('BOX',     (0,0), (-1,-1), 0.6, colors.HexColor('#1e293b')),
        ('LINEBEFORE',(1,0), (1,0), 0.6, colors.HexColor('#1e293b')),
        ('VALIGN',  (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 0),
        ('BOTTOMPADDING',(0,0), (-1,-1), 0),
    ]))

    # â”€â”€ Items table â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    item_hdr = [
        Paragraph('Sl<br/>No', st_cellc),
        Paragraph('PO No.', st_cellc),
        Paragraph('PO Date', st_cellc),
        Paragraph('Description of Goods', st_cellc),
        Paragraph('No.<br/>of pkt', st_cellc),
        Paragraph('Per Pkt<br/>Qty', st_cellc),
        Paragraph('Total Qty', st_cellc),
        Paragraph('COA', st_cellc),
        Paragraph('Remarks', st_cellc),
    ]

    total_box = 0
    total_recv = 0
    item_uom = items[0].uom if items else 'KG'
    rows = [item_hdr]
    for it in items:
        # Description cell â€” stacked
        desc_html = [Paragraph(it.item_name or '', st_desc_name)]
        meta_lines = []
        if it.storage_location_name:
            meta_lines.append(f'<i>Unload Location :</i> {it.storage_location_name}')
        # Manufacturer / Batch / MFG / EXP only for Raw Material GRNs
        if grn.grn_type == 'RM':
            meta_lines.append(f'<i>Batch No :</i> {it.batch_no or "N/A"}')
            mfg = it.mfg_date.strftime('%d-%m-%Y') if it.mfg_date else 'N/A'
            exp = it.expiry_date.strftime('%d-%m-%Y') if it.expiry_date else 'N/A'
            meta_lines.append(f'<i>Mfg. Date :</i> {mfg} &nbsp;/&nbsp; <i>Exp. Date :</i> {exp}')
            if it.manufacturer:
                meta_lines.append(f'<i>Manufacturer :</i> {it.manufacturer}')
        desc_html.append(Paragraph('<br/>'.join(meta_lines), st_meta))

        # COA cell
        coa_cell = Paragraph('[COA]', st_cellc) if it.coa_file else Paragraph('â€”', st_cellc)
        remarks_cell = Paragraph(it.remarks or '', st_cell)

        po_date_s = 'â€”'
        if grn.po_date and (it.po_number == grn.po_number):
            po_date_s = grn.po_date.strftime('%d-%m-%Y')

        rows.append([
            Paragraph(str(it.sr_no), st_cellc),
            Paragraph(it.po_number or 'N/A', st_cell),
            Paragraph(po_date_s, st_cellc),
            desc_html,
            Paragraph(f'{float(it.no_of_boxes or 0):.3f}', st_num),
            Paragraph(f'{float(it.per_box_qty or 0):.3f} {it.uom}', st_num),
            Paragraph(f'<b>{float(it.received_qty or 0):.3f} {it.uom}</b>', st_num_b),
            coa_cell,
            remarks_cell,
        ])
        total_box  += int(it.no_of_boxes or 0)
        total_recv += float(it.received_qty or 0)

    # Total row
    rows.append([
        '', '', '',
        Paragraph('<b>TOTAL</b>', st_cellc),
        Paragraph(f'<b>{total_box}</b>', st_num_b),
        '',
        Paragraph(f'<b>{total_recv:.3f} {item_uom}</b>', st_num_b),
        '',
        '',
    ])

    col_widths = [
        avail_w * 0.04,   # Sl No
        avail_w * 0.12,   # PO No
        avail_w * 0.08,   # PO Date
        avail_w * 0.32,   # Description
        avail_w * 0.07,   # No. of pkt
        avail_w * 0.09,   # Per Pkt
        avail_w * 0.10,   # Total Recv
        avail_w * 0.06,   # COA
        avail_w * 0.12,   # Remarks
    ]
    item_tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    item_tbl.setStyle(TableStyle([
        ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#1e293b')),
        ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#f8fafc')),
        ('BACKGROUND',   (0,-1), (-1,-1),colors.HexColor('#fafbfd')),
        ('SPAN',         (0,-1), (3,-1)),   # TOTAL spans first 4 cols
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))

    # â”€â”€ Quality Checklist block (only if any QC was ticked) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Batch Number on Product / Expiry Date Checked are RM-only.
    _is_rm_qc = (grn.grn_type == 'RM')
    qc_flags = [
        ('Test Certificate Received', grn.qc_test_certificate),
    ]
    if _is_rm_qc:
        qc_flags.append(('Batch Number on Product', grn.qc_batch_on_product))
    qc_flags.append(('Physical Condition OK', grn.qc_physical_condition))
    if _is_rm_qc:
        qc_flags.append(('Expiry Date Checked', grn.qc_expiry_date))
    qc_flags.append(('Labels Verified', grn.qc_label_checked))
    qc_block = None
    if any(v for _, v in qc_flags):
        qc_cells = []
        for label, v in qc_flags:
            mark = 'â˜‘' if v else 'â˜'
            qc_cells.append(Paragraph(f'<font size="10">{mark}</font> {label}', st_cell))
        # 3 columns
        while len(qc_cells) % 3 != 0:
            qc_cells.append('')
        qc_grid = [qc_cells[i:i+3] for i in range(0, len(qc_cells), 3)]
        qc_block = Table(
            [[Paragraph('<b>Quality Checklist</b>', st_lbl), '', '']] + qc_grid,
            colWidths=[avail_w/3]*3)
        qc_block.setStyle(TableStyle([
            ('BOX',          (0,0), (-1,-1), 0.5, colors.HexColor('#1e293b')),
            ('LINEABOVE',    (0,1), (-1,1),  0.4, colors.HexColor('#1e293b')),
            ('SPAN',         (0,0), (2,0)),
            ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#fafbfd')),
            ('LEFTPADDING',  (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING',   (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0), (-1,-1), 4),
        ]))

    # â”€â”€ Remarks block (only if filled) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    rmk_block = None
    if grn.supplier_remarks or grn.internal_remarks:
        rmk_block = Table([[
            [Paragraph('<b>Supplier Remarks</b>', st_lbl),
             Paragraph(grn.supplier_remarks or 'â€”', st_val)],
            [Paragraph('<b>Internal Remarks</b>', st_lbl),
             Paragraph(grn.internal_remarks or 'â€”', st_val)],
        ]], colWidths=[avail_w/2, avail_w/2])
        rmk_block.setStyle(TableStyle([
            ('BOX',          (0,0), (-1,-1), 0.5, colors.HexColor('#1e293b')),
            ('LINEBEFORE',   (1,0), (1,-1), 0.5, colors.HexColor('#1e293b')),
            ('VALIGN',       (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',  (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING',   (0,0), (-1,-1), 5),
            ('BOTTOMPADDING',(0,0), (-1,-1), 5),
        ]))

    # â”€â”€ Signature block â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    sig_block = Table([[
        '',
        [Paragraph('For HCP Wellness Pvt Ltd', st_val),
         Spacer(1, 28),
         Paragraph(grn.supervisor_name or grn.created_by_name or '', st_val),
         Paragraph('Received Signatory', st_addr)],
    ]], colWidths=[avail_w * 0.55, avail_w * 0.45])
    sig_block.setStyle(TableStyle([
        ('BOX',          (0,0), (-1,-1), 0.6, colors.HexColor('#1e293b')),
        ('LINEBEFORE',   (1,0), (1,0),  0.5, colors.HexColor('#1e293b')),
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING',   (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
    ]))

    # â”€â”€ Story â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    title_block = Table([
        [Paragraph('Goods Receipt Note', st_title)],
        [Paragraph(grn.grn_number_short or grn.grn_number, st_sub)],
    ], colWidths=[avail_w])
    title_block.setStyle(TableStyle([
        ('BOX',          (0,0), (-1,-1), 0.6, colors.HexColor('#1e293b')),
        ('LINEBELOW',    (0,0), (-1,0), 0.5, colors.HexColor('#1e293b')),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('LEFTPADDING',  (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING',   (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
    ]))

    story = [title_block, header_tbl, item_tbl]
    if qc_block:  story.append(qc_block)
    if rmk_block: story.append(rmk_block)
    story.extend([
        sig_block,
        Spacer(1, 8),
        Paragraph('SUBJECT TO AHMEDABAD JURISDICTION', st_footer),
        Paragraph('This is a Computer Generated Document', st_footer_b),
    ])

    # Build with footer-only page template
    def _draw_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(_font, 7.5)
        canvas.setFillColor(colors.HexColor('#64748b'))
        canvas.drawRightString(page_w - 8*mm, 5*mm, f'Page {doc_.page}')
        canvas.drawString(8*mm, 5*mm, f'GRN No: {grn.grn_number}')
        canvas.restoreState()

    body_frame = Frame(8*mm, 10*mm, avail_w, page_h - 18*mm,
                       leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
                       showBoundary=0)
    doc.addPageTemplates([
        PageTemplate(id='grn_pages', frames=[body_frame], onPage=_draw_footer),
    ])
    doc.build(story)
    buf.seek(0)
    return buf.read()


# â”€â”€ api_stock_ledger / api_batch_stock endpoints removed (UI deprecated) â”€â”€


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# RM QUARANTINE â€” listing of all scanned RM boxes still in Quarantine,
# grouped by (GRN, item, per-box-qty). Loose boxes (different per-box-qty)
# show as separate rows.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/quarantine')
@login_required
def quarantine_page():
    """RM Quarantine listing page."""
    if not _can('view'):
        abort(403)
    return render_template('grn/quarantine.html', active_page='grn')


@grn_bp.route('/api/quarantine')
@login_required
def api_quarantine():
    """Return all quarantined scans grouped by (grn, item, per-box-qty)."""
    if not _can('view'):
        return jsonify(success=False, error='Permission denied'), 403

    q          = (request.args.get('q', '') or '').strip()
    date_from  = _parse_date(request.args.get('date_from'))
    date_to    = _parse_date(request.args.get('date_to'))

    # Aggregation in SQL â€” group by grn_id, grn_item_id, qty
    from sqlalchemy import func
    qs = (db.session.query(
            GrnScanLog.grn_id,
            GrnScanLog.grn_number,
            GrnScanLog.grn_item_id,
            GrnScanLog.material_id,
            GrnScanLog.item_code,
            GrnScanLog.item_name,
            GrnScanLog.batch_no,
            GrnScanLog.uom,
            GrnScanLog.qty.label('per_box_qty'),
            func.count(GrnScanLog.id).label('no_of_boxes'),
            func.sum(GrnScanLog.qty).label('total_qty'),
            func.min(GrnScanLog.scanned_at).label('first_at'),
            func.max(GrnScanLog.scanned_at).label('last_at'),
          )
          .filter(GrnScanLog.is_deleted == False,
                  GrnScanLog.status     == SCAN_STATUS_QUARANTINE))

    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            GrnScanLog.grn_number.ilike(like),
            GrnScanLog.item_name.ilike(like),
            GrnScanLog.item_code.ilike(like),
            GrnScanLog.batch_no.ilike(like),
        ))

    qs = qs.group_by(GrnScanLog.grn_id, GrnScanLog.grn_number,
                     GrnScanLog.grn_item_id, GrnScanLog.material_id,
                     GrnScanLog.item_code, GrnScanLog.item_name,
                     GrnScanLog.batch_no, GrnScanLog.uom, GrnScanLog.qty)
    qs = qs.order_by(func.max(GrnScanLog.scanned_at).desc())

    # Fetch grn_date in one shot (avoid N+1)
    rows = qs.all()
    grn_ids = list({r.grn_id for r in rows})
    grn_dates = {}
    if grn_ids:
        gs = GrnMaster.query.filter(GrnMaster.id.in_(grn_ids)).all()
        for g in gs:
            grn_dates[g.id] = g

    out = []
    for r in rows:
        g = grn_dates.get(r.grn_id)
        # Apply GRN-date filter here (after group, against parent GRN date)
        if g is None: continue
        if date_from and g.grn_date and g.grn_date < date_from: continue
        if date_to   and g.grn_date and g.grn_date > date_to:   continue
        out.append({
            'grn_id':       r.grn_id,
            'grn_number':   r.grn_number or '',
            'grn_date':     g.grn_date.strftime('%d-%m-%Y') if g and g.grn_date else '',
            'grn_type':     g.grn_type if g else '',
            'supplier_name':g.supplier_name if g else '',
            'item_code':    r.item_code or '',
            'item_name':    r.item_name or '',
            'batch_no':     r.batch_no  or '',
            'uom':          r.uom       or 'KG',
            'per_box_qty':  float(r.per_box_qty or 0),
            'no_of_boxes':  int(r.no_of_boxes or 0),
            'total_qty':    float(r.total_qty or 0),
            'first_scanned':r.first_at.strftime('%d-%m-%Y %H:%M') if r.first_at else '',
            'last_scanned': r.last_at.strftime('%d-%m-%Y %H:%M')  if r.last_at  else '',
        })

    # Grand totals
    grand_boxes = sum(r['no_of_boxes'] for r in out)
    grand_total = sum(r['total_qty']   for r in out)

    return jsonify(success=True,
                   results=out,
                   row_count=len(out),
                   grand_boxes=grand_boxes,
                   grand_total=grand_total)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STOCK VIEW PAGES â€” Current Stock (batch-wise) & Stock Movement Ledger
# Permission: same as GRN view (anyone who can view GRN).
# Data source: tbl_grn_batch_stock (running totals) + tbl_grn_stock_ledger (movements),
# both populated by the scan flow for non-RM GRNs (PM/COR/SLV/FG).
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _stock_type():
    """Stock view material-type from query string: 'rm', 'pm', or '' (all)."""
    return (request.args.get('type', '') or '').strip().lower()


def _material_ids_by_type(mat_type):
    """material_ids for the stock view.
       'rm' â†’ RM items only; 'pm' â†’ non-RM (PM/COR/SLV/FG); else â†’ None (no filter)."""
    if mat_type not in ('rm', 'pm'):
        return None
    abbr = func.upper(func.coalesce(MaterialType.abbreviation, ''))
    base = (db.session.query(Material.id)
            .outerjoin(MaterialType, Material.material_type_id == MaterialType.id))
    base = base.filter(abbr == 'RM') if mat_type == 'rm' else base.filter(abbr != 'RM')
    return [row[0] for row in base.all()]


@grn_bp.route('/stock')
@login_required
def stock_page():
    """Current Stock (batch-wise) listing page."""
    if not _can('view'):
        abort(403)
    return render_template('grn/stock.html', active_page='grn', mat_type=_stock_type())


@grn_bp.route('/stock-ledger')
@login_required
def stock_ledger_page():
    """Stock Movement Ledger page."""
    if not _can('view'):
        abort(403)
    return render_template('grn/stock_ledger.html', active_page='grn', mat_type=_stock_type())


@grn_bp.route('/api/batch-stock')
@login_required
def api_batch_stock():
    """Return current batch-wise stock for the Stock View page.

    Query params:
      q             â€” search item_name / item_code / batch_no  (LIKE)
      location_id   â€” filter by storage location
      expiry_filter â€” 'expired' | 'expiring' | 'fresh'
      page, limit   â€” pagination (default 1, 50)
    """
    if not _can('view'):
        return jsonify(success=False, error='Permission denied'), 403

    page  = max(int(request.args.get('page',  1) or 1), 1)
    limit = max(min(int(request.args.get('limit', 50) or 50), 200), 1)
    q             = (request.args.get('q', '') or '').strip()
    location_id   = (request.args.get('location_id', '') or '').strip()
    expiry_filter = (request.args.get('expiry_filter', '') or '').strip().lower()

    qs = GrnBatchStock.query.filter(GrnBatchStock.qty_on_hand > 0)

    _mat_ids = _material_ids_by_type(_stock_type())
    if _mat_ids is not None:
        qs = qs.filter(GrnBatchStock.material_id.in_(_mat_ids or [-1]))

    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            GrnBatchStock.item_name.ilike(like),
            GrnBatchStock.item_code.ilike(like),
            GrnBatchStock.batch_no.ilike(like),
        ))

    if location_id:
        try:
            qs = qs.filter(GrnBatchStock.location_id == int(location_id))
        except ValueError:
            pass

    today = date.today()
    if expiry_filter == 'expired':
        qs = qs.filter(GrnBatchStock.expiry_date.isnot(None),
                       GrnBatchStock.expiry_date < today)
    elif expiry_filter == 'expiring':
        from datetime import timedelta
        cutoff = today + timedelta(days=90)
        qs = qs.filter(GrnBatchStock.expiry_date.isnot(None),
                       GrnBatchStock.expiry_date >= today,
                       GrnBatchStock.expiry_date < cutoff)
    elif expiry_filter == 'fresh':
        from datetime import timedelta
        cutoff = today + timedelta(days=90)
        qs = qs.filter(or_(GrnBatchStock.expiry_date.is_(None),
                           GrnBatchStock.expiry_date >= cutoff))

    total = qs.count()
    rows = (qs.order_by(desc(GrnBatchStock.last_inward_at),
                        desc(GrnBatchStock.id))
              .offset((page - 1) * limit).limit(limit).all())

    results = [{
        'material_id'   : r.material_id,
        'item_name'     : r.item_name or '',
        'item_code'     : r.item_code or '',
        'batch_no'      : r.batch_no  or '',
        'location'      : r.location_name or '',
        'location_id'   : r.location_id,
        'qty_on_hand'   : float(r.qty_on_hand   or 0),
        'qty_reserved'  : float(r.qty_reserved  or 0),
        'qty_available' : float(r.qty_available or 0),
        'uom'           : r.uom or 'KG',
        'avg_rate'      : float(r.avg_rate or 0),
        'mfg_date'      : r.mfg_date.strftime('%d-%m-%Y')       if r.mfg_date       else '',
        'expiry_date'   : r.expiry_date.strftime('%d-%m-%Y')    if r.expiry_date    else '',
        'last_inward_at': r.last_inward_at.strftime('%d-%m-%Y %H:%M') if r.last_inward_at else '',
    } for r in rows]

    return jsonify(success=True, total=total, page=page, limit=limit, results=results)


@grn_bp.route('/api/stock-ledger')
@login_required
def api_stock_ledger():
    """Return stock-ledger transactions for the Stock Ledger page.

    Query params:
      q          â€” search item_name / item_code / batch_no / txn_ref_no  (LIKE)
      txn_type   â€” 'GRN_IN' | 'GRN_REVERSE' | ''  (empty = all)
      date_from  â€” yyyy-mm-dd inclusive
      date_to    â€” yyyy-mm-dd inclusive
      page, limit
    """
    if not _can('view'):
        return jsonify(success=False, error='Permission denied'), 403

    page  = max(int(request.args.get('page',  1) or 1), 1)
    limit = max(min(int(request.args.get('limit', 50) or 50), 200), 1)
    q         = (request.args.get('q', '') or '').strip()
    txn_type  = (request.args.get('txn_type', '') or '').strip()
    date_from = _parse_date(request.args.get('date_from'))
    date_to   = _parse_date(request.args.get('date_to'))

    qs = GrnStockLedger.query

    _mat_ids = _material_ids_by_type(_stock_type())
    if _mat_ids is not None:
        qs = qs.filter(GrnStockLedger.material_id.in_(_mat_ids or [-1]))

    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            GrnStockLedger.item_name.ilike(like),
            GrnStockLedger.item_code.ilike(like),
            GrnStockLedger.batch_no.ilike(like),
            GrnStockLedger.txn_ref_no.ilike(like),
        ))
    if txn_type:
        qs = qs.filter(GrnStockLedger.txn_type == txn_type)
    if date_from:
        qs = qs.filter(GrnStockLedger.txn_date >= datetime.combine(date_from, time.min))
    if date_to:
        qs = qs.filter(GrnStockLedger.txn_date <= datetime.combine(date_to, time.max))

    total = qs.count()
    rows = (qs.order_by(desc(GrnStockLedger.txn_date), desc(GrnStockLedger.id))
              .offset((page - 1) * limit).limit(limit).all())

    results = [{
        'txn_date'  : r.txn_date.strftime('%d-%m-%Y %H:%M') if r.txn_date else '',
        'txn_type'  : r.txn_type   or '',
        'ref_no'    : r.txn_ref_no or '',
        'item_name' : r.item_name  or '',
        'item_code' : r.item_code  or '',
        'batch_no'  : r.batch_no   or '',
        'location'  : r.location_name or '',
        'qty_in'    : float(r.qty_in  or 0),
        'qty_out'   : float(r.qty_out or 0),
        'uom'       : r.uom  or 'KG',
        'rate'      : float(r.rate   or 0),
        'amount'    : float(r.amount or 0),
        'actor'     : r.actor_name or '',
    } for r in rows]

    return jsonify(success=True, total=total, page=page, limit=limit, results=results)



@grn_bp.route('/stock-by-grn')
@login_required
def stock_by_grn_page():
    """Stock attributed to each GRN â€” one row per (GRN Ã— item Ã— location)."""
    if not _can('view'):
        abort(403)
    return render_template('grn/stock_by_grn.html', active_page='grn', mat_type=_stock_type())


def _stock_by_grn_rm(page, limit, q, location_id, date_from, date_to):
    """RM 'Stock by GRN' â€” sourced from the QC stock-in ledger (TRS_QC_IN),
       resolved back to the originating GRN via TrsMaster.grn_id."""
    from models.trs import TrsMaster

    rm_ids = _material_ids_by_type('rm')
    led_q = GrnStockLedger.query.filter(GrnStockLedger.txn_type == 'TRS_QC_IN')
    if rm_ids is not None:
        led_q = led_q.filter(GrnStockLedger.material_id.in_(rm_ids or [-1]))
    if location_id:
        try:
            led_q = led_q.filter(GrnStockLedger.location_id == int(location_id))
        except ValueError:
            pass
    ledgers = led_q.all()

    # Resolve TRS â†’ GRN in two batched lookups
    trs_ids = list({l.txn_ref_id for l in ledgers if l.txn_ref_id})
    trs_map = {t.id: t for t in TrsMaster.query.filter(TrsMaster.id.in_(trs_ids)).all()} if trs_ids else {}
    grn_ids = list({t.grn_id for t in trs_map.values() if t.grn_id})
    grn_map = {g.id: g for g in GrnMaster.query.filter(GrnMaster.id.in_(grn_ids)).all()} if grn_ids else {}

    # Aggregate by (grn_id, material_id, location_id)
    agg = {}
    for l in ledgers:
        trs = trs_map.get(l.txn_ref_id)
        gid = trs.grn_id if trs else None
        g   = grn_map.get(gid)
        if date_from and (not g or not g.grn_date or g.grn_date < date_from):
            continue
        if date_to and (not g or not g.grn_date or g.grn_date > date_to):
            continue
        key = (gid, l.material_id, l.location_id)
        a = agg.get(key)
        if not a:
            a = {
                'grn_id': gid,
                'grn_number': (g.grn_number if g else (l.txn_ref_no or '')),
                'grn_date': g.grn_date.strftime('%d-%m-%Y') if g and g.grn_date else '',
                'grn_type': (g.grn_type if g else 'RM') or 'RM',
                'grn_status': g.status if g else '',
                'supplier_name': g.supplier_name if g else '',
                'material_id': l.material_id, 'item_code': l.item_code or '',
                'item_name': l.item_name or '', 'location_id': l.location_id,
                'location_name': l.location_name or '', 'uom': l.uom or 'KG',
                'box_count': 0, 'total_qty': 0.0, 'total_amount': 0.0,
                '_rate_sum': 0.0, '_rate_n': 0,
                'first_at': l.txn_date, 'last_at': l.txn_date,
            }
            agg[key] = a
        a['box_count']    += 1
        a['total_qty']    += float(l.qty_in or 0)
        a['total_amount'] += float(l.amount or 0)
        a['_rate_sum']    += float(l.rate or 0)
        a['_rate_n']      += 1
        if l.txn_date:
            if not a['first_at'] or l.txn_date < a['first_at']:
                a['first_at'] = l.txn_date
            if not a['last_at'] or l.txn_date > a['last_at']:
                a['last_at'] = l.txn_date

    out = []
    for a in agg.values():
        if q:
            ql  = q.lower()
            hay = ' '.join([a['grn_number'], a['item_name'], a['item_code'], a['supplier_name']]).lower()
            if ql not in hay:
                continue
        avg_rate = (a['_rate_sum'] / a['_rate_n']) if a['_rate_n'] else 0
        out.append({
            'grn_id': a['grn_id'], 'grn_number': a['grn_number'], 'grn_date': a['grn_date'],
            'grn_type': a['grn_type'], 'grn_status': a['grn_status'], 'supplier_name': a['supplier_name'],
            'material_id': a['material_id'], 'item_code': a['item_code'], 'item_name': a['item_name'],
            'location_id': a['location_id'], 'location_name': a['location_name'], 'uom': a['uom'],
            'box_count': a['box_count'], 'total_qty': a['total_qty'], 'avg_rate': avg_rate,
            'total_amount': a['total_amount'],
            'first_inward': a['first_at'].strftime('%d-%m-%Y %H:%M') if a['first_at'] else '',
            'last_inward': a['last_at'].strftime('%d-%m-%Y %H:%M') if a['last_at'] else '',
            '_last_at_iso': a['last_at'].isoformat() if a['last_at'] else '',
        })

    out.sort(key=lambda d: d['_last_at_iso'], reverse=True)
    for d in out:
        d.pop('_last_at_iso', None)

    summary = {
        'grn_count' : len({d['grn_id']      for d in out}),
        'item_count': len({d['material_id'] for d in out if d['material_id']}),
        'qty_sum'   : sum(d['total_qty']    for d in out),
        'box_sum'   : sum(d['box_count']    for d in out),
    }
    total = len(out)
    page_rows = out[(page - 1) * limit : page * limit]
    return jsonify(success=True, total=total, page=page, limit=limit,
                   results=page_rows, summary=summary)


@grn_bp.route('/api/stock-by-grn')
@login_required
def api_stock_by_grn():
    """Aggregated stock-in per GRN, grouped by (grn_id, material_id, location_id).

    Source: tbl_grn_scan_log (status = Stocked-In, is_deleted = False).
    The scan log is the source of truth because each row has direct grn_id;
    tbl_grn_stock_ledger only points at scan-log ids via txn_ref_id.

    Query params:
      q           â€” search grn_number / item_name / item_code / supplier_name
      grn_type    â€” PM / COR / SLV / FG  (case-insensitive)
      location_id â€” filter by storage location
      date_from   â€” GRN date >= this (yyyy-mm-dd)
      date_to     â€” GRN date <= this (yyyy-mm-dd)
      page, limit
    """
    if not _can('view'):
        return jsonify(success=False, error='Permission denied'), 403

    page  = max(int(request.args.get('page',  1) or 1), 1)
    limit = max(min(int(request.args.get('limit', 50) or 50), 200), 1)
    q          = (request.args.get('q', '') or '').strip()
    grn_type   = (request.args.get('grn_type', '') or '').strip().upper()
    location_id= (request.args.get('location_id', '') or '').strip()
    date_from  = _parse_date(request.args.get('date_from'))
    date_to    = _parse_date(request.args.get('date_to'))

    # RM stock-by-GRN is sourced from the QC stock-in ledger (TRS_QC_IN),
    # since RM never reaches the 'Stocked-In' scan status (it is QC-approved in).
    if _stock_type() == 'rm':
        return _stock_by_grn_rm(page, limit, q, location_id, date_from, date_to)

    # Aggregate scan log â†’ one row per (grn, item, location)
    base = (db.session.query(
                GrnScanLog.grn_id.label('grn_id'),
                GrnScanLog.grn_number.label('grn_number'),
                GrnScanLog.material_id.label('material_id'),
                GrnScanLog.item_code.label('item_code'),
                GrnScanLog.item_name.label('item_name'),
                GrnScanLog.location_id.label('location_id'),
                GrnScanLog.location_name.label('location_name'),
                GrnScanLog.uom.label('uom'),
                func.count(GrnScanLog.id).label('box_count'),
                func.sum(GrnScanLog.qty).label('total_qty'),
                func.sum(GrnScanLog.amount).label('total_amount'),
                func.avg(GrnScanLog.rate).label('avg_rate'),
                func.min(GrnScanLog.scanned_at).label('first_at'),
                func.max(GrnScanLog.scanned_at).label('last_at'),
            )
            .filter(GrnScanLog.is_deleted == False,
                    GrnScanLog.status     == SCAN_STATUS_STOCKED_IN))

    if location_id:
        try:
            base = base.filter(GrnScanLog.location_id == int(location_id))
        except ValueError:
            pass

    base = base.group_by(GrnScanLog.grn_id, GrnScanLog.grn_number,
                         GrnScanLog.material_id, GrnScanLog.item_code,
                         GrnScanLog.item_name, GrnScanLog.location_id,
                         GrnScanLog.location_name, GrnScanLog.uom)

    rows = base.all()

    # Hydrate GRN master metadata in one shot
    grn_ids = list({r.grn_id for r in rows if r.grn_id})
    grn_map = {}
    if grn_ids:
        for g in GrnMaster.query.filter(GrnMaster.id.in_(grn_ids)).all():
            grn_map[g.id] = g

    out = []
    for r in rows:
        g = grn_map.get(r.grn_id)
        # GRN-level filters (type, date, supplier-aware search)
        if grn_type and (not g or (g.grn_type or '').upper() != grn_type):
            continue
        if date_from and (not g or not g.grn_date or g.grn_date < date_from):
            continue
        if date_to   and (not g or not g.grn_date or g.grn_date > date_to):
            continue
        if q:
            ql = q.lower()
            hay = ' '.join([
                (r.grn_number or ''), (r.item_name or ''), (r.item_code or ''),
                (g.supplier_name if g else '') or '',
            ]).lower()
            if ql not in hay:
                continue
        out.append({
            'grn_id'       : r.grn_id,
            'grn_number'   : r.grn_number or '',
            'grn_date'     : g.grn_date.strftime('%d-%m-%Y') if g and g.grn_date else '',
            'grn_type'     : g.grn_type if g else '',
            'grn_status'   : g.status   if g else '',
            'supplier_name': g.supplier_name if g else '',
            'material_id'  : r.material_id,
            'item_code'    : r.item_code or '',
            'item_name'    : r.item_name or '',
            'location_id'  : r.location_id,
            'location_name': r.location_name or '',
            'uom'          : r.uom or 'KG',
            'box_count'    : int(r.box_count or 0),
            'total_qty'    : float(r.total_qty or 0),
            'avg_rate'     : float(r.avg_rate or 0),
            'total_amount' : float(r.total_amount or 0),
            'first_inward' : r.first_at.strftime('%d-%m-%Y %H:%M') if r.first_at else '',
            'last_inward'  : r.last_at.strftime('%d-%m-%Y %H:%M')  if r.last_at  else '',
            '_last_at_iso' : r.last_at.isoformat() if r.last_at else '',
        })

    # Sort: most recent inward first
    out.sort(key=lambda d: d['_last_at_iso'], reverse=True)
    for d in out: d.pop('_last_at_iso', None)

    # Summary across the full filtered set (not just the page)
    summary = {
        'grn_count' : len({d['grn_id']      for d in out}),
        'item_count': len({d['material_id'] for d in out if d['material_id']}),
        'qty_sum'   : sum(d['total_qty']    for d in out),
        'box_sum'   : sum(d['box_count']    for d in out),
    }

    total = len(out)
    page_rows = out[(page - 1) * limit : page * limit]

    return jsonify(success=True, total=total, page=page, limit=limit,
                   results=page_rows, summary=summary)



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# QR SCAN â€” receive boxes one by one
# QR payload format: "{TYPE}{GRN_ITEM_ID}-{PO_ITEM_ID}-{BOX_NO}"
#   e.g.  RM15-23-1   â† Raw Material, GrnItem.id=15, PO item id=23, box 1
# Behaviour:
#   â€¢ RM GRN  â†’ logged with status='Quarantine' (no stock ledger, no batch stock)
#   â€¢ Other types â†’ status='Stocked-In' + GrnStockLedger row + GrnBatchStock upsert
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
import re
_QR_RE = re.compile(r'^([A-Z]+)(\d+)-(\d+)-(\d+)$')


@grn_bp.route('/<int:grn_id>/scan')
@login_required
def scan_page(grn_id):
    """The QR scan & receive screen for a SPECIFIC GRN. Camera/generic flow
    has been removed â€” this is a per-GRN page with item-wise box visualization."""
    if not _can('create'):
        abort(403)
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        flash('GRN has been deleted.', 'danger')
        return redirect(url_for('grn.index'))
    if grn.status != GRN_STATUS_COMPLETED:
        flash(f'Scanning is only available on Completed GRNs. This one is "{grn.status}".', 'warning')
        return redirect(url_for('grn.view_grn', grn_id=grn.id))
    items = grn.items.order_by(GrnItem.sr_no).all()
    return render_template('grn/scan.html',
                           active_page='grn',
                           grn=grn,
                           items=items)


@grn_bp.route('/api/<int:grn_id>/scan-status')
@login_required
def api_grn_scan_status(grn_id):
    """For the scan page: return item-wise pending/scanned boxes for this GRN."""
    if not _can('view'):
        return jsonify(success=False, error='Permission denied'), 403
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        return jsonify(success=False, error='GRN deleted'), 404

    # All scans for this GRN (newest first)
    scans = (GrnScanLog.query
             .filter_by(grn_id=grn.id, is_deleted=False)
             .order_by(GrnScanLog.id.desc())
             .all())

    # Build a set: (grn_item_id, box_no) â†’ scan_dict for quick lookup
    scanned_map = {}
    for s in scans:
        scanned_map[(s.grn_item_id, s.box_no or 1)] = s

    items_out = []
    total_boxes = 0
    scanned_boxes = 0
    grn_type_prefix = (grn.grn_type or '').upper()
    for it in grn.items.order_by(GrnItem.sr_no).all():
        n_boxes = max(int(it.no_of_boxes or 0), 0)
        po_item = it.po_item_id or 0
        boxes = []
        for b in range(1, n_boxes + 1):
            # Expected QR code for this box â€” same format as labels.html
            expected_qr = f'{grn_type_prefix}{it.id}-{po_item}-{b}'
            key = (it.id, b)
            if key in scanned_map:
                s = scanned_map[key]
                boxes.append({
                    'box_no': b, 'scanned': True,
                    'qr_code': s.qr_code or expected_qr,
                    'expected_qr': expected_qr,
                    'scan_id': s.id,
                    'status': s.status, 'status_color': s.status_color,
                    'scanned_at': s.scanned_at.strftime('%d-%m-%Y %H:%M:%S') if s.scanned_at else '',
                    'scanned_by': s.scanned_by_name or '',
                })
                scanned_boxes += 1
            else:
                boxes.append({
                    'box_no': b, 'scanned': False,
                    'qr_code': expected_qr,
                    'expected_qr': expected_qr,
                })
            total_boxes += 1
        items_out.append({
            'id': it.id, 'sr_no': it.sr_no,
            'item_name': it.item_name or '',
            'item_code': it.item_code or '',
            'batch_no':  it.batch_no or '',
            'uom':       it.uom or 'KG',
            'per_box_qty': float(it.per_box_qty or 0),
            'no_of_boxes': n_boxes,
            'received_qty': float(it.received_qty or 0),
            'boxes': boxes,
            'scanned_count': sum(1 for b in boxes if b['scanned']),
        })

    # All scans for this GRN (newest first) â€” no limit, show all in right column
    recent = [s.to_dict() for s in scans]

    # Proper status-wise counts (from full list, not limited)
    quarantine_count = sum(1 for s in scans if s.status == SCAN_STATUS_QUARANTINE)
    stocked_in_count = sum(1 for s in scans if s.status == SCAN_STATUS_STOCKED_IN)

    return jsonify(
        success=True,
        grn={
            'id': grn.id,
            'grn_number': grn.grn_number,
            'grn_number_short': grn.grn_number_short or grn.grn_number,
            'grn_type': grn.grn_type or '',
            'supplier_name': grn.supplier_name or '',
            'grn_date': grn.grn_date.strftime('%d-%m-%Y') if grn.grn_date else '',
            'status': grn.status,
        },
        items=items_out,
        recent_scans=recent,
        total_boxes=total_boxes,
        scanned_boxes=scanned_boxes,
        pending_boxes=total_boxes - scanned_boxes,
        quarantine_count=quarantine_count,
        stocked_in_count=stocked_in_count,
    )


@grn_bp.route('/api/scan', methods=['POST'])
@login_required
def api_scan():
    """Process a single QR code scan.
    Optional `grn_id` in payload restricts the scan to that GRN only.
    """
    if not _can('create'):
        return jsonify(success=False, error='Permission denied'), 403

    data = request.get_json(silent=True) or {}
    qr_raw       = (data.get('qr_code') or '').strip().upper()
    source       = (data.get('source')  or 'manual').strip()
    expected_grn = data.get('grn_id')   # optional; if set, scanned QR must belong to this GRN

    if not qr_raw:
        return jsonify(success=False, error='QR code is empty'), 400

    # Parse
    m = _QR_RE.match(qr_raw)
    if not m:
        return jsonify(success=False,
                       error=f'Invalid QR format: "{qr_raw}". Expected like RM15-23-1.'), 400
    qr_type, grn_item_id_s, po_item_id_s, box_no_s = m.group(1), m.group(2), m.group(3), m.group(4)
    grn_item_id = int(grn_item_id_s)
    po_item_id  = int(po_item_id_s)
    box_no      = int(box_no_s)

    # Duplicate check
    existing = GrnScanLog.query.filter_by(qr_code=qr_raw, is_deleted=False).first()
    if existing:
        return jsonify(
            success=False,
            error=f'Already scanned at {existing.scanned_at.strftime("%d-%m-%Y %H:%M:%S")} by {existing.scanned_by_name or "unknown"}',
            duplicate=True,
            scan=existing.to_dict(),
        ), 400

    # Look up GRN item
    it = GrnItem.query.filter_by(id=grn_item_id).first()
    if not it:
        return jsonify(success=False, error=f'GRN item #{grn_item_id} not found'), 404

    grn = GrnMaster.query.get(it.grn_id)
    if not grn or grn.is_deleted:
        return jsonify(success=False, error='Parent GRN not found or deleted'), 404
    if grn.status != GRN_STATUS_COMPLETED:
        return jsonify(success=False,
                       error=f'GRN {grn.grn_number} is "{grn.status}" â€” scan only works on Completed GRNs'), 400

    # If scoped to a particular GRN, enforce it
    if expected_grn:
        try:
            if int(expected_grn) != grn.id:
                return jsonify(success=False,
                               error=f'This QR belongs to GRN {grn.grn_number}, not the one you are scanning into.'), 400
        except (ValueError, TypeError):
            pass

    # Per-box qty
    per_box = float(it.per_box_qty or 0)
    if per_box <= 0:
        n_boxes = max(int(it.no_of_boxes or 1), 1)
        per_box = float(it.received_qty or 0) / n_boxes
    rate    = float(it.rate or 0)
    amount  = round(per_box * rate, 2)

    # Decide status by GRN type
    is_rm = (grn.grn_type or '').upper() == 'RM'
    status = SCAN_STATUS_QUARANTINE if is_rm else SCAN_STATUS_STOCKED_IN

    try:
        slog = GrnScanLog(
            qr_code      = qr_raw,
            grn_type     = qr_type,
            grn_item_id  = it.id,
            po_item_id   = po_item_id or None,
            box_no       = box_no,
            grn_id       = grn.id,
            grn_number   = grn.grn_number or '',
            material_id  = it.material_id,
            item_code    = it.item_code or '',
            item_name    = it.item_name or '',
            batch_no     = it.batch_no or '',
            mfg_date     = it.mfg_date,
            expiry_date  = it.expiry_date,
            uom          = it.uom or 'KG',
            qty          = per_box,
            rate         = rate,
            amount       = amount,
            location_id  = it.storage_location_id,
            location_name= it.storage_location_name or '',
            status       = status,
            scanned_by_id   = getattr(current_user, 'id', None),
            scanned_by_name = _username(),
            scan_source     = source,
        )
        db.session.add(slog)
        db.session.flush()

        # For non-RM â†’ also push into stock_ledger + batch_stock
        if not is_rm:
            ledger = GrnStockLedger(
                txn_date     = datetime.utcnow(),
                txn_type     = 'GRN_IN',
                txn_ref_type = 'GRN_SCAN',
                txn_ref_id   = slog.id,
                txn_ref_no   = f'{grn.grn_number} Â· Box {box_no} Â· {qr_raw}',
                material_id  = it.material_id,
                item_code    = it.item_code or '',
                item_name    = it.item_name or '',
                batch_no     = it.batch_no or '',
                location_id  = it.storage_location_id,
                location_name= it.storage_location_name or '',
                qty_in       = per_box,
                qty_out      = 0,
                uom          = it.uom or 'KG',
                rate         = rate,
                amount       = amount,
                remarks      = f'Box scan {qr_raw}',
                actor_name   = _username(),
            )
            db.session.add(ledger)
            db.session.flush()
            slog.stock_ledger_id = ledger.id

            bs = (GrnBatchStock.query
                  .filter_by(material_id = it.material_id,
                             batch_no    = it.batch_no or '',
                             location_id = it.storage_location_id)
                  .first())
            if not bs:
                bs = GrnBatchStock(
                    material_id   = it.material_id,
                    item_code     = it.item_code or '',
                    item_name     = it.item_name or '',
                    batch_no      = it.batch_no or '',
                    location_id   = it.storage_location_id,
                    location_name = it.storage_location_name or '',
                    mfg_date      = it.mfg_date,
                    expiry_date   = it.expiry_date,
                    qty_on_hand   = 0,
                    qty_available = 0,
                    uom           = it.uom or 'KG',
                    avg_rate      = rate,
                )
                db.session.add(bs)
                db.session.flush()
            old_qty  = float(bs.qty_on_hand or 0)
            old_rate = float(bs.avg_rate or 0)
            new_qty  = old_qty + per_box
            new_avg  = ((old_qty * old_rate) + (per_box * rate)) / new_qty if new_qty > 0 else rate
            bs.qty_on_hand    = new_qty
            bs.qty_available  = new_qty - float(bs.qty_reserved or 0)
            bs.avg_rate       = new_avg
            bs.last_inward_at = datetime.utcnow()
            slog.batch_stock_id = bs.id

        db.session.commit()
        return jsonify(success=True, scan=slog.to_dict())

    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•


# â•â•â•â•â• Print Labels â•â•â•â•â•
@grn_bp.route('/<int:grn_id>/labels')
@login_required
def labels_grn(grn_id):
    """Print batch labels for items in a GRN â€” 2-up A4 layout."""
    if not _can('view'):
        abort(403)
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        abort(404)
    items = grn.items.order_by(GrnItem.sr_no).all()
    items_data = [it.to_dict() for it in items]  # serialize once for JS
    return render_template('grn/labels.html',
                           active_page='grn',
                           grn=grn,
                           items=items,
                           items_data=items_data)



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# CHECK LIST MATERIAL FORM â€” Item-wise printable checklist (no data stored)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@grn_bp.route('/<int:grn_id>/checklist')
@login_required
def checklist_grn(grn_id):
    """Print item-wise Check List Material Form for an RM GRN.

    Nothing is stored â€” the form is rendered from existing GRN/item data
    and is meant to be printed on paper and physically ticked.
    """
    if not _can('view'):
        abort(403)
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        abort(404)
    items = grn.items.order_by(GrnItem.sr_no).all()
    return render_template('grn/checklist.html',
                           active_page='grn',
                           grn=grn,
                           items=items)


# â•â•â•â•â• Excel Export â•â•â•â•â•
@grn_bp.route('/export')
@login_required
def export_excel():
    """Export GRN listing to Excel."""
    from flask import send_file
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify(error='openpyxl not installed. Run: pip install openpyxl'), 500

    wb = Workbook()
    ws = wb.active

    # Common header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1E40AF')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_header(row_cells):
        for c in row_cells:
            c.font = header_font; c.fill = header_fill
            c.alignment = center; c.border = border

    # Branch by `kind` â€” listing (default) / batch_stock / stock_ledger
    kind = (request.args.get('kind', '') or '').strip().lower()

    if kind == 'batch_stock':
        ws.title = 'Current Stock'
        headers = ['#', 'Item Code', 'Item Name', 'Location',
                   'On Hand', 'Reserved', 'Available', 'UOM',
                   'Avg Rate', 'Last Inward']
        ws.append(headers)
        _style_header(ws[1])

        qs = GrnBatchStock.query.filter(GrnBatchStock.qty_on_hand > 0)
        q          = (request.args.get('q','') or '').strip()
        loc_id     = (request.args.get('location_id','') or '').strip()
        if q:
            like = f'%{q}%'
            qs = qs.filter(or_(GrnBatchStock.item_name.ilike(like),
                               GrnBatchStock.item_code.ilike(like)))
        if loc_id:
            try: qs = qs.filter(GrnBatchStock.location_id == int(loc_id))
            except ValueError: pass

        for i, r in enumerate(qs.order_by(desc(GrnBatchStock.last_inward_at)).all(), 1):
            ws.append([
                i, r.item_code or '', r.item_name or '',
                r.location_name or 'â€”',
                float(r.qty_on_hand or 0), float(r.qty_reserved or 0),
                float(r.qty_available or 0), r.uom or 'KG',
                float(r.avg_rate or 0),
                r.last_inward_at.strftime('%d-%m-%Y %H:%M') if r.last_inward_at else '',
            ])
        widths = [5, 16, 36, 18, 12, 12, 12, 8, 12, 18]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.border = border
        ws.freeze_panes = 'A2'
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'Current_Stock_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if kind == 'stock_by_grn':
        ws.title = 'Stock by GRN'
        headers = ['#', 'GRN No', 'Type', 'GRN Date', 'Supplier',
                   'Item Code', 'Item Name', 'Location',
                   'Boxes', 'Total Qty', 'UOM',
                   'Avg Rate', 'Amount', 'Last Inward', 'GRN Status']
        ws.append(headers)
        _style_header(ws[1])

        q          = (request.args.get('q','') or '').strip()
        grn_type   = (request.args.get('grn_type','') or '').strip().upper()
        loc_id     = (request.args.get('location_id','') or '').strip()
        df         = _parse_date(request.args.get('date_from'))
        dt         = _parse_date(request.args.get('date_to'))

        base = (db.session.query(
                    GrnScanLog.grn_id, GrnScanLog.grn_number,
                    GrnScanLog.material_id, GrnScanLog.item_code, GrnScanLog.item_name,
                    GrnScanLog.location_id, GrnScanLog.location_name, GrnScanLog.uom,
                    func.count(GrnScanLog.id).label('box_count'),
                    func.sum(GrnScanLog.qty).label('total_qty'),
                    func.sum(GrnScanLog.amount).label('total_amount'),
                    func.avg(GrnScanLog.rate).label('avg_rate'),
                    func.max(GrnScanLog.scanned_at).label('last_at'),
                )
                .filter(GrnScanLog.is_deleted == False,
                        GrnScanLog.status == SCAN_STATUS_STOCKED_IN))
        if loc_id:
            try: base = base.filter(GrnScanLog.location_id == int(loc_id))
            except ValueError: pass
        base = base.group_by(GrnScanLog.grn_id, GrnScanLog.grn_number,
                             GrnScanLog.material_id, GrnScanLog.item_code,
                             GrnScanLog.item_name, GrnScanLog.location_id,
                             GrnScanLog.location_name, GrnScanLog.uom)
        agg_rows = base.all()

        gids = list({r.grn_id for r in agg_rows if r.grn_id})
        gmap = {}
        if gids:
            for g in GrnMaster.query.filter(GrnMaster.id.in_(gids)).all():
                gmap[g.id] = g

        i = 0
        # Sort newest-first by last scan
        agg_rows = sorted(agg_rows, key=lambda r: r.last_at or datetime.min, reverse=True)
        for r in agg_rows:
            g = gmap.get(r.grn_id)
            if grn_type and (not g or (g.grn_type or '').upper() != grn_type): continue
            if df and (not g or not g.grn_date or g.grn_date < df): continue
            if dt and (not g or not g.grn_date or g.grn_date > dt): continue
            if q:
                ql = q.lower()
                hay = ' '.join([(r.grn_number or ''), (r.item_name or ''),
                                (r.item_code or ''),
                                (g.supplier_name if g else '') or '']).lower()
                if ql not in hay: continue
            i += 1
            ws.append([
                i, r.grn_number or '',
                g.grn_type if g else '',
                g.grn_date.strftime('%d-%m-%Y') if g and g.grn_date else '',
                g.supplier_name if g else '',
                r.item_code or '', r.item_name or '',
                r.location_name or 'â€”',
                int(r.box_count or 0), float(r.total_qty or 0), r.uom or 'KG',
                float(r.avg_rate or 0), float(r.total_amount or 0),
                r.last_at.strftime('%d-%m-%Y %H:%M') if r.last_at else '',
                g.status if g else '',
            ])
        widths = [5, 22, 7, 12, 28, 16, 32, 18, 8, 14, 8, 12, 14, 18, 14]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.border = border
        ws.freeze_panes = 'A2'
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'Stock_by_GRN_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if kind == 'stock_ledger':
        ws.title = 'Stock Ledger'
        headers = ['#', 'Txn Date', 'Type', 'Ref', 'Item Code', 'Item Name',
                   'Location', 'Qty In', 'Qty Out', 'UOM',
                   'Rate', 'Amount', 'By']
        ws.append(headers)
        _style_header(ws[1])

        qs = GrnStockLedger.query
        q          = (request.args.get('q','') or '').strip()
        txn_type   = (request.args.get('txn_type','') or '').strip()
        df         = _parse_date(request.args.get('date_from'))
        dt         = _parse_date(request.args.get('date_to'))
        if q:
            like = f'%{q}%'
            qs = qs.filter(or_(GrnStockLedger.item_name.ilike(like),
                               GrnStockLedger.item_code.ilike(like),
                               GrnStockLedger.txn_ref_no.ilike(like)))
        if txn_type:
            qs = qs.filter(GrnStockLedger.txn_type == txn_type)
        if df:
            qs = qs.filter(GrnStockLedger.txn_date >= datetime.combine(df, time.min))
        if dt:
            qs = qs.filter(GrnStockLedger.txn_date <= datetime.combine(dt, time.max))

        for i, r in enumerate(qs.order_by(desc(GrnStockLedger.txn_date)).all(), 1):
            ws.append([
                i,
                r.txn_date.strftime('%d-%m-%Y %H:%M') if r.txn_date else '',
                r.txn_type or '', r.txn_ref_no or '',
                r.item_code or '', r.item_name or '',
                r.location_name or 'â€”',
                float(r.qty_in or 0), float(r.qty_out or 0),
                r.uom or 'KG',
                float(r.rate or 0), float(r.amount or 0),
                r.actor_name or '',
            ])
        widths = [5, 18, 12, 26, 16, 32, 18, 12, 12, 8, 12, 14, 16]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.border = border
        ws.freeze_panes = 'A2'
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'Stock_Ledger_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Default: GRN listing export (existing behavior)
    ws.title = 'GRN Listing'
    headers = ['#', 'GRN No', 'Date', 'Type', 'Supplier', 'PO No',
               'Invoice No', 'Invoice Date', 'Total Boxes',
               'Total Received', 'Total Amount', 'Status', 'Created By', 'Created At']
    ws.append(headers)
    _style_header(ws[1])
    qs = GrnMaster.query.filter_by(is_deleted=False)
    gtype = (request.args.get('grn_type','') or '').upper().strip()
    status = (request.args.get('status','') or '').strip()
    q = (request.args.get('q','') or '').strip()
    date_from = _parse_date(request.args.get('date_from'))
    date_to   = _parse_date(request.args.get('date_to'))
    if gtype: qs = qs.filter(GrnMaster.grn_type == gtype)
    if status: qs = qs.filter(GrnMaster.status == status)
    if date_from: qs = qs.filter(GrnMaster.grn_date >= date_from)
    if date_to:   qs = qs.filter(GrnMaster.grn_date <= date_to)
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            GrnMaster.grn_number.ilike(like),
            GrnMaster.po_number.ilike(like),
            GrnMaster.supplier_name.ilike(like),
            GrnMaster.invoice_no.ilike(like),
        ))
    for i, r in enumerate(qs.order_by(desc(GrnMaster.id)).all(), 1):
        ws.append([
            i, r.grn_number or '',
            r.grn_date.strftime('%d-%m-%Y') if r.grn_date else '',
            r.grn_type or '', r.supplier_name or '', r.po_number or '',
            r.invoice_no or '',
            r.invoice_date.strftime('%d-%m-%Y') if r.invoice_date else '',
            int(r.total_box_qty or 0),
            float(r.total_received_qty or 0),
            float(r.total_amount or 0),
            r.status or '', r.created_by_name or '',
            r.created_at.strftime('%d-%m-%Y %H:%M') if r.created_at else '',
        ])
    widths = [5, 22, 12, 7, 30, 22, 16, 12, 12, 16, 14, 16, 14, 18]

    # Apply column widths + borders
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'GRN_Listing_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@grn_bp.route('/api/<int:grn_id>/items')
@login_required
def api_grn_items(grn_id):
    """Returns items belonging to a saved GRN â€” used by listing's row-expand."""
    if not _can('view'):
        return jsonify(error='Permission denied'), 403
    grn = GrnMaster.query.get_or_404(grn_id)
    if grn.is_deleted:
        return jsonify(error='GRN deleted'), 404
    items = grn.items.order_by(GrnItem.sr_no).all()
    return jsonify(
        grn_id=grn.id,
        grn_number=grn.grn_number,
        grn_type=grn.grn_type or '',
        items=[it.to_dict() for it in items],
    )


@grn_bp.route('/api/material/<int:mat_id>')
@login_required
def api_material(mat_id):
    m = Material.query.get_or_404(mat_id)
    return jsonify(
        id=m.id,
        item_name=m.material_name or '',
        item_code=m.code or '',
        category=m.category or '',
        hsn_code=m.hsn_code or '',
        uom=m.uom or 'KG',
        gst_rate=float(m.gst_rate or 0),
        last_purchase_rate=float(m.last_purchase_rate or 0),
    )


@grn_bp.route('/api/search-items')
@login_required
def api_search_items():
    """Search items for the per-row picker.
       - If po_id given: returns ONLY that PO's pending items
       - Else (direct receive / NA flow): returns materials filtered by
         the GRN type (RM GRN â†’ RM materials, PM â†’ PM, etc.)
    """
    q        = (request.args.get('q', '') or '').strip()
    po_id    = (request.args.get('po_id', '') or '').strip()
    grn_type = (request.args.get('grn_type', '') or '').strip().upper()

    if po_id:
        try:
            po = PurchaseOrder.query.get(int(po_id))
        except (ValueError, TypeError):
            return jsonify(results=[])
        if not po or po.is_deleted:
            return jsonify(results=[])
        items = []
        for it in po.items.order_by(PurchaseOrderItem.sr_no).all():
            ordered = float(it.quantity or 0)
            already = float(getattr(it, 'received_qty', 0) or 0)
            pending = max(ordered - already, 0)
            if pending <= 0:
                continue   # skip fully-received items
            name = it.item_name or ''
            code = it.item_code or ''
            if q:
                ql = q.lower()
                if ql not in name.lower() and ql not in code.lower():
                    continue
            items.append({
                'id':   it.id,    # po_item_id (used as select2 value)
                'text': f'{name} â€” {code} (pending {pending:.3f} {it.uom or "KG"})',
                'po_item_id':  it.id,
                'po_id':       po.id,
                'po_number':   po.po_number,
                'po_date':     po.po_date.strftime('%Y-%m-%d') if po.po_date else '',
                'material_id': it.material_id,
                'item_code':   code,
                'item_name':   name,
                'category':    it.category or '',
                'hsn_code':    it.hsn_code or '',
                'uom':         it.uom or 'KG',
                'ordered_qty': ordered,
                'already_received_qty': already,
                'remaining_qty': pending,
                'rate':    float(it.rate or 0),
                'gst_pct': float(it.gst_pct or 0),
            })
        return jsonify(results=items)

    # Direct GRN â€” filter materials by GRN type abbreviation (RM/PM/COR/SLV/FG)
    qs = Material.query.filter_by(is_deleted=False)
    if grn_type:
        # Match the GRN type to material_types.abbreviation (case-insensitive).
        # If no matching type exists, return zero results rather than ALL.
        qs = (qs.join(MaterialType, Material.material_type_id == MaterialType.id)
                .filter(db.func.upper(MaterialType.abbreviation) == grn_type)
                .filter(MaterialType.is_deleted == False))
    if q:
        like = f'%{q}%'
        qs = qs.filter(or_(
            Material.material_name.ilike(like),
            Material.code.ilike(like),
        ))
    rows = qs.order_by(Material.material_name).limit(50).all()
    return jsonify(results=[{
        'id': f'mat-{m.id}',
        'text': f'{m.material_name} â€” {m.code}',
        'po_item_id':  None,
        'po_id':       None,
        'po_number':   '',
        'material_id': m.id,
        'item_code':   m.code or '',
        'item_name':   m.material_name or '',
        'category':    m.category or '',
        'hsn_code':    m.hsn_code or '',
        'uom':         m.uom or 'KG',
        'ordered_qty': 0,
        'already_received_qty': 0,
        'remaining_qty': 0,
        'rate':    float(m.last_purchase_rate or 0),
        'gst_pct': float(m.gst_rate or 0),
    } for m in rows])


