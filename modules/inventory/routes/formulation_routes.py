"""
formulation_routes.py â€” Formulation Master Module
Blueprint: formulation at /formulation

Features
========
1. List page (search, filter by brand, view linked/source badge)
2. Create New formulation manually with ingredient rows
       - Enter by % concentration  OR  Qty by batch size
3. Link to Existing formulation  (shared ingredients, own batch / brand)
4. Edit / soft-delete
5. Excel Import (XLSX) â€” multi-sheet, one formulation per sheet
       Header row trigger: a row whose Col-A is "Sr. No." or
       contains a serial integer â‰¥ 1 in a body row. We scan rows
       from the top, find the header, then read body rows where
       Col A is non-empty. Body fields:
            Col A  â†’ sr_no
            Col B  â†’ ingredient_name
            Col C  â†’ supplier_name
            Col D  â†’ percentage  (% w/w)
            Col E  â†’ qty_kg
       Stop on first empty Col A *after* we've started reading.
       Total / blank rows are auto-skipped.
6. Excel Export â€” single formulation OR all formulations, formatted
   the same way the user already uses (header block + table).
"""
from datetime import datetime
import io
import re

from flask import (Blueprint, render_template, request, jsonify, abort,
                   send_file, Response)
from flask_login import login_required, current_user

from models import (db, Formulation, FormulationIngredient,
                    Material, ClientBrand)
from core.permissions import get_perm

formulation_bp = Blueprint('formulation', __name__, url_prefix='/formulation')

# â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _cu():   return getattr(current_user, 'username', '') or ''
def _role(): return getattr(current_user, 'role', '') or ''

def _can(action):
    if _role() in ('admin', 'manager'):
        return True
    p = get_perm('formulation')
    return bool(p and getattr(p, f'can_{action}', False))


def _num(val, default=0.0):
    """Forgiving numeric parser â€” accepts '12', '12.5', '12%', '  '."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '').replace('%', '').replace('\u00a0', '')
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


# â”€â”€â”€ Pages â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/')
@formulation_bp.route('')
@login_required
def index():
    if not _can('view'):
        abort(403)

    # Brand list â€” try ClientBrand master first, fall back to distinct values
    brands = []
    try:
        brands = [b.brand_name for b in
                  ClientBrand.query.filter_by(is_active=True)
                                   .order_by(ClientBrand.brand_name).all()]
    except Exception:
        pass
    if not brands:
        rows = (db.session.query(Formulation.brand)
                .filter(Formulation.is_deleted == False,
                        Formulation.brand != '',
                        Formulation.brand != None)
                .distinct().all())
        brands = sorted({r[0] for r in rows if r[0]})

    return render_template(
        'formulation/index.html',
        active_page='formulation',
        role=_role(),
        brands=brands,
        can_add    = _can('add'),
        can_edit   = _can('edit'),
        can_delete = _can('delete'),
        can_export = _can('export'),
        user_name=getattr(current_user, 'full_name', '') or _cu(),
    )


# â”€â”€â”€ API: List â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/list')
@login_required
def api_list():
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    status = (request.args.get('status') or 'active').strip().lower()
    if status == 'deleted':
        q = Formulation.query.filter_by(is_deleted=True)
    elif status == 'all':
        q = Formulation.query
    else:  # active (default)
        q = Formulation.query.filter_by(is_deleted=False)

    search = (request.args.get('search') or '').strip()
    brand  = (request.args.get('brand')  or '').strip()
    link_filter = (request.args.get('link') or '').strip().lower()   # '', 'source', 'linked'

    if search:
        like = f'%{search}%'
        q = q.filter(
            (Formulation.name.ilike(like)) |
            (Formulation.product_code.ilike(like)) |
            (Formulation.brand.ilike(like))
        )
    if brand:
        q = q.filter(Formulation.brand == brand)
    if link_filter == 'source':
        q = q.filter(Formulation.source_id == None)
    elif link_filter == 'linked':
        q = q.filter(Formulation.source_id != None)

    q = q.order_by(Formulation.created_at.desc())
    items = []
    for f in q.all():
        d = f.to_dict()
        d['deleted_at'] = f.deleted_at.strftime('%d-%m-%Y %H:%M') if f.deleted_at else ''
        items.append(d)
    return jsonify({'status': 'ok', 'items': items, 'view_status': status})


# â”€â”€â”€ API: Get One (with resolved ingredients) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/<int:fid>')
@login_required
def api_get(fid):
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    f = Formulation.query.get_or_404(fid)
    if f.is_deleted:
        return jsonify({'status': 'error', 'message': 'Deleted'}), 404
    return jsonify({'status': 'ok', 'item': f.to_dict(include_ingredients=True)})


# â”€â”€â”€ API: Sources list â€” for "Link to Existing" dropdown â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/sources')
@login_required
def api_sources():
    """Only formulations that can be a source â€” non-deleted, non-linked."""
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    rows = (Formulation.query
            .filter(Formulation.is_deleted == False,
                    Formulation.source_id == None)
            .order_by(Formulation.name).all())
    return jsonify({
        'status': 'ok',
        'items': [{'id': r.id,
                   'name': r.name,
                   'product_code': r.product_code or '',
                   'batch_size': float(r.batch_size or 0),
                   'brand': r.brand or '',
                   'ingredient_count': r.ingredient_count()} for r in rows]
    })


# â”€â”€â”€ API: Create (manual) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/create', methods=['POST'])
@login_required
def api_create():
    if not _can('add'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d = request.get_json(silent=True) or {}

    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'Formulation name is required'}), 400

    # Uniqueness on name (case-insensitive)
    dup = (Formulation.query
           .filter(db.func.lower(Formulation.name) == name.lower(),
                   Formulation.is_deleted == False)
           .first())
    if dup:
        return jsonify({'status': 'error',
                        'message': f'Formulation "{name}" already exists'}), 400

    ings = d.get('ingredients') or []

    f = Formulation(
        name                  = name,
        product_code          = (d.get('product_code') or '').strip(),
        batch_size            = _num(d.get('batch_size'), 0),
        batch_uom             = (d.get('batch_uom') or 'KG').strip(),
        brand                 = (d.get('brand') or '').strip(),
        manufacturing_process = (d.get('manufacturing_process') or '').strip(),
        specifications        = (d.get('specifications') or '').strip(),
        created_by            = _cu(),
        updated_by            = _cu(),
    )
    db.session.add(f)
    db.session.flush()           # need id

    sr = 0
    for raw in ings:
        nm = (raw.get('ingredient_name') or '').strip()
        if not nm:
            continue
        sr += 1
        pct = _num(raw.get('percentage'), 0)
        rm  = _num(raw.get('rm_rate_per_kg'), 0)
        # auto-derive bulk_rate if not provided: % Ã— rm_rate
        bulk = _num(raw.get('bulk_rate_per_kg'), pct * rm)
        db.session.add(FormulationIngredient(
            formulation_id   = f.id,
            sr_no            = raw.get('sr_no') or sr,
            ingredient_name  = nm,
            supplier_name    = (raw.get('supplier_name') or '').strip(),
            percentage       = pct,
            qty_kg           = _num(raw.get('qty_kg'), 0),
            uom              = (raw.get('uom') or 'KG').strip(),
            is_additional    = bool(raw.get('is_additional')),
            rm_rate_per_kg   = rm,
            bulk_rate_per_kg = bulk,
            material_id      = raw.get('material_id') or None,
        ))

    db.session.commit()
    return jsonify({'status': 'ok', 'id': f.id, 'item': f.to_dict()})


# â”€â”€â”€ API: Link to existing â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/link', methods=['POST'])
@login_required
def api_link():
    if not _can('add'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d = request.get_json(silent=True) or {}

    name      = (d.get('name') or '').strip()
    source_id = d.get('source_id')

    if not name:
        return jsonify({'status': 'error', 'message': 'Formulation name is required'}), 400
    if not source_id:
        return jsonify({'status': 'error', 'message': 'Source formulation is required'}), 400

    src = Formulation.query.get(source_id)
    if not src or src.is_deleted:
        return jsonify({'status': 'error', 'message': 'Source formulation not found'}), 400
    if src.source_id is not None:
        return jsonify({'status': 'error',
                        'message': 'Selected source is itself a linked formulation. '
                                   'Pick the original source.'}), 400

    # Uniqueness
    dup = (Formulation.query
           .filter(db.func.lower(Formulation.name) == name.lower(),
                   Formulation.is_deleted == False)
           .first())
    if dup:
        return jsonify({'status': 'error',
                        'message': f'Formulation "{name}" already exists'}), 400

    f = Formulation(
        name                  = name,
        product_code          = (d.get('product_code') or '').strip(),
        batch_size            = _num(d.get('batch_size'), float(src.batch_size or 0)),
        batch_uom             = (d.get('batch_uom') or src.batch_uom or 'KG').strip(),
        brand                 = (d.get('brand') or src.brand or '').strip(),
        manufacturing_process = (d.get('manufacturing_process') or src.manufacturing_process or '').strip(),
        source_id             = src.id,
        created_by            = _cu(),
        updated_by            = _cu(),
    )
    db.session.add(f)
    db.session.commit()
    return jsonify({'status': 'ok', 'id': f.id, 'item': f.to_dict()})


# â”€â”€â”€ API: Update â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/<int:fid>', methods=['PUT'])
@login_required
def api_update(fid):
    if not _can('edit'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    f = Formulation.query.get_or_404(fid)
    if f.is_deleted:
        return jsonify({'status': 'error', 'message': 'Deleted'}), 404

    d = request.get_json(silent=True) or {}

    if 'name' in d:
        new_name = (d['name'] or '').strip()
        if not new_name:
            return jsonify({'status': 'error', 'message': 'Name cannot be empty'}), 400
        # Uniqueness â€” allow same id
        dup = (Formulation.query
               .filter(db.func.lower(Formulation.name) == new_name.lower(),
                       Formulation.is_deleted == False,
                       Formulation.id != fid)
               .first())
        if dup:
            return jsonify({'status': 'error',
                            'message': f'Another formulation named "{new_name}" exists'}), 400
        f.name = new_name

    if 'product_code'          in d: f.product_code          = (d['product_code'] or '').strip()
    if 'batch_size'            in d: f.batch_size            = _num(d['batch_size'], 0)
    if 'batch_uom'             in d: f.batch_uom             = (d['batch_uom'] or 'KG').strip()
    if 'brand'                 in d: f.brand                 = (d['brand'] or '').strip()
    if 'manufacturing_process' in d: f.manufacturing_process = (d['manufacturing_process'] or '').strip()
    if 'specifications'        in d: f.specifications        = (d['specifications'] or '').strip()

    # Only update ingredients on rows that own them (source rows)
    if 'ingredients' in d and not f.is_linked:
        # Wipe + re-add (simplest, fewest edge cases)
        FormulationIngredient.query.filter_by(formulation_id=f.id).delete()
        sr = 0
        for raw in (d['ingredients'] or []):
            nm = (raw.get('ingredient_name') or '').strip()
            if not nm:
                continue
            sr += 1
            pct = _num(raw.get('percentage'), 0)
            rm  = _num(raw.get('rm_rate_per_kg'), 0)
            bulk = _num(raw.get('bulk_rate_per_kg'), pct * rm)
            db.session.add(FormulationIngredient(
                formulation_id   = f.id,
                sr_no            = raw.get('sr_no') or sr,
                ingredient_name  = nm,
                supplier_name    = (raw.get('supplier_name') or '').strip(),
                percentage       = pct,
                qty_kg           = _num(raw.get('qty_kg'), 0),
                uom              = (raw.get('uom') or 'KG').strip(),
                is_additional    = bool(raw.get('is_additional')),
                rm_rate_per_kg   = rm,
                bulk_rate_per_kg = bulk,
                material_id      = raw.get('material_id') or None,
            ))

    f.updated_by = _cu()
    db.session.commit()
    return jsonify({'status': 'ok', 'item': f.to_dict(include_ingredients=True)})


# â”€â”€â”€ API: Delete (soft) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/<int:fid>', methods=['DELETE'])
@login_required
def api_delete(fid):
    if not _can('delete'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    f = Formulation.query.get_or_404(fid)
    if f.is_deleted:
        return jsonify({'status': 'ok', 'message': 'Already deleted'})

    # Block if any active linked formulation depends on this
    if f.source_id is None:
        live_links = (Formulation.query
                      .filter(Formulation.source_id == f.id,
                              Formulation.is_deleted == False)
                      .count())
        if live_links:
            return jsonify({'status': 'error',
                            'message': f'Cannot delete â€” {live_links} linked '
                                       f'formulation(s) depend on this. Delete '
                                       f'or unlink them first.'}), 400

    f.is_deleted = True
    f.deleted_at = datetime.utcnow()
    f.updated_by = _cu()
    db.session.commit()
    return jsonify({'status': 'ok'})


# â”€â”€â”€ API: Restore (un-delete a soft-deleted record) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/<int:fid>/restore', methods=['POST'])
@login_required
def api_restore(fid):
    if not _can('edit'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    f = Formulation.query.get_or_404(fid)
    if not f.is_deleted:
        return jsonify({'status': 'ok', 'message': 'Already active'})

    # If a record with the same name now exists active, block restore
    clash = (Formulation.query
             .filter(db.func.lower(Formulation.name) == (f.name or '').lower(),
                     Formulation.is_deleted == False,
                     Formulation.id != fid)
             .first())
    if clash:
        return jsonify({'status': 'error',
                        'message': f'Cannot restore â€” an active formulation named '
                                   f'"{f.name}" already exists. Rename or delete the active one first.'}), 400

    # If this row was linked to a now-deleted source, surface that as a warning
    warning = ''
    if f.source_id:
        src = Formulation.query.get(f.source_id)
        if not src or src.is_deleted:
            warning = 'Source formulation is missing or deleted â€” this row is restored as a standalone formulation (source link cleared).'
            f.source_id = None

    f.is_deleted = False
    f.deleted_at = None
    f.updated_by = _cu()
    db.session.commit()
    return jsonify({'status': 'ok', 'warning': warning})


# â”€â”€â”€ API: Permanent delete (hard) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@formulation_bp.route('/api/<int:fid>/permanent', methods=['DELETE'])
@login_required
def api_permanent_delete(fid):
    if not _can('delete'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    f = Formulation.query.get_or_404(fid)

    # Block hard-delete if there are live (non-deleted) linked formulations
    if f.source_id is None:
        live_links = (Formulation.query
                      .filter(Formulation.source_id == f.id,
                              Formulation.is_deleted == False)
                      .count())
        if live_links:
            return jsonify({'status': 'error',
                            'message': f'Cannot permanently delete â€” {live_links} active '
                                       f'linked formulation(s) still reference this source. '
                                       f'Delete or unlink them first.'}), 400

    # For soft-deleted links pointing here, clear the FK (FK is ON DELETE SET NULL,
    # but be explicit so we don't rely on SQL cascade behaviour).
    db.session.execute(
        db.text("UPDATE formulations SET source_id = NULL WHERE source_id = :sid"),
        {'sid': f.id}
    )

    db.session.delete(f)        # cascades to formulation_ingredients via FK
    db.session.commit()
    return jsonify({'status': 'ok'})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  EXCEL IMPORT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _is_header_row(row):
    """True if the row's first 5 cells look like the column header
    (Sr No / Ingredient / Supplier / Qty% / Actual Qty)."""
    cells = [str(c or '').strip().lower() for c in row[:5]]
    if 'ingredient' in cells[1] and ('supplier' in cells[2] or 'name' in cells[2]):
        return True
    if 'sr' in cells[0] and ('ingredient' in cells[1] or 'name' in cells[1]):
        return True
    return False


def _find_data_start(ws):
    """Locate the header row.  Returns the *next* row (1-based, the first
    body row).  Falls back to 13 if not found (per user spec)."""
    for r in range(1, min(ws.max_row, 30) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        if _is_header_row(row):
            return r + 1
    return 13


def _extract_header_meta(ws):
    """Pull product name, product code, batch size, batch UOM from the
    top of the sheet â€” based on the template the user uses.
    Returns dict with keys: name, product_code, batch_size, batch_uom."""
    meta = {'name': '', 'product_code': '', 'batch_size': 0.0, 'batch_uom': 'KG'}
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            sv = str(v)
            low = sv.lower()
            # Product Name
            if 'product name' in low and not meta['name']:
                m = re.split(r'product\s*name\s*:?', sv, flags=re.I, maxsplit=1)
                if len(m) == 2 and m[1].strip():
                    meta['name'] = m[1].strip().rstrip(':').strip()
                else:
                    # look right
                    nxt = ws.cell(row=r, column=c+1).value
                    if nxt:
                        meta['name'] = str(nxt).strip()
            # Product Code
            elif 'product code' in low and not meta['product_code']:
                m = re.split(r'product\s*code\s*:?', sv, flags=re.I, maxsplit=1)
                if len(m) == 2 and m[1].strip():
                    meta['product_code'] = m[1].strip().rstrip(':').strip()
                else:
                    nxt = ws.cell(row=r, column=c+1).value
                    if nxt:
                        meta['product_code'] = str(nxt).strip()
            # Batch Size â€” number is usually a couple of cells right
            elif 'batch size' in low and meta['batch_size'] == 0:
                for cc in range(c + 1, c + 6):
                    nxt = ws.cell(row=r, column=cc).value
                    if isinstance(nxt, (int, float)) and nxt > 0:
                        meta['batch_size'] = float(nxt)
                        # Try UOM in next cell
                        uom = ws.cell(row=r, column=cc+1).value or ws.cell(row=r, column=cc+2).value
                        if uom and isinstance(uom, str) and uom.strip():
                            meta['batch_uom'] = uom.strip()
                        break
    return meta


def _parse_sheet(ws):
    """Parse a single worksheet â†’ (meta_dict, list of ingredient_dicts, process_text)."""
    meta = _extract_header_meta(ws)
    start = _find_data_start(ws)

    rows_out = []
    seen_any = False
    last_data_row = start
    for r in range(start, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        col_b = ws.cell(row=r, column=2).value

        # Stop condition â€” empty Col A after we've started reading
        if col_a is None or (isinstance(col_a, str) and not col_a.strip()):
            if seen_any:
                last_data_row = r
                break
            continue

        # Skip a "Total" row that may slip in
        if isinstance(col_b, str) and col_b.strip().lower() in ('total', 'totals'):
            continue

        nm = ws.cell(row=r, column=2).value
        if nm is None or (isinstance(nm, str) and not nm.strip()):
            continue

        seen_any = True
        last_data_row = r
        rows_out.append({
            'sr_no'          : int(col_a) if isinstance(col_a, (int, float))
                                          else _num(col_a, len(rows_out) + 1),
            'ingredient_name': str(nm).strip(),
            'supplier_name'  : (str(ws.cell(row=r, column=3).value).strip()
                                if ws.cell(row=r, column=3).value else ''),
            'percentage'     : _num(ws.cell(row=r, column=4).value, 0),
            'qty_kg'         : _num(ws.cell(row=r, column=5).value, 0),
        })

    process = _extract_process_text(ws, last_data_row + 1)
    specs   = _extract_specifications(ws, last_data_row + 1)
    return meta, rows_out, process, specs


# â”€â”€â”€ Auto-extract the QC Specifications table â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _extract_specifications(ws, start_row):
    """Find the 'Specs' / 'Specifications' section below the ingredient
    table and capture its rows.  Each row may have:
        Col A  â†’ Sr.No
        Col B  â†’ Spec name      (e.g. "Appearance", "pH")
        Col C/D â†’ Range          (the acceptable value or range)
        Col E/F â†’ Result         (filled by QC team â€” usually blank in template)
        Col G  â†’ Remarks        (also usually blank)
    Returns an HTML table string ready for CKEditor display, or ''
    if no specs section is detected.
    """
    found_header = False
    rows = []
    max_col = min((ws.max_column or 1) + 1, 14)

    for r in range(start_row, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, max_col)]
        row_text = ' '.join(str(c).strip() for c in cells if c not in (None, '')).strip()
        if not row_text:
            continue
        low = row_text.lower()

        # Section header detection
        if not found_header:
            # Match 'specs', 'specs.', 'specifications', 'specs. range result remarks'
            if re.search(r'\bspecs?\.?\s', low + ' ') or 'specification' in low:
                # Avoid matching the Manufacturing Process header itself
                if 'process' not in low:
                    found_header = True
            continue

        # Terminators
        if re.search(r'batch\s*incharge|material\s*dispenser|approved\s*by|after\s*approval', low, re.I):
            break

        # Skip "Manufacturing Process" if encountered AFTER specs (rare but possible)
        if re.search(r'manufactur\w*\s*process', low):
            break

        # Extract row data â€” Col A (sr_no), Col B (spec name), rest (range/result/remarks)
        col_a = cells[0]
        col_b = cells[1] if len(cells) > 1 else None

        if col_b is None or (isinstance(col_b, str) and not col_b.strip()):
            continue   # need a spec name to register a row

        spec_name = str(col_b).strip()
        # Collect everything else (col C onward) as the Range/Result/Remarks tuple
        # Index 0 = col A (sr_no), 1 = col B (name), so look at 2..end
        extras = [str(c).strip() if c is not None else '' for c in cells[2:]]
        # Strip trailing empties
        while extras and not extras[-1]:
            extras.pop()

        # Heuristic: the FIRST non-empty extra is "Range", second is "Result", third is "Remarks"
        non_empty = [e for e in extras if e]
        range_val  = non_empty[0] if len(non_empty) > 0 else ''
        result_val = non_empty[1] if len(non_empty) > 1 else ''
        remark_val = non_empty[2] if len(non_empty) > 2 else ''

        sr = int(col_a) if isinstance(col_a, (int, float)) else len(rows) + 1
        rows.append((sr, spec_name, range_val, result_val, remark_val))

    if not rows:
        return ''

    # Return plain-text format (one per line) â€” easier to display & edit in
    # the per-sheet textarea.  On commit we convert this back to a clean
    # HTML table.  Existing HTML data already in DB still renders fine.
    lines = []
    for sr, nm, rg, rs, rm in rows:
        bits = [f'{sr}. {nm}: {rg}']
        if rs: bits.append(f'Result: {rs}')
        if rm: bits.append(f'Remarks: {rm}')
        lines.append(' | '.join(bits))
    return '\n'.join(lines)


def _specs_text_to_html(text):
    """Convert plain-text specs into a clean HTML table for storage.
    Lines may be either:
        "1. Apperance: Opaque Viscous Liquid"
        "Apperance: Opaque Viscous Liquid"
        "Apperance: Opaque Viscous Liquid | Result: ok | Remarks: -"
    """
    if not text or not text.strip():
        return ''
    if '<' in text:
        return text                            # already HTML â€” pass through
    import html as _html
    parts = ['<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:100%; font-size:13px;">']
    parts.append('<thead style="background:#1e3a5f; color:#fff;"><tr>'
                 '<th style="width:40px;">#</th><th>Specs</th><th>Range</th>'
                 '<th style="width:120px;">Result</th><th style="width:120px;">Remarks</th>'
                 '</tr></thead><tbody>')
    sr = 0
    for ln in text.split('\n'):
        ln = ln.strip()
        if not ln:
            continue
        sr += 1
        # Strip leading sr. number if present
        body = re.sub(r'^\d+\s*[.)]\s*', '', ln)
        # Split on pipes â€” first part is "Name: Range", rest may be Result/Remarks
        seg = [s.strip() for s in body.split('|')]
        name, rng, result, remark = '', '', '', ''
        if seg:
            m = re.match(r'^(.+?):\s*(.+)$', seg[0])
            if m:
                name, rng = m.group(1).strip(), m.group(2).strip()
            else:
                name = seg[0]
            for extra in seg[1:]:
                m2 = re.match(r'^(result|remark[s]?)\s*:\s*(.+)$', extra, re.I)
                if m2:
                    key = m2.group(1).lower()
                    if 'result' in key: result = m2.group(2).strip()
                    else:               remark = m2.group(2).strip()
        parts.append(
            f'<tr><td>{sr}</td>'
            f'<td><b>{_html.escape(name)}</b></td>'
            f'<td>{_html.escape(rng)}</td>'
            f'<td>{_html.escape(result)}</td>'
            f'<td>{_html.escape(remark)}</td></tr>'
        )
    parts.append('</tbody></table>')
    return ''.join(parts)


# â”€â”€â”€ Auto-extract the "MANUFACTURING PROCESS" section â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _extract_process_text(ws, start_row):
    """Scan rows from start_row downward.  When we find a row containing
    'MANUFACTURING PROCESS' (case-insensitive), capture every subsequent
    row as a numbered step until we hit a known section break (Specs,
    BATCH INCHARGE, etc.) or the end of the sheet.

    Returns a plain-text string with one step per line, e.g.:
        1. Weigh all ingredients as per formulation sheet
        2. Make sure all vessels should be neat & cleaned with DM water
        ...
    """
    found_header   = False
    steps          = []
    max_col        = min((ws.max_column or 1) + 1, 12)

    for r in range(start_row, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, max_col)]
        row_text = ' '.join(str(c).strip() for c in cells if c not in (None, '')).strip()

        # Blank row: skip but don't terminate
        if not row_text:
            continue

        low = row_text.lower()

        # Section header â€” start capturing
        if not found_header:
            if re.search(r'manufactur\w*\s*process|^\s*process\s*$', low):
                found_header = True
            continue

        # Known section terminators
        if re.search(r'\bspecs?\b|\bspecification', low, re.I):
            break
        if re.search(r'batch\s*incharge|material\s*dispenser|approved\s*by', low, re.I):
            break

        # Inside process section â€” capture a step
        col_a = cells[0]

        # Variant 1: Sr.No in Col A (numeric), step text in Col B+
        if isinstance(col_a, (int, float)) and 0 < col_a < 200:
            txt = ' '.join(str(c).strip() for c in cells[1:] if c not in (None, '')).strip()
            if txt:
                steps.append(f'{int(col_a)}. {txt}')
            continue

        # Variant 2: Col A is empty â†’ continuation of the previous step
        # (very common in user's sheets â€” a step's text spans 2 rows)
        if (col_a is None or (isinstance(col_a, str) and not col_a.strip())) and steps:
            steps[-1] = steps[-1].rstrip() + ' ' + row_text
            continue

        # Variant 3: stray rows â€” skip very short labels
        if len(row_text) <= 4:
            continue

        # Variant 4: text already starts with "1.", "2)" etc.
        if re.match(r'^\d+\s*[.)\-]', row_text):
            steps.append(row_text)
        else:
            steps.append(f'{len(steps) + 1}. {row_text}')

    return '\n'.join(steps)


def _process_text_to_html(text):
    """Convert plain numbered text into a clean <ol> HTML block for
    CKEditor.  Lines starting with '1.', '2)', '3-' etc. become list
    items; everything else becomes a paragraph."""
    if not text or not text.strip():
        return ''
    import html as _html
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    list_items, free_paras = [], []
    for ln in lines:
        m = re.match(r'^(\d+)\s*[.)\-]\s*(.+)$', ln)
        if m:
            list_items.append(_html.escape(m.group(2)))
        else:
            free_paras.append(_html.escape(ln))
    parts = []
    if list_items:
        parts.append('<ol>' + ''.join(f'<li>{s}</li>' for s in list_items) + '</ol>')
    for p in free_paras:
        parts.append(f'<p>{p}</p>')
    return ''.join(parts)


@formulation_bp.route('/api/import/preview', methods=['POST'])
@login_required
def api_import_preview():
    """Receive XLSX upload â†’ return one preview block per worksheet."""
    if not _can('add'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    f = request.files.get('file')
    if not f:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400

    fname = (f.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xlsm', '.xls')):
        return jsonify({'status': 'error',
                        'message': 'Only .xlsx / .xlsm / .xls allowed'}), 400

    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
        import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Cannot read file: {e}'}), 400

    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            meta, rows, process, specs = _parse_sheet(ws)
        except Exception as e:
            sheets.append({'sheet_name': sheet_name,
                           'valid': False,
                           'error': f'Parse error: {e}',
                           'ingredients': [],
                           'manufacturing_process': '',
                           'specifications': ''})
            continue

        # Decide the canonical formulation name â€” prefer sheet name (your UI)
        formulation_name = sheet_name.strip() or meta.get('name') or 'Untitled'

        # Already exists?
        exists = (Formulation.query
                  .filter(db.func.lower(Formulation.name) == formulation_name.lower(),
                          Formulation.is_deleted == False).first())

        # Count specs rows (plain-text format: one row per non-empty line)
        spec_row_count = len([l for l in (specs or '').split('\n') if l.strip()]) if specs else 0

        sheets.append({
            'sheet_name'           : sheet_name,
            'formulation_name'     : formulation_name,
            'product_name'         : meta.get('name', ''),
            'product_code'         : meta.get('product_code', ''),
            'batch_size'           : meta.get('batch_size', 0),
            'batch_uom'            : meta.get('batch_uom', 'KG'),
            'ingredients'          : rows,
            'ingredient_count'     : len(rows),
            'manufacturing_process': process,
            'process_step_count'   : len([l for l in (process or '').split('\n') if l.strip()]),
            'specifications'       : specs,
            'specs_row_count'      : spec_row_count,
            'already_exists'       : bool(exists),
            'existing_id'          : exists.id if exists else None,
            'valid'                : len(rows) > 0,
            'error'                : '' if rows else 'No ingredient rows detected',
        })

    return jsonify({
        'status': 'ok',
        'file_name': f.filename,
        'sheets': sheets,
        'total_sheets': len(sheets),
        'valid_sheets': sum(1 for s in sheets if s.get('valid')),
        'invalid_sheets': sum(1 for s in sheets if not s.get('valid')),
    })


@formulation_bp.route('/api/import/commit', methods=['POST'])
@login_required
def api_import_commit():
    """Persist selected sheets â€” each sheet becomes one Formulation."""
    if not _can('add'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d = request.get_json(silent=True) or {}
    sheets = d.get('sheets') or []
    brand  = (d.get('brand') or '').strip()
    # Global process kept as fallback when a sheet doesn't supply its own
    global_process = (d.get('manufacturing_process') or '').strip()

    if not sheets:
        return jsonify({'status': 'error', 'message': 'No sheets to import'}), 400

    created = []
    skipped = []

    for s in sheets:
        nm = (s.get('formulation_name') or '').strip()
        if not nm:
            skipped.append({'sheet': s.get('sheet_name', '?'), 'reason': 'No name'})
            continue
        dup = (Formulation.query
               .filter(db.func.lower(Formulation.name) == nm.lower(),
                       Formulation.is_deleted == False).first())
        if dup:
            skipped.append({'sheet': nm, 'reason': 'Already exists'})
            continue

        # Per-sheet process overrides the global one. If the value looks like
        # plain text (no HTML tags), auto-convert into <ol><li>â€¦</li></ol>
        # so CKEditor renders it as a real numbered list.
        sheet_process = (s.get('manufacturing_process') or '').strip() or global_process
        if sheet_process and '<' not in sheet_process:
            sheet_process = _process_text_to_html(sheet_process)

        # Specifications â€” convert to clean HTML table whether the parser
        # gave us plain text or the user typed text manually.  Existing
        # HTML (with tags) is passed through untouched.
        sheet_specs = (s.get('specifications') or '').strip()
        if sheet_specs and '<' not in sheet_specs:
            sheet_specs = _specs_text_to_html(sheet_specs)

        f = Formulation(
            name                  = nm,
            product_code          = (s.get('product_code') or '').strip(),
            batch_size            = _num(s.get('batch_size'), 0),
            batch_uom             = (s.get('batch_uom') or 'KG').strip(),
            brand                 = brand,
            manufacturing_process = sheet_process,
            specifications        = sheet_specs,
            created_by            = _cu(),
            updated_by            = _cu(),
        )
        db.session.add(f)
        db.session.flush()

        sr = 0
        for raw in (s.get('ingredients') or []):
            nmi = (raw.get('ingredient_name') or '').strip()
            if not nmi:
                continue
            sr += 1
            pct = _num(raw.get('percentage'), 0)
            rm  = _num(raw.get('rm_rate_per_kg'), 0)
            bulk = _num(raw.get('bulk_rate_per_kg'), pct * rm)
            db.session.add(FormulationIngredient(
                formulation_id   = f.id,
                sr_no            = raw.get('sr_no') or sr,
                ingredient_name  = nmi,
                supplier_name    = (raw.get('supplier_name') or '').strip(),
                percentage       = pct,
                qty_kg           = _num(raw.get('qty_kg'), 0),
                uom              = (raw.get('uom') or 'KG').strip(),
                is_additional    = bool(raw.get('is_additional')),
                rm_rate_per_kg   = rm,
                bulk_rate_per_kg = bulk,
            ))
        created.append({'id': f.id, 'name': f.name, 'rows': sr})

    db.session.commit()
    return jsonify({'status': 'ok', 'created': created, 'skipped': skipped})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  EXCEL EXPORT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _process_html_to_steps(html_text):
    """Convert stored manufacturing_process HTML back to a list of plain-text
    step strings (one per <li>, or one per non-blank line as a fallback)."""
    if not html_text or not html_text.strip():
        return []
    import html as _html
    out = []
    items = re.findall(r'<li[^>]*>(.*?)</li>', html_text, re.DOTALL | re.IGNORECASE)
    if items:
        for it in items:
            t = re.sub(r'<[^>]+>', ' ', it)
            t = re.sub(r'\s+', ' ', t).strip()
            t = _html.unescape(t)
            if t:
                out.append(t)
        return out
    tmp = re.sub(r'<\s*br\s*/?\s*>', '\n', html_text, flags=re.IGNORECASE)
    tmp = re.sub(r'</\s*p\s*>', '\n', tmp, flags=re.IGNORECASE)
    tmp = re.sub(r'<[^>]+>', '', tmp)
    tmp = _html.unescape(tmp)
    for ln in tmp.split('\n'):
        ln = re.sub(r'^\s*\d+\s*[.):\-]\s*', '', ln).strip()
        if ln:
            out.append(ln)
    return out


def _specs_html_to_rows(html_text):
    """Convert stored specifications HTML back to a list of
    (name, range, result, remarks) tuples for the export sheet."""
    if not html_text or not html_text.strip():
        return []
    import html as _html
    out = []
    tr_blocks = re.findall(r'<tr[^>]*>(.*?)</tr>', html_text, re.DOTALL | re.IGNORECASE)
    if tr_blocks:
        for blk in tr_blocks:
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', blk, re.DOTALL | re.IGNORECASE)
            cells = [_html.unescape(re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', c)).strip())
                     for c in cells]
            if not cells:
                continue
            joined = ' '.join(cells).lower()
            if ('specs' in joined or 'specification' in joined) and 'range' in joined:
                continue   # header row
            if cells and re.fullmatch(r'\d+\.?', cells[0] or ''):
                cells = cells[1:]
            if not cells or not cells[0]:
                continue
            name   = cells[0]
            rng    = cells[1] if len(cells) > 1 else ''
            result = cells[2] if len(cells) > 2 else ''
            remark = cells[3] if len(cells) > 3 else ''
            out.append((name, rng, result, remark))
        return out
    # Plain-text fallback: "Name: range | Result: x | Remarks: y"
    for ln in html_text.split('\n'):
        ln = re.sub(r'^\s*\d+\s*[.)]\s*', '', ln).strip()
        if not ln:
            continue
        segs = [s.strip() for s in ln.split('|')]
        name, rng, res, rem = '', '', '', ''
        m = re.match(r'^(.+?):\s*(.+)$', segs[0]) if segs else None
        if m:
            name, rng = m.group(1).strip(), m.group(2).strip()
        else:
            name = segs[0] if segs else ''
        for extra in segs[1:]:
            m2 = re.match(r'^(result|remark[s]?)\s*:\s*(.+)$', extra, re.I)
            if m2:
                if 'result' in m2.group(1).lower():
                    res = m2.group(2).strip()
                else:
                    rem = m2.group(2).strip()
        if name:
            out.append((name, rng, res, rem))
    return out


def _build_workbook_for_view(formulations, view):
    """Dispatcher: route to the right builder based on the chosen view.
    view âˆˆ {'default', 'supplier', 'percentage', 'all'}.
    'all' returns the full Beardo production sheet; the others a simpler
    clean table with only the requested columns."""
    view = (view or 'all').strip().lower()
    if view == 'all':
        return _build_workbook(formulations)
    return _build_workbook_simple(formulations, view)


def _build_workbook_simple(formulations, view):
    """Build a clean, column-filtered ingredient sheet per formulation.
    No manufacturing process / specs / approval box â€” those live in the
    full ('all') view only.  view âˆˆ {'default','supplier','percentage'}.

      â€¢ default     : #  Addl.  Item  Qty  UOM
      â€¢ supplier    : #  Addl.  Item  Supplier  Qty  UOM
      â€¢ percentage  : #  Addl.  Item  Qty (%)   Qty  UOM
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Column layout per view (column_letter, header, key_or_callable, width)
    COLUMNS = {
        'default': [
            ('A', 'Sr. No.',           lambda i: i.sr_no,                                     7),
            ('B', 'Addl.',             lambda i: 'Yes' if i.is_additional else '',           8),
            ('C', 'Item',              lambda i: i.ingredient_name or '',                    38),
            ('D', 'Qty',               lambda i: float(i.qty_kg or 0),                       14),
            ('E', 'UOM',               lambda i: i.uom or 'KG',                              8),
        ],
        'supplier': [
            ('A', 'Sr. No.',           lambda i: i.sr_no,                                     7),
            ('B', 'Addl.',             lambda i: 'Yes' if i.is_additional else '',           8),
            ('C', 'Item',              lambda i: i.ingredient_name or '',                    34),
            ('D', 'Supplier',          lambda i: i.supplier_name or '',                      20),
            ('E', 'Qty',               lambda i: float(i.qty_kg or 0),                       14),
            ('F', 'UOM',               lambda i: i.uom or 'KG',                              8),
        ],
        'percentage': [
            ('A', 'Sr. No.',           lambda i: i.sr_no,                                     7),
            ('B', 'Addl.',             lambda i: 'Yes' if i.is_additional else '',           8),
            ('C', 'Item',              lambda i: i.ingredient_name or '',                    38),
            ('D', 'Qty (%)',           '__pct__',                                            12),
            ('E', 'Qty',               lambda i: float(i.qty_kg or 0),                       14),
            ('F', 'UOM',               lambda i: i.uom or 'KG',                              8),
        ],
    }
    cols = COLUMNS.get(view, COLUMNS['default'])
    n_cols = len(cols)
    last_col_letter = cols[-1][0]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Styles
    f_title  = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
    f_hdr    = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    f_bold12 = Font(name='Calibri', size=12, bold=True)
    f_norm12 = Font(name='Calibri', size=12, bold=False)
    fill_title  = PatternFill('solid', fgColor='1E3A5F')
    fill_header = PatternFill('solid', fgColor='2563EB')
    fill_total  = PatternFill('solid', fgColor='F1F5F9')

    thin    = Side(style='thin',   color='94A3B8')
    medium  = Side(style='medium', color='1E3A5F')
    b_all_thin   = Border(left=thin, right=thin, top=thin, bottom=thin)

    a_center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    a_left_center = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    a_right_center= Alignment(horizontal='right',  vertical='center')

    used_names = set()
    for f in formulations:
        ings = f.resolved_ingredients()

        # Sheet name
        safe = re.sub(r'[:\\/?*\[\]]', '_', f.name)[:31] or f'Formulation_{f.id}'
        base = safe; n = 2
        while safe in used_names:
            suff = f' ({n})'; safe = (base[:31 - len(suff)] + suff); n += 1
        used_names.add(safe)
        ws = wb.create_sheet(safe)

        # â”€â”€â”€ Title row 1 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        cc = ws.cell(row=1, column=1, value='HCP WELLNESS PVT. LTD.')
        cc.font = f_title; cc.alignment = a_center; cc.fill = fill_title
        ws.row_dimensions[1].height = 28

        # â”€â”€â”€ Meta row 3 (Product / Brand / Batch Size / Code) â”€â”€â”€â”€
        meta_parts = []
        if f.name:         meta_parts.append(f'Product: {f.name}')
        if f.brand:        meta_parts.append(f'Brand: {f.brand}')
        if f.product_code: meta_parts.append(f'Code: {f.product_code}')
        if f.batch_size:   meta_parts.append(f'Batch Size: {float(f.batch_size):g} {f.batch_uom or "KG"}')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
        cc = ws.cell(row=3, column=1, value='  â€¢  '.join(meta_parts))
        cc.font = f_bold12; cc.alignment = a_left_center
        ws.row_dimensions[3].height = 20

        # â”€â”€â”€ Header row 5 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        for i, (letter, label, _value, width) in enumerate(cols, start=1):
            cc = ws.cell(row=5, column=i, value=label)
            cc.font = f_hdr; cc.fill = fill_header
            cc.alignment = a_center
            cc.border = b_all_thin
            ws.column_dimensions[letter].width = width
        ws.row_dimensions[5].height = 24

        # â”€â”€â”€ Body rows from row 6 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Percentage scaling heuristic (same as full builder)
        raw_pcts = [float(i.percentage or 0) for i in ings]
        pct_scale = 0.01 if (raw_pcts and max(raw_pcts) > 1.5) else 1.0

        r = 6
        first_row = r
        for ing in ings:
            for ci, (letter, _label, value, _w) in enumerate(cols, start=1):
                if value == '__pct__':
                    v = float(ing.percentage or 0) * pct_scale
                    cc = ws.cell(row=r, column=ci, value=v)
                    cc.number_format = '0.00%'
                    cc.alignment = a_right_center
                else:
                    v = value(ing)
                    cc = ws.cell(row=r, column=ci, value=v)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        cc.alignment = a_right_center
                        if ci > 1:   # Sr.No stays integer; qty/percentage gets decimals
                            cc.number_format = '0.0000'
                    else:
                        cc.alignment = a_left_center if ci >= 3 else a_center
                cc.font = f_norm12
                cc.border = b_all_thin
            r += 1

        # â”€â”€â”€ Total row â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if ings:
            total_row = r
            # Find columns to sum: Qty% (if present) and Qty
            cc = ws.cell(row=total_row, column=1, value='')
            cc.fill = fill_total; cc.border = b_all_thin
            cc = ws.cell(row=total_row, column=2, value='')
            cc.fill = fill_total; cc.border = b_all_thin
            cc = ws.cell(row=total_row, column=3, value='Total')
            cc.font = f_bold12; cc.fill = fill_total
            cc.alignment = Alignment(horizontal='right', vertical='center')
            cc.border = b_all_thin

            for ci, (letter, _label, value, _w) in enumerate(cols, start=1):
                if ci <= 3:
                    continue
                cell = ws.cell(row=total_row, column=ci)
                cell.fill = fill_total
                cell.border = b_all_thin
                if value == '__pct__':
                    cell.value = f'=SUM({letter}{first_row}:{letter}{r - 1})'
                    cell.font = f_bold12; cell.number_format = '0.00%'
                    cell.alignment = a_right_center
                elif _label == 'Qty':
                    cell.value = f'=SUM({letter}{first_row}:{letter}{r - 1})'
                    cell.font = f_bold12; cell.number_format = '0.0000'
                    cell.alignment = a_right_center

        # Apply outer medium border around the whole table
        from openpyxl.utils import get_column_letter
        max_r = r if ings else 5
        for col_n in range(1, n_cols + 1):
            top    = ws.cell(row=5, column=col_n)
            bottom = ws.cell(row=max_r, column=col_n)
            top.border    = Border(left=top.border.left,    right=top.border.right,    top=medium,    bottom=top.border.bottom)
            bottom.border = Border(left=bottom.border.left, right=bottom.border.right, top=bottom.border.top, bottom=medium)
        for row_n in range(5, max_r + 1):
            left  = ws.cell(row=row_n, column=1)
            right = ws.cell(row=row_n, column=n_cols)
            left.border  = Border(left=medium,  right=left.border.right,  top=left.border.top,  bottom=left.border.bottom)
            right.border = Border(left=right.border.left, right=medium, top=right.border.top, bottom=right.border.bottom)

        # Freeze header row
        ws.freeze_panes = 'A6'

    if not wb.sheetnames:
        wb.create_sheet('Empty')
    return wb


def _build_workbook(formulations):
    """Build an openpyxl workbook with one sheet per formulation, matching
    the Beardo costing template (rows 1â€“63, columns Aâ€“J) pixel-for-pixel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # drop default sheet

    # â”€â”€ Fonts â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    f_title  = Font(name='Calibri', size=25, bold=False)
    f_bold12 = Font(name='Calibri', size=12, bold=True)
    f_norm12 = Font(name='Calibri', size=12, bold=False)
    f_bold11 = Font(name='Calibri', size=11, bold=True)

    # â”€â”€ Borders â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    thin   = Side(style='thin',   color='000000')
    medium = Side(style='medium', color='000000')

    b_all_thin   = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_all_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    # â”€â”€ Alignment â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    a_center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    a_center_nw   = Alignment(horizontal='center', vertical='center')
    a_left_center = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    a_left_top    = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    a_center_top  = Alignment(horizontal='center', vertical='top',    wrap_text=True)

    # Indian Rupee currency format used for hidden RM rate / bulk cost columns
    inr_fmt = '_ "â‚¹"\\ * #,##0.00_ ;_ "â‚¹"\\ * \\-#,##0.00_ ;_ "â‚¹"\\ * "-"??_ ;_ @_ '

    used_names = set()
    for f in formulations:
        ings = f.resolved_ingredients()

        # Excel limits sheet names to 31 chars and forbids : \ / ? * [ ]
        safe = re.sub(r'[:\\/?*\[\]]', '_', f.name)[:31] or f'Formulation_{f.id}'
        base = safe; n = 2
        while safe in used_names:
            suff = f' ({n})'; safe = (base[:31 - len(suff)] + suff); n += 1
        used_names.add(safe)
        ws = wb.create_sheet(safe)

        # â”€â”€ (1) Title  A1:J3 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        ws.merge_cells('A1:J3')
        c = ws['A1']
        c.value = 'HCP WELLNESS PVT. LTD.'
        c.font  = f_title
        c.alignment = a_center_nw
        for row_n in (1, 2, 3):
            for col_n in range(1, 11):
                ws.cell(row=row_n, column=col_n).border = Border(
                    left   = medium if col_n == 1  else None,
                    right  = medium if col_n == 10 else None,
                    top    = medium if row_n == 1  else None,
                    bottom = medium if row_n == 3  else None,
                )

        # Empty separator row 4 (matches Beardo)
        ws.merge_cells('A4:J4')

        # "Main sheet" label at column K row 5 (small note, off-screen)
        ws['K5'] = 'Main sheet'
        ws['K5'].font = Font(name='Calibri', size=12, bold=False)

        # â”€â”€ (2) Header info rows 6â€“10 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Left labels rows 6-9 ; Right labels rows 7-10 (shifted by 1)
        fd = (f.created_at.strftime('%d/%m/%Y')
              if getattr(f, 'created_at', None) else '')
        left_labels = [
            (6, f'Product Name: {f.name}'),
            (7, f'Variant: {f.brand or ""}'),
            (8, f'Product Code: {f.product_code or ""}'),
            (9, f'Formulation Date : {fd}'),
        ]
        right_labels = [
            (7,  'Batch No.: '),
            (8,  'Batch Size: '),
            (9,  'Batch Date: '),
            (10, 'Exp. Date: '),
        ]
        for r, lbl in left_labels:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cc = ws.cell(row=r, column=1, value=lbl)
            cc.font = f_bold12; cc.alignment = a_left_center
            for col_n in range(1, 5):
                ws.cell(row=r, column=col_n).border = Border(
                    left  = thin if col_n == 1 else None,
                    right = thin if col_n == 4 else None,
                    top=thin, bottom=thin,
                )

        for r, lbl in right_labels:
            cc = ws.cell(row=r, column=5, value=lbl)
            cc.font = f_bold12; cc.alignment = a_left_center
            cc.border = b_all_thin

        # H7:J7, H9:J9, H10:J10 â€” merged value cells (empty / filled)
        for r in (7, 9, 10):
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
            for col_n in range(8, 11):
                ws.cell(row=r, column=col_n).border = Border(
                    left  = thin if col_n == 8  else None,
                    right = thin if col_n == 10 else None,
                    top=thin, bottom=thin,
                )

        # Row 8 special: H8:I8 holds batch-size number, J8 holds UOM
        ws.merge_cells('H8:I8')
        ws['H8'] = float(f.batch_size or 0)
        ws['H8'].font = f_bold12
        ws['H8'].alignment = a_center_nw
        for col_n in (8, 9):
            ws.cell(row=8, column=col_n).border = Border(
                left  = thin if col_n == 8 else None,
                right = thin if col_n == 9 else None,
                top=thin, bottom=thin,
            )
        ws['J8'] = f.batch_uom or 'KG'
        ws['J8'].font = f_bold12
        ws['J8'].alignment = a_center_nw
        ws['J8'].border = b_all_thin

        # Row 9 â€” current batch date in H9 (merged H9:J9)
        ws['H9'] = datetime.now().strftime('%d-%m-%Y')
        ws['H9'].font = f_bold12
        ws['H9'].alignment = a_center_nw

        # â”€â”€ (3) "FORMULATION" heading row 11 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        ws.merge_cells('A11:J11')
        c = ws['A11']
        c.value = 'FORMULATION'
        c.font  = f_bold12
        c.alignment = a_center_nw
        for col_n in range(1, 11):
            ws.cell(row=11, column=col_n).border = Border(
                left   = medium if col_n == 1  else None,
                right  = medium if col_n == 10 else None,
                top    = medium, bottom = None,
            )

        # â”€â”€ (4) Ingredient table header row 12 â€” 10 columns â”€â”€â”€â”€
        headers = ['Sr. No.', 'Ingredients', 'Supplier Name', 'Qty (%)',
                   'Actual Qty. in kg', 'START TIME', 'END TIME',
                   'TEMPERATURE', 'RPM', 'Process Name']
        for i, h in enumerate(headers, start=1):
            cc = ws.cell(row=12, column=i, value=h)
            cc.font = f_bold12
            cc.alignment = a_center
            cc.border = Border(
                left   = medium if i == 1  else thin,
                right  = medium if i == 10 else thin,
                top    = medium,
                bottom = thin,
            )
        ws.row_dimensions[12].height = 31.5

        # â”€â”€ (5) Ingredient rows â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Heuristic: if values look like percentage numbers (max > 1.5),
        # divide by 100 so the '0%' format renders correctly.
        raw_pcts = [float(i.percentage or 0) for i in ings]
        scale = 0.01 if (raw_pcts and max(raw_pcts) > 1.5) else 1.0

        r = 13
        first_row = r
        for ing in ings:
            ws.cell(row=r, column=1, value=ing.sr_no).font = f_norm12
            ws.cell(row=r, column=1).alignment = a_center_top
            ws.cell(row=r, column=1).border = Border(left=medium, right=thin, top=thin, bottom=thin)

            ws.cell(row=r, column=2, value=ing.ingredient_name).font = f_norm12
            ws.cell(row=r, column=2).alignment = a_left_top
            ws.cell(row=r, column=2).border = b_all_thin

            ws.cell(row=r, column=3, value=ing.supplier_name or '').font = f_norm12
            ws.cell(row=r, column=3).alignment = a_left_top
            ws.cell(row=r, column=3).border = b_all_thin

            # Qty (%) â€” decimal fraction, displayed with 2 decimals (e.g. 4.00%)
            ws.cell(row=r, column=4, value=float(ing.percentage or 0) * scale).font = f_norm12
            ws.cell(row=r, column=4).number_format = '0.00%'
            ws.cell(row=r, column=4).alignment = a_center_top
            ws.cell(row=r, column=4).border = b_all_thin

            # Actual Qty (kg) â€” formula = Qty% Ã— batch_size (H8)
            ws.cell(row=r, column=5, value=f'=D{r}*$H$8').font = f_norm12
            ws.cell(row=r, column=5).alignment = a_center_top
            ws.cell(row=r, column=5).border = b_all_thin

            # START TIME / END TIME / TEMPERATURE / RPM / Process Name â€” blanks
            for col_n in (6, 7, 8, 9, 10):
                cc = ws.cell(row=r, column=col_n, value='')
                cc.font = f_norm12
                cc.alignment = a_center_top
                cc.border = Border(
                    left=thin,
                    right=medium if col_n == 10 else thin,
                    top=thin, bottom=thin,
                )

            # Off-screen rate columns M (RM rate) & N (bulk cost)
            rm = float(ing.rm_rate_per_kg or 0)
            ws.cell(row=r, column=13, value=rm if rm else None)
            ws.cell(row=r, column=13).number_format = inr_fmt

            bulk = float(ing.bulk_rate_per_kg or 0)
            ws.cell(row=r, column=14, value=bulk if bulk else None)
            ws.cell(row=r, column=14).number_format = inr_fmt
            r += 1

        # â”€â”€ (6) Total row â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        total_row = r
        if ings:
            ws.cell(row=total_row, column=1).border = Border(
                left=medium, right=thin, top=None, bottom=medium,
            )
            cc = ws.cell(row=total_row, column=2, value='Total')
            cc.font = f_bold12
            cc.alignment = Alignment(horizontal='center', vertical='center')
            cc.border = Border(left=thin, right=thin, top=None, bottom=medium)

            ws.cell(row=total_row, column=3).border = Border(
                left=thin, right=thin, top=None, bottom=medium,
            )

            cc = ws.cell(row=total_row, column=4, value=f'=SUM(D{first_row}:D{r - 1})')
            cc.font = f_bold12; cc.number_format = '0.00%'
            cc.alignment = a_center_nw
            cc.border = Border(left=thin, right=thin, top=None, bottom=medium)

            cc = ws.cell(row=total_row, column=5, value=f'=SUM(E{first_row}:E{r - 1})')
            cc.font = f_bold12
            cc.alignment = a_center_nw
            cc.border = Border(left=thin, right=thin, top=None, bottom=medium)

            for col_n in (6, 7, 8, 9, 10):
                ws.cell(row=total_row, column=col_n).border = Border(
                    left=thin,
                    right=medium if col_n == 10 else thin,
                    top=None, bottom=medium,
                )

            cc = ws.cell(row=total_row, column=14, value=f'=SUM(N{first_row}:N{r - 1})')
            cc.font = f_bold12
            cc.number_format = inr_fmt

        # â”€â”€ (6.5) "Total Extract Used" label  (matches B27 in Beardo) â”€
        extract_row = total_row + 2
        cc = ws.cell(row=extract_row, column=2, value='Total Extract Used')
        cc.font = f_bold12
        cc.alignment = a_left_center
        for cn in range(2, 11):
            ws.cell(row=extract_row, column=cn).border = Border(
                left   = thin if cn == 2  else None,
                right  = thin if cn == 10 else None,
                top    = thin, bottom = thin,
            )
        ws.merge_cells(start_row=extract_row, start_column=2,
                       end_row=extract_row,   end_column=10)

        # â”€â”€ (7) Approval signature box â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        appr_hdr = total_row + 4 if total_row >= 25 else 29
        appr_end = appr_hdr + 5   # 2-row header + 4-row signature

        for (start_col, end_col, label) in [(1, 3, 'BATCH INCHARGE'),
                                            (4, 7, 'MATERIAL DISPENSER'),
                                            (8, 10, 'APPROVED BY')]:
            ws.merge_cells(start_row=appr_hdr, start_column=start_col,
                           end_row=appr_hdr + 1, end_column=end_col)
            cc = ws.cell(row=appr_hdr, column=start_col, value=label)
            cc.font = f_bold12
            cc.alignment = a_center_nw

            ws.merge_cells(start_row=appr_hdr + 2, start_column=start_col,
                           end_row=appr_end,       end_column=end_col)

            for rr in range(appr_hdr, appr_end + 1):
                for cn in range(start_col, end_col + 1):
                    ws.cell(row=rr, column=cn).border = Border(
                        left   = medium if cn == start_col else None,
                        right  = medium if cn == end_col   else None,
                        top    = medium if rr == appr_hdr  else None,
                        bottom = medium if rr == appr_end  else None,
                    )

        # â”€â”€ (8) Manufacturing process â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        proc_hdr = appr_end + 3 if total_row >= 25 else 37
        steps = _process_html_to_steps(f.manufacturing_process or '')

        # Title row â€” CONCATENATE formula matches Beardo exactly
        ws.merge_cells(start_row=proc_hdr, start_column=2, end_row=proc_hdr, end_column=10)
        cc = ws.cell(row=proc_hdr, column=2,
                     value='=CONCATENATE(RIGHT((A6), LEN(A6)-13), "-MANUFACTURING PROCESS")')
        cc.font = f_bold11
        cc.alignment = a_center_nw
        cc.border = Border(left=thin, right=thin, top=medium, bottom=thin)

        # Step rows â€” each fully bordered like Beardo
        step_r = proc_hdr + 1
        for i, txt in enumerate(steps, start=1):
            cc = ws.cell(row=step_r, column=1, value=i)
            cc.font = f_norm12
            cc.alignment = a_center_nw
            cc.border = b_all_thin

            ws.merge_cells(start_row=step_r, start_column=2, end_row=step_r, end_column=10)
            cc = ws.cell(row=step_r, column=2, value=txt)
            cc.font = f_norm12
            cc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for cn in range(2, 11):
                ws.cell(row=step_r, column=cn).border = Border(
                    left   = thin if cn == 2  else None,
                    right  = thin if cn == 10 else None,
                    top=thin, bottom=thin,
                )
            step_r += 1

        # â”€â”€ (9) Specifications table â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        spec_hdr = step_r + 1
        # Headers: Specs (B:C) | Range (D:G) | Result (H) | Remarks (I:J)
        ws.merge_cells(start_row=spec_hdr, start_column=2, end_row=spec_hdr, end_column=3)
        cc = ws.cell(row=spec_hdr, column=2, value='Specs.')
        cc.font = f_bold11; cc.alignment = a_center_nw
        for cn in (2, 3):
            ws.cell(row=spec_hdr, column=cn).border = Border(
                left=thin if cn == 2 else None,
                right=thin if cn == 3 else None,
                top=medium, bottom=thin,
            )

        ws.merge_cells(start_row=spec_hdr, start_column=4, end_row=spec_hdr, end_column=7)
        cc = ws.cell(row=spec_hdr, column=4, value='Range')
        cc.font = f_bold11; cc.alignment = a_center_nw
        for cn in (4, 5, 6, 7):
            ws.cell(row=spec_hdr, column=cn).border = Border(
                left=thin if cn == 4 else None,
                right=thin if cn == 7 else None,
                top=medium, bottom=thin,
            )

        cc = ws.cell(row=spec_hdr, column=8, value='Result')
        cc.font = f_bold11; cc.alignment = a_center_nw
        cc.border = Border(left=thin, right=thin, top=medium, bottom=thin)

        ws.merge_cells(start_row=spec_hdr, start_column=9, end_row=spec_hdr, end_column=10)
        cc = ws.cell(row=spec_hdr, column=9, value='Remarks')
        cc.font = f_bold11; cc.alignment = a_center_nw
        for cn in (9, 10):
            ws.cell(row=spec_hdr, column=cn).border = Border(
                left=thin if cn == 9 else None,
                right=thin if cn == 10 else None,
                top=medium, bottom=thin,
            )

        # Spec data rows
        specs_list = _specs_html_to_rows(f.specifications or '')
        spec_r = spec_hdr + 1
        for idx, (name, rng, res, rem) in enumerate(specs_list, start=1):
            cc = ws.cell(row=spec_r, column=1, value=idx)
            cc.font = f_norm12
            cc.alignment = a_center_nw
            cc.border = b_all_thin

            ws.merge_cells(start_row=spec_r, start_column=2, end_row=spec_r, end_column=3)
            cc = ws.cell(row=spec_r, column=2, value=name)
            cc.font = f_bold11
            cc.alignment = a_center_nw
            for cn in (2, 3):
                ws.cell(row=spec_r, column=cn).border = Border(
                    left=thin if cn == 2 else None,
                    right=thin if cn == 3 else None,
                    top=thin, bottom=thin,
                )

            ws.merge_cells(start_row=spec_r, start_column=4, end_row=spec_r, end_column=7)
            cc = ws.cell(row=spec_r, column=4, value=rng)
            cc.font = f_norm12
            cc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for cn in (4, 5, 6, 7):
                ws.cell(row=spec_r, column=cn).border = Border(
                    left=thin if cn == 4 else None,
                    right=thin if cn == 7 else None,
                    top=thin, bottom=thin,
                )

            cc = ws.cell(row=spec_r, column=8, value=res)
            cc.font = f_norm12
            cc.alignment = a_center_nw
            cc.border = b_all_thin

            cc = ws.cell(row=spec_r, column=9, value=rem)
            cc.font = f_norm12
            cc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cc.border = b_all_thin
            spec_r += 1

        # â”€â”€ (10) After Approval Batch Yield â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        yield_r = spec_r + 8
        ws.merge_cells(start_row=yield_r, start_column=1, end_row=yield_r, end_column=2)
        cc = ws.cell(row=yield_r, column=1, value='After Approval Batch Yield qty.')
        cc.font = f_bold11
        cc.alignment = a_center_nw
        cc.border = b_all_medium
        ws.cell(row=yield_r, column=2).border = b_all_medium

        ws.merge_cells(start_row=yield_r, start_column=3, end_row=yield_r, end_column=10)
        for cn in range(3, 11):
            ws.cell(row=yield_r, column=cn).border = Border(
                left   = medium if cn == 3  else None,
                right  = medium if cn == 10 else None,
                top    = medium, bottom = medium,
            )

        # â”€â”€ (11) Column widths (match Beardo exactly) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        ws.column_dimensions['A'].width = 4.43
        ws.column_dimensions['B'].width = 33.86
        ws.column_dimensions['C'].width = 14.14
        ws.column_dimensions['D'].width = 11.14
        ws.column_dimensions['E'].width = 12.86
        ws.column_dimensions['F'].width = 13.57
        # G..J left at Excel default ~ 8.43 (same as Beardo)

    if not wb.sheetnames:                                  # safety
        wb.create_sheet('Empty')
    return wb


@formulation_bp.route('/api/export/<int:fid>')
@login_required
def api_export_one(fid):
    if not _can('export'):
        abort(403)
    f = Formulation.query.get_or_404(fid)
    if f.is_deleted:
        abort(404)
    view = (request.args.get('view') or 'all').strip().lower()
    wb = _build_workbook_for_view([f], view)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = re.sub(r'[^A-Za-z0-9_\-]+', '_', f.name).strip('_') or f'formulation_{fid}'
    suffix = '' if view == 'all' else f'_{view}'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{safe}{suffix}.xlsx')


@formulation_bp.route('/api/export-all')
@login_required
def api_export_all():
    if not _can('export'):
        abort(403)
    q = Formulation.query.filter_by(is_deleted=False)
    brand = (request.args.get('brand') or '').strip()
    if brand:
        q = q.filter(Formulation.brand == brand)
    formulations = q.order_by(Formulation.name).all()
    if not formulations:
        return jsonify({'status': 'error', 'message': 'No formulations to export'}), 400
    view = (request.args.get('view') or 'all').strip().lower()
    wb = _build_workbook_for_view(formulations, view)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    suffix = '' if view == 'all' else f'_{view}'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'formulations{suffix}_{stamp}.xlsx')


