"""
packing_bom_routes.py
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
Packing Material BOM module.

A single BOM per FG (unique constraint on fg_material_id) â€” saving a new BOM
for an FG that already has one will UPSERT (replace items).

Endpoints
â”€â”€â”€â”€â”€â”€â”€â”€â”€
GET    /packing-bom/                      â†’ list page
GET    /packing-bom/api/list              â†’ JSON list (search, deleted filter)
GET    /packing-bom/api/<id>              â†’ JSON view (with items)
GET    /packing-bom/api/by-fg/<fg_id>     â†’ look up existing BOM by FG
POST   /packing-bom/api/save              â†’ create or upsert
DELETE /packing-bom/api/<id>              â†’ soft delete
POST   /packing-bom/api/<id>/restore      â†’ undo soft delete

GET    /packing-bom/api/fg-list           â†’ all FG materials (for dropdown)
GET    /packing-bom/api/pm-list           â†’ all PM/Corrugation/Sleeves materials
GET    /packing-bom/api/uom-list          â†’ all UOMs from master

GET    /packing-bom/api/export/<id>       â†’ single BOM .xlsx download
GET    /packing-bom/api/export-all        â†’ all active BOMs in one workbook
GET    /packing-bom/api/template          â†’ empty Excel template for import
POST   /packing-bom/api/import/preview    â†’ upload + parse, return preview
POST   /packing-bom/api/import/commit     â†’ commit selected previewed sheets
"""
import io
from datetime import datetime
import re
from flask import (Blueprint, render_template, request, jsonify, abort,
                   send_file)
from flask_login import login_required, current_user

from models import db, Material, MaterialType, PackingBOM, PackingBOMItem, UOMMaster
from core.permissions import get_perm

packing_bom_bp = Blueprint('packing_bom', __name__, url_prefix='/packing-bom')


# â”€â”€â”€ Permission helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _can(action):
    """Check the 'packing_bom' permission for the current user."""
    if getattr(current_user, 'role', None) == 'admin':
        return True
    p = get_perm('packing_bom')
    if not p:
        return False
    return bool(getattr(p, f'can_{action}', False))


def _cu():
    return getattr(current_user, 'username', '') or getattr(current_user, 'name', '') or ''


def _num(v, default=0):
    """Best-effort numeric conversion."""
    try:
        if v is None or v == '':
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  PAGE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@packing_bom_bp.route('/')
@login_required
def index():
    if not _can('view'):
        abort(403)
    return render_template(
        'packing_bom/index.html',
        can_view   = True,
        can_add    = _can('add'),
        can_edit   = _can('edit'),
        can_delete = _can('delete'),
        can_export = _can('export'),
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  LIST
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@packing_bom_bp.route('/api/list')
@login_required
def api_list():
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    q = PackingBOM.query
    status = (request.args.get('status') or 'active').lower()
    if status == 'deleted':
        q = q.filter(PackingBOM.is_deleted == True)
    elif status == 'all':
        pass
    else:   # active (default)
        q = q.filter(PackingBOM.is_deleted == False)

    search = (request.args.get('search') or '').strip()
    if search:
        like = f'%{search}%'
        q = q.join(Material, PackingBOM.fg_material_id == Material.id).filter(
            db.or_(Material.material_name.ilike(like),
                   Material.code.ilike(like),
                   Material.brand.ilike(like))
        )

    rows = q.order_by(PackingBOM.updated_at.desc()).all()
    items = [r.to_dict() for r in rows]
    return jsonify({'status': 'ok', 'items': items, 'total': len(items)})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  VIEW ONE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@packing_bom_bp.route('/api/<int:bom_id>')
@login_required
def api_view(bom_id):
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    bom = PackingBOM.query.get_or_404(bom_id)
    return jsonify({'status': 'ok', 'item': bom.to_dict(include_items=True)})


@packing_bom_bp.route('/api/by-fg/<int:fg_id>')
@login_required
def api_view_by_fg(fg_id):
    """Look up an existing BOM for a given FG material id.  Returns
    'not_found' (200) if none exists â€” used by the create form to warn
    the user that saving will overwrite."""
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    bom = PackingBOM.query.filter_by(fg_material_id=fg_id, is_deleted=False).first()
    if not bom:
        return jsonify({'status': 'not_found'})
    return jsonify({'status': 'ok', 'item': bom.to_dict(include_items=True)})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  SAVE  (create OR upsert by FG)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@packing_bom_bp.route('/api/save', methods=['POST'])
@login_required
def api_save():
    if not (_can('add') or _can('edit')):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d = request.get_json(silent=True) or {}
    fg_id  = d.get('fg_material_id')
    fg_qty = _num(d.get('fg_qty'), 0)
    fg_uom = (d.get('fg_uom') or 'PCS').strip() or 'PCS'
    notes  = (d.get('notes')  or '').strip()
    items  = d.get('items') or []

    if not fg_id:
        return jsonify({'status': 'error', 'message': 'FG Item is required'}), 400
    if not items:
        return jsonify({'status': 'error', 'message': 'Add at least one packing item'}), 400

    # Verify FG exists and is actually FG-type
    fg_mat = Material.query.get(fg_id)
    if not fg_mat:
        return jsonify({'status': 'error', 'message': 'FG material not found'}), 404

    # Upsert by fg_material_id (one BOM per FG)
    bom = PackingBOM.query.filter_by(fg_material_id=fg_id, is_deleted=False).first()
    is_new = bom is None
    if is_new:
        if not _can('add'):
            return jsonify({'status': 'error', 'message': 'Add permission required'}), 403
        bom = PackingBOM(fg_material_id=fg_id, created_by=_cu())
        db.session.add(bom)
    else:
        if not _can('edit'):
            return jsonify({'status': 'error', 'message': 'Edit permission required'}), 403

    bom.fg_qty     = fg_qty
    bom.fg_uom     = fg_uom
    bom.notes      = notes
    bom.updated_by = _cu()
    bom.is_active  = True
    bom.is_deleted = False
    bom.deleted_at = None

    # Replace items
    if not is_new:
        for old in bom.items.all():
            db.session.delete(old)
        db.session.flush()

    db.session.flush()   # ensure bom.id exists

    sr = 0
    for raw in items:
        mid = raw.get('material_id')
        qty = _num(raw.get('qty'), 0)
        if not mid:
            continue
        m = Material.query.get(mid)
        if not m:
            continue
        sr += 1
        db.session.add(PackingBOMItem(
            packing_bom_id = bom.id,
            sr_no          = raw.get('sr_no') or sr,
            material_id    = mid,
            qty            = qty,
            item_name_snap = m.material_name or '',
            uom_snap       = m.uom or 'PCS',
        ))

    db.session.commit()
    return jsonify({
        'status': 'ok',
        'mode'  : 'created' if is_new else 'updated',
        'id'    : bom.id,
        'item'  : bom.to_dict(include_items=True),
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  DELETE / RESTORE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@packing_bom_bp.route('/api/<int:bom_id>', methods=['DELETE'])
@login_required
def api_delete(bom_id):
    if not _can('delete'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    bom = PackingBOM.query.get_or_404(bom_id)
    bom.is_deleted = True
    bom.deleted_at = datetime.utcnow()
    bom.updated_by = _cu()
    db.session.commit()
    return jsonify({'status': 'ok'})


@packing_bom_bp.route('/api/<int:bom_id>/restore', methods=['POST'])
@login_required
def api_restore(bom_id):
    if not _can('edit'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    bom = PackingBOM.query.get_or_404(bom_id)

    # If another live BOM already owns this FG, refuse â€” would break the
    # unique constraint.
    conflict = (PackingBOM.query
                .filter(PackingBOM.fg_material_id == bom.fg_material_id,
                        PackingBOM.is_deleted == False,
                        PackingBOM.id != bom.id)
                .first())
    if conflict:
        return jsonify({'status': 'error',
                        'message': 'Another active BOM already exists for this FG'}), 409

    bom.is_deleted = False
    bom.deleted_at = None
    bom.updated_by = _cu()
    db.session.commit()
    return jsonify({'status': 'ok'})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  DROPDOWN HELPERS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _materials_by_abbr(abbreviations):
    """Return all live materials whose material_type.abbreviation is in the
    given set (case-insensitive)."""
    abbr_upper = {a.upper() for a in abbreviations}
    type_ids = [t.id for t in MaterialType.query.all()
                if (t.abbreviation or '').upper() in abbr_upper]
    if not type_ids:
        return []
    q = Material.query.filter(
        Material.material_type_id.in_(type_ids),
        db.or_(Material.is_deleted == False, Material.is_deleted == None),
    ).order_by(Material.material_name)
    return q.all()


@packing_bom_bp.route('/api/fg-list')
@login_required
def api_fg_list():
    """All Finished Goods materials for the FG dropdown."""
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    rows = _materials_by_abbr({'FG'})
    return jsonify({
        'status': 'ok',
        'items' : [{
            'id'            : m.id,
            'name'          : m.material_name,
            'code'          : m.code or '',
            'brand'         : m.brand or '',
            'category'      : m.category or '',
            'uom'           : m.uom or 'PCS',
            'per_box_qty'   : int(m.per_box_qty or 0),
            'image'         : m.image_data or '',
        } for m in rows],
        'total' : len(rows),
    })


@packing_bom_bp.route('/api/pm-list')
@login_required
def api_pm_list():
    """All PM-side materials (incl. Corrugation & Sleeves sub-categories).
    Optional ?cat=Corrugation|Sleeves|PM to filter by sub-category."""
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    rows = _materials_by_abbr({'PM'})

    cat = (request.args.get('cat') or '').strip().lower()
    if cat:
        rows = [m for m in rows if (m.category or '').lower() == cat]

    return jsonify({
        'status': 'ok',
        'items' : [{
            'id'            : m.id,
            'name'          : m.material_name,
            'code'          : m.code or '',
            'category'      : m.category or 'PM',
            'uom'           : m.uom or 'PCS',
            'brand'         : m.brand or '',
            'pm_material_type': m.pm_material_type or '',
        } for m in rows],
        'total' : len(rows),
    })


@packing_bom_bp.route('/api/uom-list')
@login_required
def api_uom_list():
    """All active UOMs from the UOM master table â€” used to populate the
    UOM dropdown in each packing-item row."""
    if not _can('view'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    try:
        q = UOMMaster.query.filter(
            db.or_(UOMMaster.is_deleted == False, UOMMaster.is_deleted == None),
            db.or_(UOMMaster.status == True,    UOMMaster.status == None),
        ).order_by(UOMMaster.code)
        rows = q.all()
    except Exception:
        # Fall back to no-filter query if the columns don't exist
        rows = UOMMaster.query.order_by(UOMMaster.code).all()
    return jsonify({
        'status': 'ok',
        'items' : [{'code': r.code, 'name': r.name} for r in rows if r.code],
        'total' : len(rows),
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  EXCEL EXPORT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _pb_openpyxl():
    """Lazy-import openpyxl (auto-install once if missing)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return openpyxl, Font, PatternFill, Alignment, Border, Side


def _safe_sheet_name(name, used):
    safe = re.sub(r'[:\\/?*\[\]]', '_', name or '')[:31] or 'BOM'
    base, n = safe, 2
    while safe in used:
        suff = f' ({n})'
        safe = (base[:31 - len(suff)] + suff)
        n += 1
    used.add(safe)
    return safe


def _pb_build_sheet(ws, bom, styles):
    """Render one PackingBOM onto an openpyxl worksheet."""
    Font, PatternFill, Alignment, Border, Side = styles
    thin    = Side(style='thin',   color='94A3B8')
    medium  = Side(style='medium', color='1E3A5F')
    b_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)

    f_title  = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
    f_sub    = Font(name='Calibri', size=12, bold=True)
    f_hdr    = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    f_norm   = Font(name='Calibri', size=11, bold=False)
    fill_t   = PatternFill('solid', fgColor='1E3A5F')
    fill_h   = PatternFill('solid', fgColor='2563EB')
    fill_lbl = PatternFill('solid', fgColor='F1F5F9')

    a_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    a_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    a_right  = Alignment(horizontal='right',  vertical='center')

    fg = bom.fg
    fg_name  = (fg.material_name if fg else '') or ''
    fg_brand = (fg.brand         if fg else '') or ''

    # â”€â”€ Title row 1 (merged A1:E1) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value     = 'HCP WELLNESS PVT. LTD.'
    c.font      = f_title
    c.alignment = a_center
    c.fill      = fill_t
    ws.row_dimensions[1].height = 28

    # â”€â”€ Subtitle row 2 (merged A2:E2) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.merge_cells('A2:E2')
    c = ws['A2']
    c.value     = 'Packing Bill of Materials'
    c.font      = f_sub
    c.alignment = a_center

    # â”€â”€ Header info rows (4â€“6) â€” FG Code intentionally omitted â”€â”€â”€
    def _kv(row, label, value):
        ws.cell(row=row, column=1, value=label).font = f_sub
        ws.cell(row=row, column=1).fill = fill_lbl
        ws.cell(row=row, column=1).border = b_thin
        ws.cell(row=row, column=1).alignment = a_left
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row,   end_column=5)
        cc = ws.cell(row=row, column=2, value=value)
        cc.font = f_norm
        cc.alignment = a_left
        for cn in range(2, 6):
            ws.cell(row=row, column=cn).border = b_thin

    _kv(4, 'FG Item:',  fg_name)
    _kv(5, 'Brand:',    fg_brand)
    _kv(6, 'FG Qty:',   f"{float(bom.fg_qty or 0):g} {bom.fg_uom or 'PCS'}")
    if bom.notes:
        _kv(7, 'Notes:', bom.notes)
        items_hdr_row = 9
    else:
        items_hdr_row = 8

    # â”€â”€ Items table header (5 columns: # | Name | Category | Qty | UOM) â”€
    headers = ['#', 'Item Name', 'Category', 'Qty', 'UOM']
    widths  = [6,    40,          18,         12,    10]
    for i, h in enumerate(headers, start=1):
        cc = ws.cell(row=items_hdr_row, column=i, value=h)
        cc.font = f_hdr; cc.fill = fill_h
        cc.alignment = a_center
        cc.border = b_thin
        ws.column_dimensions[chr(64 + i)].width = widths[i - 1]
    ws.row_dimensions[items_hdr_row].height = 24

    # â”€â”€ Item rows (A=#, B=Name, C=Category, D=Qty, E=UOM) â”€â”€â”€â”€â”€â”€â”€â”€â”€
    items = bom.items.all()
    r = items_hdr_row + 1
    first_r = r
    for it in items:
        m = it.material
        ws.cell(row=r, column=1, value=it.sr_no or 0).font = f_norm
        ws.cell(row=r, column=1).alignment = a_center
        ws.cell(row=r, column=1).border = b_thin

        ws.cell(row=r, column=2, value=(m.material_name if m else '') or it.item_name_snap or '').font = f_norm
        ws.cell(row=r, column=2).alignment = a_left
        ws.cell(row=r, column=2).border = b_thin

        ws.cell(row=r, column=3, value=(m.category if m else '') or 'PM').font = f_norm
        ws.cell(row=r, column=3).alignment = a_left
        ws.cell(row=r, column=3).border = b_thin

        ws.cell(row=r, column=4, value=float(it.qty or 0)).font = f_norm
        ws.cell(row=r, column=4).number_format = '0.000'
        ws.cell(row=r, column=4).alignment = a_right
        ws.cell(row=r, column=4).border = b_thin

        ws.cell(row=r, column=5, value=(m.uom if m else '') or it.uom_snap or 'PCS').font = f_norm
        ws.cell(row=r, column=5).alignment = a_center
        ws.cell(row=r, column=5).border = b_thin
        r += 1

    # â”€â”€ Total row â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if items:
        total_r = r
        for cn in range(1, 6):
            ws.cell(row=total_r, column=cn).fill = fill_lbl
            ws.cell(row=total_r, column=cn).border = b_thin
        cc = ws.cell(row=total_r, column=3, value='Total')
        cc.font = f_sub; cc.alignment = a_right
        cc = ws.cell(row=total_r, column=4, value=f'=SUM(D{first_r}:D{r - 1})')
        cc.font = f_sub; cc.number_format = '0.000'
        cc.alignment = a_right
    ws.freeze_panes = f'A{items_hdr_row + 1}'


def _pb_build_workbook(boms):
    """Build an .xlsx with one sheet per BOM."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _pb_openpyxl()
    styles = (Font, PatternFill, Alignment, Border, Side)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used = set()
    for b in boms:
        title = (b.fg.material_name if b.fg else f'BOM_{b.id}')
        ws = wb.create_sheet(_safe_sheet_name(title, used))
        _pb_build_sheet(ws, b, styles)
    if not wb.sheetnames:
        wb.create_sheet('Empty')
    return wb


@packing_bom_bp.route('/api/export/<int:bom_id>')
@login_required
def api_export_one(bom_id):
    if not _can('export'):
        abort(403)
    bom = PackingBOM.query.get_or_404(bom_id)
    if bom.is_deleted:
        abort(404)
    wb = _pb_build_workbook([bom])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    name = (bom.fg.material_name if bom.fg else f'BOM_{bom.id}')
    safe = re.sub(r'[^A-Za-z0-9_\-]+', '_', name).strip('_') or f'packing_bom_{bom.id}'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{safe}_packing_BOM.xlsx')


@packing_bom_bp.route('/api/export-all')
@login_required
def api_export_all():
    if not _can('export'):
        abort(403)
    boms = (PackingBOM.query
            .filter_by(is_deleted=False)
            .order_by(PackingBOM.id.desc()).all())
    if not boms:
        return jsonify({'status': 'error', 'message': 'No BOMs to export'}), 400
    wb = _pb_build_workbook(boms)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'packing_BOMs_{stamp}.xlsx')


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  EXCEL TEMPLATE  (empty workbook so users know the import format)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@packing_bom_bp.route('/api/template')
@login_required
def api_template():
    """Return a sample empty template â€” one sheet pre-filled with the
    structure parser expects (header info rows + items table)."""
    if not _can('view'):
        abort(403)
    openpyxl, Font, PatternFill, Alignment, Border, Side = _pb_openpyxl()
    styles = (Font, PatternFill, Alignment, Border, Side)

    # Build one fake BOM-shaped sheet using a stub object
    class _StubFG:
        material_name = 'Beardo Moisturising Body Lotion 200ml'
        brand = 'Beardo'
        category = 'FG'; uom = 'PCS'
        # 'code' kept for compatibility with the sheet builder, but is no
        # longer rendered in the layout
        code = ''
    class _StubItem:
        def __init__(self, sr, name, cat, qty, uom):
            self.sr_no = sr; self.material_id = None
            class _M: pass
            m = _M()
            m.material_name = name; m.code = ''; m.category = cat; m.uom = uom
            self.material = m
            self.qty = qty
            self.item_name_snap = name; self.uom_snap = uom
    class _StubBOM:
        id = 1
        fg = _StubFG()
        fg_qty = 100; fg_uom = 'PCS'; notes = 'Sample template â€” edit values & rows then import'
        _items = [
            _StubItem(1, 'Plastic Bottle 200ml',  'PM',          100, 'PCS'),
            _StubItem(2, 'Pump Cap',              'PM',          100, 'PCS'),
            _StubItem(3, 'Outer Carton 5+1',      'Corrugation',  20, 'PCS'),
            _StubItem(4, 'Front Sleeve',          'Sleeves',     100, 'PCS'),
        ]
        @property
        def items(self):
            class _Q:
                _list = _StubBOM._items
                def all(self): return self._list
            return _Q()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('SAMPLE â€” edit and re-upload')
    _pb_build_sheet(ws, _StubBOM(), styles)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='packing_BOM_template.xlsx')


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  EXCEL IMPORT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _is_pb_header_row(cells):
    """True if the first 6 cells look like 'Sr.No | Item | ... | Qty | UOM'."""
    low = [str(c or '').strip().lower() for c in cells[:6]]
    if not any(low):
        return False
    has_item = any('item' in x or 'name' in x or 'ingredient' in x for x in low[:3])
    has_qty  = any('qty' in x or 'quantity' in x for x in low)
    return has_item and has_qty


def _pb_find_items_start(ws):
    """Return the row number of the FIRST data row after the items header.
    Falls back to None if no header found."""
    for r in range(1, min(ws.max_row, 40) + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        if _is_pb_header_row(cells):
            return r + 1
    return None


def _pb_extract_meta(ws, end_row):
    """Scan rows 1..end_row to find FG name / brand / qty / notes.
    Check the more specific labels first so 'FG Qty:' is not eaten by the
    generic 'fg ...' fallback that matches the FG name."""
    meta = {'fg_name': '', 'fg_brand': '',
            'fg_qty': 0.0, 'fg_uom': 'PCS', 'notes': ''}
    for r in range(1, end_row):
        for c in range(1, 4):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            sv = str(v).strip()
            low = sv.lower().rstrip(':').strip()
            # Right-neighbour cell value
            right = None
            for nc in range(c + 1, min(c + 6, 10)):
                rv = ws.cell(row=r, column=nc).value
                if rv not in (None, ''):
                    right = rv; break
            if right is None:
                continue
            right_txt = str(right).strip()

            # Most specific labels first (FG Qty before generic FG ...)
            if (low in ('fg qty', 'qty', 'quantity', 'fg quantity')
                or low.startswith('fg qty')):
                m = re.match(r'^\s*([\d.,]+)\s*([A-Za-z]*)\s*$', right_txt)
                if m:
                    try: meta['fg_qty'] = float(m.group(1).replace(',', ''))
                    except (TypeError, ValueError): pass
                    if m.group(2): meta['fg_uom'] = m.group(2)
                else:
                    try: meta['fg_qty'] = float(right_txt)
                    except (TypeError, ValueError): pass
            elif low.startswith('brand'):
                if not meta['fg_brand']: meta['fg_brand'] = right_txt
            elif low.startswith('notes') or low.startswith('note'):
                if not meta['notes']: meta['notes'] = right_txt
            elif (low in ('fg item', 'fg name', 'fg', 'product name',
                          'item name', 'item', 'product')
                  or low.startswith('fg item') or low.startswith('fg name')
                  or low.startswith('product name')):
                if not meta['fg_name']: meta['fg_name'] = right_txt
    return meta


def _pb_parse_sheet(ws):
    """Return (meta_dict, list of item_dicts)."""
    start = _pb_find_items_start(ws)
    if start is None:
        return {}, []
    meta = _pb_extract_meta(ws, start - 1)

    rows_out = []
    seen_any = False
    for r in range(start, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        col_b = ws.cell(row=r, column=2).value

        # Stop on empty col A once we've seen at least one row
        if col_a is None or (isinstance(col_a, str) and not col_a.strip()):
            if seen_any: break
            continue

        # Skip a 'Total' row
        if isinstance(col_a, str) and col_a.strip().lower() in ('total', 'totals'):
            continue
        if isinstance(col_b, str) and col_b.strip().lower() in ('total', 'totals'):
            continue

        nm = ws.cell(row=r, column=2).value
        if nm is None or (isinstance(nm, str) and not str(nm).strip()):
            continue

        # Layout used by our exporter:
        #   col A=#, B=Item Name, C=Category, D=Qty, E=UOM
        seen_any = True
        try:
            sr = int(col_a)
        except (ValueError, TypeError):
            sr = len(rows_out) + 1
        cat  = ws.cell(row=r, column=3).value
        qty  = ws.cell(row=r, column=4).value
        uom  = ws.cell(row=r, column=5).value
        try:    qty_f = float(qty) if qty not in (None, '') else 0.0
        except (TypeError, ValueError): qty_f = 0.0

        rows_out.append({
            'sr_no'    : sr,
            'item_name': str(nm).strip(),
            'category' : str(cat).strip()  if cat  else '',
            'qty'      : qty_f,
            'uom'      : str(uom).strip()  if uom  else '',
        })
    return meta, rows_out


def _pb_lookup_fg(meta):
    """Resolve FG material by name."""
    name = (meta.get('fg_name') or '').strip()
    if not name:
        return None
    m = Material.query.filter(
        db.func.lower(Material.material_name) == name.lower(),
        db.or_(Material.is_deleted == False, Material.is_deleted == None),
    ).first()
    return m


def _pb_lookup_item(row):
    """Resolve a PM material by name. If multiple matches and a Category
    is given, prefer the one matching that category."""
    name = (row.get('item_name') or '').strip()
    cat  = (row.get('category')  or '').strip().lower()
    if not name:
        return None
    matches = Material.query.filter(
        db.func.lower(Material.material_name) == name.lower(),
        db.or_(Material.is_deleted == False, Material.is_deleted == None),
    ).all()
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]
    # Multiple â€” prefer one whose category matches
    if cat:
        for m in matches:
            if (m.category or '').lower() == cat:
                return m
    return matches[0]


@packing_bom_bp.route('/api/import/preview', methods=['POST'])
@login_required
def api_import_preview():
    """Upload an .xlsx; return one preview block per sheet."""
    if not _can('add'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    f = request.files.get('file')
    if not f:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
    fname = (f.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xlsm', '.xls')):
        return jsonify({'status': 'error',
                        'message': 'Only .xlsx / .xlsm / .xls allowed'}), 400

    openpyxl, *_ = _pb_openpyxl()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Cannot read file: {e}'}), 400

    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            meta, items = _pb_parse_sheet(ws)
        except Exception as e:
            sheets.append({'sheet_name': sheet_name, 'valid': False,
                           'error': f'Parse error: {e}',
                           'items': [], 'item_count': 0})
            continue

        fg = _pb_lookup_fg(meta)
        existing = None
        if fg:
            existing = (PackingBOM.query
                        .filter_by(fg_material_id=fg.id, is_deleted=False)
                        .first())

        # Resolve each item & flag unmatched
        resolved = []
        unmatched = 0
        for it in items:
            m = _pb_lookup_item(it)
            resolved.append({
                **it,
                'material_id'    : m.id if m else None,
                'resolved_name'  : m.material_name if m else '',
                'resolved_cat'   : (m.category if m else '') or it.get('category', ''),
                'resolved_uom'   : (m.uom if m else '') or it.get('uom', ''),
                'matched'        : bool(m),
            })
            if not m: unmatched += 1

        errors = []
        if not fg:           errors.append('FG item not found in master')
        if not items:        errors.append('No item rows detected')
        if unmatched:        errors.append(f'{unmatched} item(s) not matched to master')

        sheets.append({
            'sheet_name'        : sheet_name,
            'fg_name'           : meta.get('fg_name', ''),
            'fg_brand'          : meta.get('fg_brand', ''),
            'fg_qty'            : meta.get('fg_qty', 0),
            'fg_uom'            : meta.get('fg_uom', 'PCS'),
            'notes'             : meta.get('notes', ''),
            'matched_fg_id'     : fg.id   if fg else None,
            'matched_fg_name'   : fg.material_name if fg else '',
            'already_exists'    : bool(existing),
            'existing_id'       : existing.id if existing else None,
            'items'             : resolved,
            'item_count'        : len(resolved),
            'unmatched_count'   : unmatched,
            'valid'             : bool(fg) and bool(items) and unmatched == 0,
            'warnings'          : errors,
        })

    return jsonify({
        'status'         : 'ok',
        'file_name'      : f.filename,
        'sheets'         : sheets,
        'total_sheets'   : len(sheets),
        'valid_sheets'   : sum(1 for s in sheets if s.get('valid')),
        'invalid_sheets' : sum(1 for s in sheets if not s.get('valid')),
    })


@packing_bom_bp.route('/api/import/commit', methods=['POST'])
@login_required
def api_import_commit():
    """Persist the sheets the user selected from the preview."""
    if not _can('add'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d = request.get_json(silent=True) or {}
    sheets = d.get('sheets') or []
    if not sheets:
        return jsonify({'status': 'error', 'message': 'No sheets to import'}), 400

    created, updated, skipped = [], [], []
    for s in sheets:
        fg_id = s.get('matched_fg_id')
        sheet_name = s.get('sheet_name', '?')
        if not fg_id:
            skipped.append({'sheet': sheet_name, 'reason': 'FG not matched'})
            continue
        items = s.get('items') or []
        # Drop unmatched item rows
        items = [it for it in items if it.get('material_id')]
        if not items:
            skipped.append({'sheet': sheet_name, 'reason': 'No matched items'})
            continue

        existing = (PackingBOM.query
                    .filter_by(fg_material_id=fg_id, is_deleted=False)
                    .first())
        is_new = existing is None
        if is_new:
            if not _can('add'):
                skipped.append({'sheet': sheet_name, 'reason': 'No add permission'})
                continue
            bom = PackingBOM(fg_material_id=fg_id, created_by=_cu())
            db.session.add(bom)
        else:
            if not _can('edit'):
                skipped.append({'sheet': sheet_name, 'reason': 'No edit permission (would overwrite)'})
                continue
            bom = existing

        bom.fg_qty     = _num(s.get('fg_qty'), 0)
        bom.fg_uom     = (s.get('fg_uom') or 'PCS').strip() or 'PCS'
        bom.notes      = (s.get('notes')  or '').strip()
        bom.updated_by = _cu()
        bom.is_active  = True
        bom.is_deleted = False
        bom.deleted_at = None

        # Replace items
        if not is_new:
            for old in bom.items.all():
                db.session.delete(old)
            db.session.flush()
        db.session.flush()

        sr = 0
        for it in items:
            mid = it.get('material_id')
            qty = _num(it.get('qty'), 0)
            if not mid: continue
            m = Material.query.get(mid)
            if not m: continue
            sr += 1
            db.session.add(PackingBOMItem(
                packing_bom_id = bom.id,
                sr_no          = it.get('sr_no') or sr,
                material_id    = mid,
                qty            = qty,
                item_name_snap = m.material_name or '',
                uom_snap       = (it.get('uom') or m.uom or 'PCS'),
            ))

        (created if is_new else updated).append({
            'sheet': sheet_name, 'id': bom.id,
            'fg': m.material_name if m else '', 'rows': sr,
        })

    db.session.commit()
    return jsonify({
        'status' : 'ok',
        'created': created,
        'updated': updated,
        'skipped': skipped,
    })


