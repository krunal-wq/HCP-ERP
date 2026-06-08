"""
salary_routes.py — Salary Processing + Payslip PDF
Blueprint: salary_bp
  GET  /hr/salary                      → process screen (month-wise preview)
  POST /hr/salary/process              → calculate & save slips for the month
  GET  /hr/salary/slip/<emp>/<y>/<m>   → payslip PDF download (HCP format)

Calculation (attendance-linked):
  payable_days = month_days − LOP
  LOP = Absent + MIS-PUNCH (policy: salary_config key 'mp_treatment'
        = absent / half_day / present; default half_day)
  Earned = Actual × payable/month_days  (incentive full)
  Deductions = PF + ESIC + PT + TDS (employee master se)
"""
import calendar
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_file, abort)
from flask_login import login_required, current_user

from models import db, Employee
from models.payroll import SalarySlip
from models.employee import SalaryConfig
from models.hr_rules import HRLeaveApplication

salary_bp = Blueprint('salary', __name__)

COMPANY_NAME = 'HCP WELLNESS PRIVATE LIMITED'
COMPANY_ADDR = ('403, Maruti Elanza Vertex, Opp. GTPL House, B/h. Armeida,\n'
                'Sindhu Bhavan Road, Bodakdev, Ahmedabad - 380054 Gujarat, India')


# ─────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────
def _role_ok():
    return current_user.is_authenticated and \
           (current_user.role or '').lower() in ('admin', 'hr', 'manager')


def _mp_treatment():
    row = SalaryConfig.query.filter_by(key='mp_treatment').first()
    v = (row.value if row else 'half_day').strip().lower()
    return v if v in ('absent', 'half_day', 'present') else 'half_day'


def _f(x):
    try:
        return float(x or 0)
    except (TypeError, ValueError):
        return 0.0


def _month_leaves(year, month, month_start, month_end):
    """Employee-wise PL/CL/SL taken (month ke andar clipped days)."""
    apps = HRLeaveApplication.query.filter(
        HRLeaveApplication.status == 'approved',
        HRLeaveApplication.from_date <= month_end,
        HRLeaveApplication.to_date >= month_start,
    ).all()
    out = {}
    for l in apps:
        f = max(l.from_date, month_start)
        t = min(l.to_date, month_end)
        d = (t - f).days + 1
        if l.half_day and l.from_date == l.to_date:
            d = 0.5
        rec = out.setdefault(l.employee_id, {'PL': 0.0, 'CL': 0.0, 'SL': 0.0})
        rec[l.leave_type if l.leave_type in rec else 'PL'] += d
    return out


def _ensure_structure(emp):
    """Agar salary structure (basic etc.) missing hai lekin gross set hai,
    to ERP ka existing HCP auto-apply chala do."""
    if _f(emp.salary_basic) > 0 or _f(emp.salary_gross) <= 0:
        return False
    try:
        from modules.hr.routes.hr_routes import _apply_hcp_structure
        return bool(_apply_hcp_structure(emp))
    except Exception:
        return False


def _late_rules_map():
    from models.attendance import LateShiftRule
    return {r.employee_type: r for r in
            LateShiftRule.query.filter_by(is_active=True).all()}


def _late_info(rule, days):
    """Late days count + penalty (type ke slabs se).
    Returns (late_count, penalty_amount)."""
    if rule is None:
        return 0, 0.0
    band_end, r1, r2, r3 = '10:59', 120.0, 250.0, 250.0
    for s in rule.penalty_rules:
        if not s.is_active:
            continue
        if s.time_to:
            band_end = s.time_to
            if s.from_count <= 3 <= (s.to_count or 999):
                r1 = float(s.penalty_amount)
            if s.from_count >= 4:
                r2 = float(s.penalty_amount)
        else:
            r3 = float(s.penalty_amount)
    band_n = after_n = 0
    for dd in days:
        it = dd.get('in_t')
        if it and it > rule.late_after:
            if it <= band_end:
                band_n += 1
            else:
                after_n += 1
    penalty = min(band_n, 3) * r1 + max(band_n - 3, 0) * r2 + after_n * r3
    return band_n + after_n, round(penalty, 2)


def _compute_salary(emp, cnt, month_days, mp_treat, leaves, late=(0, 0.0)):
    """Ek employee ka pura calculation. cnt = monthly status counts."""
    _ensure_structure(emp)
    ab = cnt.get('AB', 0)
    mp = cnt.get('MP', 0)
    hd = cnt.get('HD', 0)
    if mp_treat == 'absent':
        mp_lop = mp
    elif mp_treat == 'present':
        mp_lop = 0
    else:
        mp_lop = mp * 0.5
    lop = ab + mp_lop + hd * 0.5
    lop = min(lop, month_days)
    payable = month_days - lop
    factor = (payable / month_days) if month_days else 0

    basic, hra  = _f(emp.salary_basic), _f(emp.salary_hra)
    conv        = _f(emp.salary_conveyance)
    medical     = _f(emp.salary_medical_allow)
    special     = _f(emp.salary_special_allow)
    incentive   = _f(emp.salary_incentive)

    e = lambda v: round(v * factor, 2)
    earned = dict(basic=e(basic), hra=e(hra), conv=e(conv),
                  medical=e(medical), special=e(special))
    total_actual = round(basic + hra + conv + medical + special, 2)

    # ── WOP / HLP = off-day par kaam → extra day ka paisa ──
    wop = cnt.get('WOP', 0)
    hlp = cnt.get('HLP', 0)
    per_day = (total_actual / month_days) if month_days else 0
    extra_pay = round(per_day * (wop + hlp), 2)

    total_earned = round(sum(earned.values()) + incentive + extra_pay, 2)

    pf   = _f(emp.salary_pf_employee)
    esic = _f(emp.salary_esic_employee)
    pt   = _f(emp.salary_professional_tax) if payable > 0 else 0
    tds  = _f(emp.salary_tds)
    late_n, late_pen = late
    if total_earned <= 0:
        # structure/gross hi nahi (jaise contractor) — penalty se negative mat banao
        late_pen = 0.0
    total_ded = round(pf + esic + pt + tds + late_pen, 2)
    net = round(total_earned - total_ded)

    lv = leaves.get(emp.id, {'PL': 0, 'CL': 0, 'SL': 0})
    return dict(
        month_days=month_days,
        present=cnt.get('P', 0), half=hd,
        absent=ab, mispunch=mp,
        weekoff=cnt.get('WO', 0), holiday=cnt.get('HOL', 0),
        leave=cnt.get('LV', 0), lop=round(lop, 1), payable=round(payable, 1),
        late_n=late_n, late_pen=late_pen,
        wop=wop, hlp=hlp, extra_pay=extra_pay,
        lv_pl=lv['PL'], lv_cl=lv['CL'], lv_sl=lv['SL'],
        basic=basic, hra=hra, conv=conv, medical=medical, special=special,
        earned=earned, incentive=incentive,
        total_actual=total_actual, total_earned=total_earned,
        pf=pf, esic=esic, pt=pt, tds=tds, late=late_pen,
        total_ded=total_ded, net=net,
    )


def _payroll_rows(emp_rows):
    """Salary list se Contractors aur Admin account hatao."""
    out = []
    for er in emp_rows:
        emp = er['emp']
        etype = (emp.employee_type or '').upper()
        code  = (emp.employee_code or emp.employee_id or '').upper()
        if 'CONTRACTOR' in etype:
            continue
        if getattr(emp, 'is_contractor', False):
            continue
        if code == 'ADMIN':
            continue
        out.append(er)
    return out


def _get_month_data(year, month, emp_type, search_q):
    """Attendance monthly builder reuse — same data, same logic."""
    from modules.hr.routes.attendance_routes import _build_monthly_data
    return _build_monthly_data(year, month, emp_type, search_q)


# ─────────────────────────────────────────────────────────────────
#  GET /hr/salary — Process screen
# ─────────────────────────────────────────────────────────────────
@salary_bp.route('/hr/salary')
@login_required
def salary_process_screen():
    if not _role_ok():
        flash('Access denied: Salary permission nahi hai.', 'error')
        return redirect('/')

    today = date.today()
    # default = pichla complete month
    d_y, d_m = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)
    try:
        year  = int(request.args.get('year',  d_y))
        month = int(request.args.get('month', d_m))
    except ValueError:
        year, month = d_y, d_m
    if not (1 <= month <= 12):
        year, month = d_y, d_m
    emp_type = (request.args.get('emp_type') or '').strip()
    search_q = (request.args.get('q') or '').strip().lower()

    month_days  = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end   = date(year, month, month_days)
    mp_treat    = _mp_treatment()

    md     = _get_month_data(year, month, emp_type, search_q)
    leaves = _month_leaves(year, month, month_start, month_end)

    existing = {s.employee_id: s for s in SalarySlip.query.filter_by(
        year=year, month=month).all()}

    lrules = _late_rules_map()
    rows, mp_total = [], 0
    for er in _payroll_rows(md['emp_rows']):
        emp = er['emp']
        late = _late_info(lrules.get(emp.employee_type or ''), er['days'])
        calc = _compute_salary(emp, er['cnt'], month_days, mp_treat, leaves, late)
        mp_total += calc['mispunch']
        rows.append(dict(emp=emp, c=calc, slip=existing.get(emp.id)))

    return render_template('hr/salary/process.html',
        rows=rows, year=year, month=month,
        month_name=calendar.month_name[month],
        emp_types=[t for t in md['emp_types'] if 'CONTRACTOR' not in t.upper()],
        emp_type=emp_type, search_q=search_q,
        mp_total=mp_total, mp_treat=mp_treat,
        processed_count=len(existing),
        years=list(range(today.year - 2, today.year + 1)),
        active_page='hr_salary')


# ─────────────────────────────────────────────────────────────────
#  POST /hr/salary/process — calculate & save
# ─────────────────────────────────────────────────────────────────
@salary_bp.route('/hr/salary/process', methods=['POST'])
@login_required
def salary_process():
    if not _role_ok():
        flash('Access denied.', 'error')
        return redirect('/')
    try:
        year  = int(request.form.get('year'))
        month = int(request.form.get('month'))
    except (TypeError, ValueError):
        flash('Invalid month.', 'error')
        return redirect(url_for('salary.salary_process_screen'))
    emp_type = (request.form.get('emp_type') or '').strip()
    search_q = (request.form.get('q') or '').strip().lower()

    month_days  = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end   = date(year, month, month_days)
    mp_treat    = _mp_treatment()

    md     = _get_month_data(year, month, emp_type, search_q)
    leaves = _month_leaves(year, month, month_start, month_end)
    existing = {s.employee_id: s for s in SalarySlip.query.filter_by(
        year=year, month=month).all()}

    lrules = _late_rules_map()
    n_new = n_upd = 0
    for er in _payroll_rows(md['emp_rows']):
        emp = er['emp']
        late = _late_info(lrules.get(emp.employee_type or ''), er['days'])
        c = _compute_salary(emp, er['cnt'], month_days, mp_treat, leaves, late)
        slip = existing.get(emp.id)
        if slip is None:
            slip = SalarySlip(employee_id=emp.id, year=year, month=month,
                              created_by=current_user.id)
            db.session.add(slip)
            n_new += 1
        else:
            n_upd += 1
        slip.month_days    = month_days
        slip.present_days  = c['present'];  slip.half_days     = c['half']
        slip.absent_days   = c['absent'];   slip.mispunch_days = c['mispunch']
        slip.weekoff_days  = c['weekoff'];  slip.holiday_days  = c['holiday']
        slip.leave_days    = c['leave'];    slip.lop_days      = c['lop']
        slip.late_days     = c['late_n'];   slip.ded_late      = c['late_pen']
        slip.wop_days      = c['wop'];      slip.hlp_days      = c['hlp']
        slip.extra_earned  = c['extra_pay']
        slip.worked_days   = c['payable']
        slip.leave_taken_pl = c['lv_pl']; slip.leave_taken_cl = c['lv_cl']
        slip.leave_taken_sl = c['lv_sl']
        slip.basic_actual   = c['basic'];   slip.basic_earned   = c['earned']['basic']
        slip.hra_actual     = c['hra'];     slip.hra_earned     = c['earned']['hra']
        slip.conv_actual    = c['conv'];    slip.conv_earned    = c['earned']['conv']
        slip.medical_actual = c['medical']; slip.medical_earned = c['earned']['medical']
        slip.special_actual = c['special']; slip.special_earned = c['earned']['special']
        slip.incentive_earned = c['incentive']
        slip.total_actual   = c['total_actual']
        slip.total_earned   = c['total_earned']
        slip.ded_pf  = c['pf'];  slip.ded_esic = c['esic']
        slip.ded_pt  = c['pt'];  slip.ded_tds  = c['tds']
        slip.total_deductions = c['total_ded']
        slip.net_pay = c['net']
        slip.status  = 'processed'
        slip.updated_at = datetime.now()
    db.session.commit()
    flash(f'✅ Salary processed — {n_new} new, {n_upd} updated '
          f'({calendar.month_name[month]} {year})', 'success')
    return redirect(url_for('salary.salary_process_screen',
                            year=year, month=month, emp_type=emp_type, q=search_q))


# ─────────────────────────────────────────────────────────────────
#  Number → Indian words
# ─────────────────────────────────────────────────────────────────
_ONES = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight',
         'Nine', 'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen',
         'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
_TENS = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy',
         'Eighty', 'Ninety']


def _two(n):
    return _ONES[n] if n < 20 else (_TENS[n // 10] + (' ' + _ONES[n % 10] if n % 10 else ''))


def _three(n):
    s = ''
    if n >= 100:
        s = _ONES[n // 100] + ' Hundred'
        if n % 100: s += ' '
    return s + _two(n % 100) if n % 100 else s


def amount_in_words(n):
    n = int(round(n))
    if n == 0: return 'Zero Only'
    parts = []
    for div, name in ((10000000, 'Crore'), (100000, 'Lakh'), (1000, 'Thousand')):
        if n >= div:
            parts.append(_three(n // div) + ' ' + name)
            n %= div
    if n:
        parts.append(_three(n))
    return ' '.join(parts) + ' Only'


# ─────────────────────────────────────────────────────────────────
#  GET /hr/salary/slip/<emp_id>/<year>/<month> — Payslip PDF
# ─────────────────────────────────────────────────────────────────
@salary_bp.route('/hr/salary/slip/<int:emp_id>/<int:year>/<int:month>')
@login_required
def salary_slip_pdf(emp_id, year, month):
    """HCP 'Form IV B' salary slip - HTML print view (browser Print -> Save PDF)."""
    if not _role_ok():
        flash('Access denied.', 'error')
        return redirect('/')
    slip = SalarySlip.query.filter_by(employee_id=emp_id, year=year,
                                      month=month).first()
    if slip is None:
        flash('Slip abhi process nahi hui - pehle Process Salary karo.', 'warning')
        return redirect(url_for('salary.salary_process_screen',
                                year=year, month=month))
    emp = slip.employee or Employee.query.get_or_404(emp_id)

    g = _f
    fmt = lambda v: f'{g(v):,.2f}'

    # ---- Earnings (full entitlement; LOP ab deduction side me jayega) ----
    # extra-day (WO/Holiday) + overtime pay -> Production Incentive
    incentive_pay = g(slip.incentive_earned) + g(getattr(slip, 'extra_earned', 0))
    earnings = [
        ('Basic + DA',           g(slip.basic_actual),   g(slip.basic_actual)),
        ('HRA',                  g(slip.hra_actual),     g(slip.hra_actual)),
        ('CONV',                 g(slip.conv_actual),    g(slip.conv_actual)),
        ('MEDICAL',              g(slip.medical_actual), g(slip.medical_actual)),
        ('Other Allowance',      g(slip.special_actual), g(slip.special_actual)),
        ('Production Incentive', None,                   incentive_pay),
        ('Arrear',               None,                   g(slip.arrears_earned)),
        ('Bonus',                None,                   g(getattr(slip, 'bonus_earned', 0))),
    ]
    gross_actual  = sum(a for _, a, _ in earnings if a is not None)
    gross_payable = sum(p for _, _, p in earnings if p is not None)

    # ---- LOP (leave) amount = per-day rate x LOP days ----
    md = g(slip.month_days) or 30
    struct_total = (g(slip.basic_actual) + g(slip.hra_actual) + g(slip.conv_actual)
                    + g(slip.medical_actual) + g(slip.special_actual))
    lop_amount = round(struct_total / md * g(slip.lop_days), 2) if md else 0.0

    # ---- Challan (fine) + Loan EMI module se ----
    from modules.hr.routes.challan_routes import challan_total_for
    from modules.hr.routes.loan_routes import loan_emi_for
    challan_total = challan_total_for(emp.id, year, month)
    loan_emi      = loan_emi_for(emp.id, year, month)

    # ---- Deductions: LOP + late-coming + challan + others -> Other Deduction ----
    other_ded = (g(slip.ded_others) + g(getattr(slip, 'ded_late', 0))
                 + lop_amount + challan_total)
    deductions = [
        ('PF',              g(slip.ded_pf)),
        ('ESIC',            g(slip.ded_esic)),
        ('PT',              g(slip.ded_pt)),
        ('LWF',             g(slip.ded_lwf)),
        ('TDS',             g(slip.ded_tds)),
        ('Loan EMI',        loan_emi + g(slip.ded_advance)),
        ('Other Deduction', other_ded),
    ]
    total_ded = sum(v for _, v in deductions)
    net = int(round(gross_payable - total_ded))

    # ---- Working details ----
    work_total = g(slip.month_days)
    working = [
        ('Working Days', g(slip.month_days) - g(slip.weekoff_days) - g(slip.holiday_days)),
        ('Weekoff',      g(slip.weekoff_days)),
        ('Pay Holiday',  g(slip.holiday_days)),
        ('Present Days', g(slip.present_days)),
        ('CL',           g(slip.leave_taken_cl)),
        ('PL',           g(slip.leave_taken_pl)),
        ('SL',           g(slip.leave_taken_sl)),
        ('ML/Adj',       g(slip.half_days)),
        ('LWP',          g(slip.lop_days)),
    ]

    mabbr = calendar.month_abbr[month]
    doj = emp.date_of_joining.strftime('%d-%b-%y') if emp.date_of_joining else '-'

    return render_template(
        'hr/salary/slip.html',
        emp=emp, slip=slip, year=year, month=month,
        month_label=f'{mabbr}- {year}',
        company_name=COMPANY_NAME, company_addr=COMPANY_ADDR,
        working=working, work_total=work_total,
        earnings=earnings, gross_actual=gross_actual, gross_payable=gross_payable,
        deductions=deductions, total_ded=total_ded,
        net=net, net_words=amount_in_words(net),
        doj=doj, fmt=fmt,
        period_label=f'{calendar.month_name[month]} {year}',
        worked_days=g(slip.worked_days), lop_days=g(slip.lop_days),
        lv_bal=(f"PL - {g(emp.paid_leave_balance):g}  "
                f"CL - {g(emp.casual_leave_balance):g}  "
                f"SL - {g(emp.sick_leave_balance):g}"),
        lv_taken=(f"PL - {g(slip.leave_taken_pl):g}  "
                  f"CL - {g(slip.leave_taken_cl):g}  "
                  f"SL - {g(slip.leave_taken_sl):g}"),
    )


# ══════════════════════════════════════════════════════════════
#  SALARY REGISTER — Excel export (Sonali format)
#  GET /hr/salary/register/<year>/<month>
#  Attendance grid + HCP earning/deduction breakup. TDS blank.
# ══════════════════════════════════════════════════════════════
@salary_bp.route('/hr/salary/register/<int:year>/<int:month>')
@login_required
def salary_register_export(year, month):
    if not _role_ok():
        flash('Access denied.', 'error')
        return redirect('/')

    import calendar as _cal
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    from modules.hr.routes.attendance_routes import _build_monthly_data
    from modules.hr.routes.challan_routes import challan_total_for
    from modules.hr.routes.loan_routes import loan_emi_for

    g = _f
    md = _build_monthly_data(year, month, '', '')
    emp_rows = _payroll_rows(md['emp_rows'])      # contractor/admin hatao
    day_list = md['day_list']
    n_days = len(day_list)
    month_name = _cal.month_name[month]
    mdays = _cal.monthrange(year, month)[1]

    # slips is month ke (emp_id -> slip)
    slips = {s.employee_id: s for s in
             SalarySlip.query.filter_by(year=year, month=month).all()}

    F = lambda h: PatternFill('solid', start_color=h)
    ctr = Alignment(horizontal='center', vertical='center')
    ctrw = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='B0B0B0')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    TITLE = F('F4B084'); HDR = F('D9E1F2'); SUMM = F('DEEAF1')
    ACT = F('FFF2CC'); EARN = F('E2EFDA'); DED = F('FCE4D6'); NET = F('C9DAF0')
    IN_F = F('E2EFDA'); OUT_F = F('FCE4D6'); HRS_F = F('EBF3FB'); OT_F = F('FFF2CC')

    def _ot_for(total):
        """OT = round(hours) - 8, sirf jab rounded >= 10 (warna 0).
        9:50 -> 10 -> OT 2 ; 8:55 -> 9 -> 0 ; 11:18 -> 11 -> OT 3."""
        if not total:
            return 0
        rh = int(float(total) + 0.5)        # round half up
        return rh - 8 if rh >= 10 else 0
    ST = {'P':('00B050','FFFFFF'),'WO':('BFBFBF','000000'),'WOP':('FFFF00','000000'),
          'HD':('00B0F0','FFFFFF'),'AB':('FF0000','FFFFFF'),'MP':('FFC000','000000'),
          'PH':('ED7D31','FFFFFF'),'HLP':('E11D48','FFFFFF'),'CL':('0070C0','FFFFFF'),
          'SL':('7030A0','FFFFFF'),'PL':('0070C0','FFFFFF')}

    ATT_HDRS = ['Present','Paid Holiday','Absent','Half Day','PL','CL','SL',
                'Week Off','WO Present','Holiday Present']
    ATT_CODE = ['P','PH','AB','HD','PL','CL','SL','WO','WOP','HLP']
    SAL = [('Actual Days',None),('Paid Days',None),('OT Hrs',None),('OT Days',None),
           ('Gross/Month',ACT),('Act Basic',ACT),('Act HRA',ACT),('Act Conv',ACT),
           ('Act Medical',ACT),('Act Other',ACT),
           ('Earn Basic',EARN),('Earn HRA',EARN),('Earn Conv',EARN),('Earn Medical',EARN),
           ('Earn Other',EARN),('Prod Incentive',EARN),
           ('Gross Earning',EARN),('Arrear',EARN),('Total Earned',EARN),
           ('PF',DED),('ESIC',DED),('PT',DED),('LWF',DED),('Advance/Loan',DED),
           ('Other Ded',DED),('TDS',DED),('Paid Salary',NET)]

    wb = Workbook(); ws = wb.active; ws.title = 'Salary Register'
    FIX = 7
    total_cols = FIX + n_days + len(ATT_HDRS) + len(SAL)

    titles = ['HCP Wellness Pvt. Ltd.',
              f'SALARY REGISTER FOR {month_name.upper()} {year}',
              'Plot No. 8, Ozone Industrial Park, Bavla-Bagodara Highway, Bhayla, Ahmedabad']
    for r, txt in enumerate(titles, start=1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(bold=True); c.fill = TITLE; c.alignment = ctr

    fixed = ['Sr.','Emp ID','Period','Employee Name','Designation','Department','Remarks']
    for c, h in enumerate(fixed, start=1):
        ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HDR; cell.alignment = ctr; cell.border = bord; cell.font = Font(bold=True)
    for i, d in enumerate(day_list):
        c1 = ws.cell(row=4, column=FIX+1+i, value=d.day)
        c2 = ws.cell(row=5, column=FIX+1+i, value=d.strftime('%a'))
        for cc in (c1, c2):
            cc.fill = HDR; cc.alignment = ctr; cc.border = bord
    base_att = FIX + n_days
    for j, h in enumerate(ATT_HDRS):
        col = base_att + 1 + j
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = SUMM; cell.alignment = ctrw; cell.border = bord; cell.font = Font(bold=True, size=9)
    base_sal = base_att + len(ATT_HDRS)
    for j, (h, fill) in enumerate(SAL):
        col = base_sal + 1 + j
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = fill or SUMM; cell.alignment = ctrw; cell.border = bord; cell.font = Font(bold=True, size=9)
    ws.row_dimensions[4].height = 26

    def daycode(dd):
        if dd['code'] == 'HOL': return 'PH'
        if dd['code'] == 'LV':  return dd.get('lv_code') or 'PL'
        return dd['code']

    r = 6
    for sr, er in enumerate(emp_rows, start=1):
        emp = er['emp']; rs, re_ = r, r + 4
        slip = slips.get(emp.id)
        fixed_vals = [sr, emp.employee_code or emp.employee_id or '',
                      'Probation' if getattr(emp, 'is_probation', False) else 'Permanent',
                      emp.full_name, emp.designation or '', emp.department or '']
        for c, v in enumerate(fixed_vals, start=1):
            ws.merge_cells(start_row=rs, start_column=c, end_row=re_, end_column=c)
            cell = ws.cell(row=rs, column=c, value=v); cell.alignment = ctrw; cell.border = bord
        for rr, (lbl, fill) in zip(range(rs, re_+1),
                [('In Time', IN_F), ('Out Time', OUT_F), ('Hours', HRS_F),
                 ('OT', OT_F), ('Status', None)]):
            gcell = ws.cell(row=rr, column=7, value=lbl); gcell.border = bord
            if fill: gcell.fill = fill

        ot_hrs = 0
        for i, dd in enumerate(er['days']):
            col = FIX + 1 + i
            cin = ws.cell(row=rs, column=col, value=dd['in_t'] or None)
            cout = ws.cell(row=rs+1, column=col, value=dd['out_t'] or None)
            chr_ = ws.cell(row=rs+2, column=col, value=round(dd['tot'], 2) if dd['tot'] else 0)
            ot = _ot_for(dd['tot'])
            cot = ws.cell(row=rs+3, column=col, value=ot)
            code = daycode(dd)
            cst = ws.cell(row=rs+4, column=col, value=code)
            cin.fill, cout.fill, chr_.fill, cot.fill = IN_F, OUT_F, HRS_F, OT_F
            bg, fg = ST.get(code, ('FFFFFF', '000000'))
            cst.fill = F(bg); cst.font = Font(bold=True, color=fg)
            for cc in (cin, cout, chr_, cot, cst):
                cc.alignment = ctr; cc.border = bord
            ot_hrs += ot

        # attendance summary (COUNTIF on status row)
        d1 = get_column_letter(FIX+1); d2 = get_column_letter(FIX+n_days)
        strow = rs + 4
        for j, code in enumerate(ATT_CODE):
            col = base_att + 1 + j
            ws.merge_cells(start_row=rs, start_column=col, end_row=re_, end_column=col)
            cell = ws.cell(row=rs, column=col, value=f'=COUNTIF({d1}{strow}:{d2}{strow},"{code}")')
            cell.fill = SUMM; cell.alignment = ctr; cell.border = bord

        # salary breakup (slip se; reduced/earned model — reference jaisa)
        if slip:
            ba,ha,ca,ma,oa = g(slip.basic_actual),g(slip.hra_actual),g(slip.conv_actual),g(slip.medical_actual),g(slip.special_actual)
            be,he,ce,me_,oe = g(slip.basic_earned),g(slip.hra_earned),g(slip.conv_earned),g(slip.medical_earned),g(slip.special_earned)
            inc = g(slip.incentive_earned) + g(getattr(slip,'extra_earned',0))
            gm = round(ba+ha+ca+ma+oa, 2)
            gross_earn = round(be+he+ce+me_+oe+inc, 2)
            arr = g(slip.arrears_earned)
            tot_earned = round(gross_earn + arr, 2)
            pf,esic,pt,lwf = g(slip.ded_pf),g(slip.ded_esic),g(slip.ded_pt),g(slip.ded_lwf)
            advance = round(g(slip.ded_advance) + loan_emi_for(emp.id, year, month), 2)
            other = round(g(slip.ded_others) + g(getattr(slip,'ded_late',0)) + challan_total_for(emp.id, year, month), 2)
            tot_ded = round(pf+esic+pt+lwf+advance+other, 2)
            paid_days = round(g(slip.month_days) - g(slip.lop_days), 2) or mdays
            # Paid Salary = Total Earned - SUM(PF..TDS) as FORMULA, taaki manual
            # TDS daalte hi auto minus ho jaye. (Total Earned idx18, PF idx19, TDS idx25)
            te_L  = get_column_letter(base_sal + 19)
            pf_L  = get_column_letter(base_sal + 20)
            tds_L = get_column_letter(base_sal + 26)
            paid_f = f'={te_L}{rs}-SUM({pf_L}{rs}:{tds_L}{rs})'
            sal_vals = [mdays, paid_days, round(ot_hrs,1), round(ot_hrs/8,2),
                        gm, ba, ha, ca, ma, oa,
                        be, he, ce, me_, oe, inc,
                        gross_earn, arr, tot_earned,
                        pf, esic, pt, lwf, advance, other, None, paid_f]
        else:
            sal_vals = [mdays, None, round(ot_hrs,1), round(ot_hrs/8,2)] + [None]*23

        for j, v in enumerate(sal_vals):
            col = base_sal + 1 + j
            ws.merge_cells(start_row=rs, start_column=col, end_row=re_, end_column=col)
            cell = ws.cell(row=rs, column=col, value=v)
            cell.fill = SAL[j][1] or SUMM; cell.alignment = ctr; cell.border = bord
            if SAL[j][0] in ('Total Earned', 'Paid Salary'):
                cell.font = Font(bold=True)
            if isinstance(v, float) or (isinstance(v, str) and v.startswith('=')):
                cell.number_format = '#,##0.00'
        r += 5

    for col, w in zip('ABCDEFG', [4, 11, 10, 22, 16, 14, 9]):
        ws.column_dimensions[col].width = w
    for i in range(n_days):
        ws.column_dimensions[get_column_letter(FIX+1+i)].width = 6
    for j in range(len(ATT_HDRS)):
        ws.column_dimensions[get_column_letter(base_att+1+j)].width = 8
    for j in range(len(SAL)):
        ws.column_dimensions[get_column_letter(base_sal+1+j)].width = 11
    ws.freeze_panes = 'H6'

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f'HCP_Salary_Register_{month_name}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
