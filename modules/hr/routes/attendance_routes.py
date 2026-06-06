"""
attendance_routes.py â€” Attendance API + Dashboard

API:
  POST /api/receive_logs          â† PHP push_to_live() yahan data bhejta hai

Routes:
  GET  /hr/attendance/            â† Dashboard
  GET  /hr/attendance/logs        â† Raw punch logs table
  GET  /hr/attendance/report      â† Monthly attendance report
"""
import json
from decimal import Decimal
from datetime import datetime, date, timedelta
from flask import Blueprint, request, jsonify, render_template, abort, flash, redirect, url_for
from flask_login import login_required, current_user
from models import db, Employee
from models.attendance import RawPunchLog, Attendance
from core.permissions import get_sub_perm

attendance_bp = Blueprint('attendance', __name__)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Helper: gate decorator-style for sub-perm + admin bypass
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _require_sub_perm(module, key, redirect_to='attendance.attendance_dashboard'):
    """
    Returns True if access allowed. If denied, calls flash() and returns False.
    Caller should: `if not _require_sub_perm(...): return redirect(...)`
    Admin role bypasses all checks.
    """
    if not current_user.is_authenticated:
        return False
    if current_user.role == 'admin':
        return True
    return bool(get_sub_perm(module, key))

# â”€â”€ Auth key â€” PHP push_to_live() mein jo Authorization header hai â”€â”€
PUSH_API_KEY = "HCP_PUSH_2024"

# â”€â”€ Shift timing â”€â”€
SHIFT_START = (9, 0)    # 9:00 AM
SHIFT_END   = (18, 0)   # 6:00 PM
HALF_DAY_HOURS = 4.0    # 4 ghante se kam = Half Day


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HELPER: Device API response fields parse karo
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _get_field(entry, *keys):
    """
    Multiple possible key names try karo (case-insensitive).
    Device APIs har vendor ka alag JSON shape use karta hai â€”
    isiliye CamelCase, snake_case, lowercase sab try hoti hain.
    """
    if not isinstance(entry, dict):
        return None
    # First: exact match (fast path)
    for k in keys:
        v = entry.get(k)
        if v is not None and str(v).strip() != '':
            return str(v).strip()
    # Fallback: case-insensitive match
    lower_map = {str(k).lower(): k for k in entry.keys()}
    for k in keys:
        actual = lower_map.get(str(k).lower())
        if actual is not None:
            v = entry.get(actual)
            if v is not None and str(v).strip() != '':
                return str(v).strip()
    return None


def _parse_datetime(val):
    """
    Bahut saare formats try karo. Device APIs alag-alag format mein
    dates bhejti hain â€” Z-suffix UTC, milliseconds, ISO with offset, etc.
    """
    if not val:
        return None
    s = str(val).strip()
    # Strip 'Z' (UTC marker) and milliseconds for parsing
    if s.endswith('Z'):
        s = s[:-1]
    if '.' in s and 'T' in s:
        # e.g. 2026-04-30T08:55:00.123 â†’ strip ms
        try:
            dot = s.index('.')
            tplus = s.find('+', dot)
            tminus = s.find('-', dot)
            cut = min([x for x in [tplus, tminus, len(s)] if x > 0])
            s = s[:dot] + s[cut:] if cut < len(s) else s[:dot]
        except Exception:
            pass

    formats = (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M:%S%z',
        '%d/%m/%Y %H:%M:%S',
        '%m/%d/%Y %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d-%m-%Y %H:%M:%S',
        '%d-%m-%Y %H:%M',
        '%Y/%m/%d %H:%M:%S',
        '%d.%m.%Y %H:%M:%S',
    )
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue

    # Last try: epoch millis or seconds
    try:
        n = float(s)
        if n > 1e12:   # milliseconds
            return datetime.fromtimestamp(n / 1000.0)
        if n > 1e9:    # seconds
            return datetime.fromtimestamp(n)
    except Exception:
        pass
    return None


def _get_punch_direction(entry):
    """IN ya OUT detect karo device log se."""
    val = _get_field(entry, 'Direction', 'PunchType', 'punch_type',
                     'Type', 'type', 'punch_direction')
    if val:
        v = val.upper()
        if v in ('IN', '0', 'CHECKIN', 'CHECK IN', 'ENTRY', 'E'):
            return 'IN'
        if v in ('OUT', '1', 'CHECKOUT', 'CHECK OUT', 'EXIT', 'X'):
            return 'OUT'
    return 'IN'   # default IN


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# CORE: attendance table update karo (first IN + last OUT)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _update_attendance(employee_code, log_date_only):
    """
    raw_punch_logs se us employee ka us din ka
    PEHLA punch = punch_in
    AAKHRI punch = punch_out
    calculate karke attendance table mein save karo.
    """
    # Us din ke saare punches sorted by time
    punches = RawPunchLog.query.filter(
        RawPunchLog.employee_code == employee_code,
        db.func.date(RawPunchLog.log_date) == log_date_only
    ).order_by(RawPunchLog.log_date.asc()).all()

    if not punches:
        return

    first_punch = punches[0]
    last_punch  = punches[-1]

    punch_in    = first_punch.log_date
    in_device   = first_punch.serial_number

    # punch_out sirf tab jab ek se zyada punch ho
    punch_out   = last_punch.log_date  if len(punches) > 1 else None
    out_device  = last_punch.serial_number if len(punches) > 1 else None

    # Total hours calculate karo
    total_hours = None
    if punch_in and punch_out and punch_out > punch_in:
        diff_minutes = (punch_out - punch_in).total_seconds() / 3600
        total_hours  = round(diff_minutes, 2)

    # Status determine karo
    if punch_out is None:
        status = 'MIS-PUNCH'   # Sirf ek hi punch hai â€” in ya out pata nahi
    elif total_hours is not None and total_hours < HALF_DAY_HOURS:
        status = 'Half Day'
    else:
        status = 'Present'

    # Attendance record upsert karo
    att = Attendance.query.filter_by(
        employee_code=employee_code,
        attendance_date=log_date_only
    ).first()

    # Employee link â€” try employee_id first, then employee_code
    from models import Employee as _Emp
    emp_obj = _Emp.query.filter_by(employee_id=employee_code).first()
    if not emp_obj:
        emp_obj = _Emp.query.filter_by(employee_code=employee_code).first()

    if not att:
        att = Attendance(
            employee_code=employee_code,
            attendance_date=log_date_only
        )
        db.session.add(att)

    att.punch_in    = punch_in
    att.punch_out   = punch_out
    att.in_device   = in_device
    att.out_device  = out_device
    att.total_hours = total_hours
    att.status      = status
    att.updated_at  = datetime.now()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API: POST /api/receive_logs
# PHP push_to_live() yahan call karta hai
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/api/receive_logs', methods=['POST'])
def receive_logs():
    import sys

    # â”€â”€ DEBUG: log incoming request basics â”€â”€
    sys.stderr.write(f"\n========== RECEIVE_LOGS DEBUG ==========\n")
    sys.stderr.write(f"  From IP      : {request.remote_addr}\n")
    sys.stderr.write(f"  Auth header  : '{request.headers.get('Authorization', '<MISSING>')}'\n")
    sys.stderr.write(f"  Content-Type : '{request.headers.get('Content-Type', '<MISSING>')}'\n")
    sys.stderr.write(f"  Body bytes   : {len(request.get_data())}\n")
    sys.stderr.flush()

    # â”€â”€ Auth check â”€â”€
    auth = request.headers.get('Authorization', '')
    if auth != PUSH_API_KEY:
        sys.stderr.write(f"  âŒ Auth FAILED â€” expected '{PUSH_API_KEY}'\n")
        sys.stderr.flush()
        return jsonify({'error': 'Unauthorized'}), 401
    sys.stderr.write(f"  âœ… Auth OK\n")
    sys.stderr.flush()

    # â”€â”€ JSON parse â”€â”€
    try:
        data = request.get_json(force=True)
    except Exception as ex:
        sys.stderr.write(f"  âŒ JSON parse failed: {ex}\n")
        sys.stderr.flush()
        return jsonify({'error': 'Invalid JSON'}), 400

    if not data:
        return jsonify({'error': 'Empty payload'}), 400

    # â”€â”€ List normalize karo â”€â”€
    if isinstance(data, dict):
        logs_list = (data.get('data') or data.get('logs') or
                     data.get('DeviceLogs') or data.get('Records') or
                     data.get('Result')   or data.get('result')  or
                     [data])
    elif isinstance(data, list):
        logs_list = data
    else:
        return jsonify({'error': 'Unexpected format'}), 400

    # â”€â”€ DEBUG: dump first 2 entries' structure â”€â”€
    sys.stderr.write(f"  payload type : {type(data).__name__}\n")
    sys.stderr.write(f"  logs_list len: {len(logs_list)}\n")
    for i, entry in enumerate(logs_list[:2]):
        sys.stderr.write(f"  entry[{i}] type: {type(entry).__name__}\n")
        if isinstance(entry, dict):
            sys.stderr.write(f"  entry[{i}] keys: {list(entry.keys())}\n")
            sys.stderr.write(f"  entry[{i}] data: {entry}\n")
    sys.stderr.flush()

    inserted       = 0
    skipped        = 0
    skip_reasons   = {'no_emp': 0, 'no_dt': 0, 'duplicate': 0, 'other': 0}
    errors         = []
    to_update      = set()
    sample_skipped = None    # pehla skipped entry yaad rakhne ke liye

    # â”€â”€ Extended key list â€” alag alag biometric brands ke saath compatible â”€â”€
    # eSSL, ZKTeco, Realtime, Matrix, Anviz, etc.
    EMP_CODE_KEYS = (
        'EmployeeCode', 'employee_code', 'Employee_Code',
        'CardNo', 'card_no', 'CardNumber',
        'UserID', 'UserId', 'userid', 'user_id', 'UserCode',
        'EmpCode', 'emp_code', 'EmpId', 'emp_id', 'EmployeeID',
        'PIN', 'pin', 'Pin', 'BadgeNumber', 'badge_number',
        'StaffCode', 'staff_code', 'StaffID',
    )
    LOG_DT_KEYS = (
        'LogTime', 'log_date', 'log_time', 'LogDate',
        'PunchTime', 'punch_time', 'punchTime',
        'DateTime', 'datetime', 'DateAndTime',
        'Time', 'time', 'Timestamp', 'timestamp',
        'AttendanceTime', 'attendance_time',
        'CheckTime', 'check_time', 'EventTime',
    )

    for entry in logs_list:
        try:
            if not isinstance(entry, dict):
                skipped += 1
                skip_reasons['other'] += 1
                if sample_skipped is None: sample_skipped = repr(entry)[:200]
                continue

            emp_code = _get_field(entry, *EMP_CODE_KEYS)
            log_dt   = _parse_datetime(_get_field(entry, *LOG_DT_KEYS))

            if not emp_code:
                skipped += 1
                skip_reasons['no_emp'] += 1
                if sample_skipped is None: sample_skipped = f"no emp_code â†’ keys={list(entry.keys())}"
                continue
            if not log_dt:
                skipped += 1
                skip_reasons['no_dt'] += 1
                if sample_skipped is None:
                    raw_dt = _get_field(entry, *LOG_DT_KEYS)
                    sample_skipped = f"no log_dt (raw={raw_dt!r}) â†’ keys={list(entry.keys())}"
                continue

            # Duplicate check
            exists = RawPunchLog.query.filter_by(
                employee_code=emp_code,
                log_date=log_dt
            ).first()
            if exists:
                skipped += 1
                skip_reasons['duplicate'] += 1
                continue

            punch = RawPunchLog(
                employee_code     = emp_code,
                log_date          = log_dt,
                serial_number     = _get_field(entry, 'SerialNumber', 'serial_number',
                                               'DeviceId', 'device_id', 'MachineNo',
                                               'TerminalId', 'terminal_id'),
                punch_direction   = _get_punch_direction(entry),
                temperature       = _get_field(entry, 'Temperature', 'temperature') or 0.00,
                temperature_state = _get_field(entry, 'TemperatureState', 'temperature_state'),
                synced_at         = datetime.now(),
            )
            db.session.add(punch)
            to_update.add((emp_code, log_dt.date()))
            inserted += 1

        except Exception as e:
            errors.append(str(e))
            skip_reasons['other'] += 1
            continue

    # â”€â”€ DEBUG: summary print â”€â”€
    sys.stderr.write(f"  â”€ SUMMARY â”€\n")
    sys.stderr.write(f"  inserted     : {inserted}\n")
    sys.stderr.write(f"  skipped      : {skipped} â†’ {skip_reasons}\n")
    if sample_skipped:
        sys.stderr.write(f"  first skip   : {sample_skipped}\n")
    if errors:
        sys.stderr.write(f"  errors       : {errors[:3]}\n")
    sys.stderr.flush()

    # â”€â”€ Commit raw punches â”€â”€
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        sys.stderr.write(f"  âŒ DB commit failed: {e}\n")
        sys.stderr.flush()
        return jsonify({'error': f'DB error: {str(e)}'}), 500

    # â”€â”€ attendance table update â€” SKIPPED in receive_logs â”€â”€
    # 7000+ entries pe loop chalane se gunicorn worker timeout ho jata hai.
    # Solution: HR admin /attendance_log/sync screen se "Run Summary Sync"
    # press kare â€” woh batch processing optimal hai aur progress dikhta hai.
    #
    # Agar real-time chahiye toh background queue (Celery/RQ) chahiye â€”
    # for now ye trade-off acceptable hai kyunki PHP cron 1 din mein 1-2
    # baar chalti hai aur sync 30 second mein ho jaata hai manually.
    sys.stderr.write(f"  â„¹ï¸  Skipping inline _update_attendance() to avoid "
                     f"timeout â€” use /attendance_log/sync to recompute.\n")
    sys.stderr.flush()

    return jsonify({
        'status':       'ok',
        'inserted':     inserted,
        'skipped':      skipped,
        'skip_reasons': skip_reasons,
        'total':        len(logs_list),
        'errors':       errors[:5],
        'sample_skip':  sample_skipped,
        'note':         'Raw punches saved. Run /attendance_log/sync to update attendance table.',
    }), 200


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DASHBOARD: GET /hr/attendance/
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/')
@attendance_bp.route('/hr/attendance')
@login_required
def attendance_dashboard():
    from models.hr_rules import HRLeaveApplication

    # ── Selected date (?date=YYYY-MM-DD, default = today) ──
    sel_date_str = request.args.get('date', date.today().isoformat())
    try:
        sel_date = datetime.strptime(sel_date_str, '%Y-%m-%d').date()
    except ValueError:
        sel_date = date.today()

    # ── Active employees ──
    employees = Employee.query.filter(
        Employee.status == 'active',
        db.or_(Employee.is_deleted.is_(False), Employee.is_deleted.is_(None)),
    ).all()
    total_employees = len(employees)

    # ── Selected day ka attendance + approved leaves ──
    att_map = {r.employee_code: r for r in Attendance.query.filter_by(
        attendance_date=sel_date).all()}

    leaves = HRLeaveApplication.query.filter(
        HRLeaveApplication.status == 'approved',
        HRLeaveApplication.from_date <= sel_date,
        HRLeaveApplication.to_date   >= sel_date,
    ).all()
    leave_by_emp = {l.employee_id: l for l in leaves}

    # ── Per-employee status resolve ──
    # Attendance row > approved leave > week off > Absent
    def _resolve(emp):
        att = att_map.get(emp.employee_code) or att_map.get(emp.employee_id)
        if att:
            return att.status, att
        if emp.id in leave_by_emp:
            return 'On Leave', None
        if _is_week_off(emp.employee_type or '', sel_date):
            return 'Week Off', None
        return 'Absent', None

    STATUS_COLORS = {
        'Present':   '#10b981', 'Half Day': '#8b5cf6', 'MIS-PUNCH': '#f59e0b',
        'Absent':    '#ef4444', 'On Leave': '#3b82f6', 'Week Off':  '#94a3b8',
        'WOP':       '#06b6d4', 'Holiday':  '#64748b',
    }

    overview_counts = {}
    cat_grp, dept_grp, shift_grp = {}, {}, {}
    gender_present = {}
    present_n = absent_n = mispunch_n = on_leave_n = 0

    def _bump(grp, key, field):
        g = grp.setdefault(key or '—', dict(total=0, present=0, absent=0,
                                            mispunch=0, on_leave=0))
        g['total'] += 1
        if field:
            g[field] += 1

    for emp in employees:
        st, att = _resolve(emp)
        overview_counts[st] = overview_counts.get(st, 0) + 1

        if st in ('Present', 'WOP'):
            present_n += 1;  fld = 'present'
            gkey = (emp.gender or 'Other').title()
            gender_present[gkey] = gender_present.get(gkey, 0) + 1
        elif st == 'Half Day':
            present_n += 1;  fld = 'present'
            gkey = (emp.gender or 'Other').title()
            gender_present[gkey] = gender_present.get(gkey, 0) + 1
        elif st == 'MIS-PUNCH':
            mispunch_n += 1; fld = 'mispunch'
        elif st == 'On Leave':
            on_leave_n += 1; fld = 'on_leave'
        elif st == 'Absent':
            absent_n += 1;   fld = 'absent'
        else:                       # Week Off / Holiday — kisi KPI bucket mein nahi
            fld = None

        _bump(cat_grp,   emp.employee_type, fld)
        _bump(dept_grp,  emp.department,    fld)
        _bump(shift_grp, emp.shift,         fld)

    def _pct(n):
        return round(n * 100.0 / total_employees, 1) if total_employees else 0

    # ── Working days (selected month — Sundays exclude) ──
    month_start = sel_date.replace(day=1)
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    days_in_month = (next_month - month_start).days
    working_days = sum(1 for i in range(days_in_month)
                       if (month_start + timedelta(days=i)).weekday() != 6)

    # ── 7-day trend (sel_date tak) ──
    trend_start = sel_date - timedelta(days=6)
    trend_rows = Attendance.query.filter(
        Attendance.attendance_date >= trend_start,
        Attendance.attendance_date <= sel_date,
    ).all()
    by_day = {}
    for r in trend_rows:
        d = by_day.setdefault(r.attendance_date, dict(present=0, absent=0, mispunch=0))
        if r.status in ('Present', 'Half Day', 'WOP'):
            d['present'] += 1
        elif r.status == 'Absent':
            d['absent'] += 1
        elif r.status == 'MIS-PUNCH':
            d['mispunch'] += 1

    trend_leaves = HRLeaveApplication.query.filter(
        HRLeaveApplication.status == 'approved',
        HRLeaveApplication.from_date <= sel_date,
        HRLeaveApplication.to_date   >= trend_start,
    ).all()

    trend = []
    for i in range(6, -1, -1):
        d = sel_date - timedelta(days=i)
        day = by_day.get(d, {})
        lv = sum(1 for l in trend_leaves if l.from_date <= d <= l.to_date)
        trend.append({
            'label':    d.strftime('%d %b'),
            'present':  day.get('present', 0),
            'absent':   day.get('absent', 0),
            'mispunch': day.get('mispunch', 0),
            'on_leave': lv,
        })

    # ── Top mis-punch list (selected day) ──
    top_mispunch = []
    for code, att in att_map.items():
        if att.status != 'MIS-PUNCH':
            continue
        emp = att.employee
        top_mispunch.append({
            'code':  code,
            'name':  (emp.full_name if emp else code),
            'in':    att.punch_in.strftime('%I:%M %p') if att.punch_in else '—',
            'issue': 'Punch Out Missing' if att.punch_in else 'Punch In Missing',
        })
        if len(top_mispunch) >= 8:
            break

    # ── On-leave list (selected day) ──
    leave_rows = [{
        'code': (l.employee.employee_code if l.employee else l.employee_id),
        'name': (l.employee.full_name if l.employee else str(l.employee_id)),
        'type': l.type_label,
        'days': float(l.days or 0),
    } for l in leaves[:8]]

    # ── Group dicts → sorted lists ──
    def _to_rows(grp):
        return [dict(label=k, **v) for k, v in
                sorted(grp.items(), key=lambda kv: -kv[1]['total'])]

    category_summary = _to_rows(cat_grp)
    dept_summary     = _to_rows(dept_grp)
    shift_summary    = _to_rows(shift_grp)
    category_bar     = category_summary[:8]

    overview = [{'label': k, 'value': v, 'color': STATUS_COLORS.get(k, '#94a3b8')}
                for k, v in sorted(overview_counts.items(), key=lambda kv: -kv[1])
                if v > 0]
    GENDER_COLORS = {'Male': '#6366f1', 'Female': '#ec4899'}
    gender = [{'label': k, 'value': v, 'color': GENDER_COLORS.get(k, '#94a3b8')}
              for k, v in sorted(gender_present.items(), key=lambda kv: -kv[1])]

    # ── Last sync ──
    last_raw  = RawPunchLog.query.order_by(RawPunchLog.synced_at.desc()).first()
    last_sync = last_raw.synced_at.strftime('%d %b %Y, %I:%M %p') if last_raw else 'No data yet'

    a = {
        'total_employees': total_employees,
        'present':         present_n,   'present_pct':  _pct(present_n),
        'absent':          absent_n,    'absent_pct':   _pct(absent_n),
        'mispunch':        mispunch_n,  'mispunch_pct': _pct(mispunch_n),
        'on_leave':        on_leave_n,  'on_leave_pct': _pct(on_leave_n),
        'working_days':    working_days,
        'overview':         overview,
        'gender':           gender,
        'trend':            trend,
        'category_bar':     category_bar,
        'category_summary': category_summary,
        'dept_summary':     dept_summary,
        'shift_summary':    shift_summary,
        'top_mispunch':     top_mispunch,
        'leave_rows':       leave_rows,
        'last_sync':        last_sync,
    }

    return render_template(
        'hr/attendance/dashboard.html',
        sel_date    = sel_date,
        analytics   = a,          # template line 71: {% set a = analytics ... %}
        active_page = 'hr_attendance',
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# RAW LOGS PAGE: GET /hr/attendance/logs
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/logs')
@login_required
def attendance_logs():
    page       = request.args.get('page', 1, type=int)
    emp_search = request.args.get('emp', '').strip()
    date_from  = request.args.get('from', '')
    date_to    = request.args.get('to', '')

    q = RawPunchLog.query

    if emp_search:
        q = q.filter(RawPunchLog.employee_code.ilike(f'%{emp_search}%'))
    if date_from:
        try:
            q = q.filter(db.func.date(RawPunchLog.log_date) >=
                         datetime.strptime(date_from, '%Y-%m-%d').date())
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(db.func.date(RawPunchLog.log_date) <=
                         datetime.strptime(date_to, '%Y-%m-%d').date())
        except Exception:
            pass

    pagination = q.order_by(RawPunchLog.log_date.desc()).paginate(
        page=page, per_page=50, error_out=False
    )

    return render_template(
        'hr/attendance/logs.html',
        logs        = pagination.items,
        pagination  = pagination,
        emp_search  = emp_search,
        date_from   = date_from,
        date_to     = date_to,
        active_page = 'hr_attendance',
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# REPORT PAGE: GET /hr/attendance/report
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/report')
@login_required
def attendance_report():
    month_str  = request.args.get('month', date.today().strftime('%Y-%m'))
    emp_search = request.args.get('emp', '').strip()
    page       = request.args.get('page', 1, type=int)

    try:
        month_start = datetime.strptime(month_str + '-01', '%Y-%m-%d').date()
    except Exception:
        month_start = date.today().replace(day=1)

    # Month end
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(days=1)

    status_filter = request.args.get('status', '').strip()

    q = Attendance.query.filter(
        Attendance.attendance_date >= month_start,
        Attendance.attendance_date <= month_end
    )
    if emp_search:
        q = q.filter(Attendance.employee_code.ilike(f'%{emp_search}%'))
    if status_filter:
        q = q.filter(Attendance.status == status_filter)

    pagination = q.order_by(
        Attendance.attendance_date.desc(),
        Attendance.employee_code
    ).paginate(page=page, per_page=50, error_out=False)

    # Month summary stats
    month_stats = dict(
        db.session.query(Attendance.status, db.func.count(Attendance.id))
        .filter(
            Attendance.attendance_date >= month_start,
            Attendance.attendance_date <= month_end
        ).group_by(Attendance.status).all()
    )

    return render_template(
        'hr/attendance/report.html',
        records       = pagination.items,
        pagination    = pagination,
        month_str     = month_str,
        month_start   = month_start,
        month_end     = month_end,
        emp_search    = emp_search,
        status_filter = status_filter,
        month_stats   = month_stats,
        active_page   = 'hr_attendance',
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# AJAX: GET /hr/attendance/api/daily-summary
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/api/daily-summary')
@login_required
def api_daily_summary():
    month_str = request.args.get('month', date.today().strftime('%Y-%m'))
    try:
        month_start = datetime.strptime(month_str + '-01', '%Y-%m-%d').date()
    except Exception:
        month_start = date.today().replace(day=1)

    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(days=1)

    rows = db.session.query(
        Attendance.attendance_date,
        db.func.count(db.case(
            (Attendance.status == 'Present', 1))).label('present'),
        db.func.count(db.case(
            (Attendance.status == 'Absent', 1))).label('absent'),
        db.func.count(db.case(
            (Attendance.status == 'Half Day', 1))).label('half_day'),
        db.func.count(db.case(
            (Attendance.status == 'MIS-PUNCH', 1))).label('mis_punch'),
    ).filter(
        Attendance.attendance_date >= month_start,
        Attendance.attendance_date <= month_end,
    ).group_by(Attendance.attendance_date).order_by(Attendance.attendance_date).all()

    return jsonify({'data': [
        {
            'date':      r.attendance_date.strftime('%d %b'),
            'present':   r.present,
            'absent':    r.absent,
            'half_day':  r.half_day,
            'mis_punch': r.mis_punch,
        } for r in rows
    ]})


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MANUAL ATTENDANCE ENTRY & EDIT
# GET/POST /hr/attendance/manual
# GET/POST /hr/attendance/<int:id>/edit
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/manual', methods=['GET', 'POST'])
@login_required
def attendance_manual():
    from flask_login import current_user
    if current_user.role not in ('admin', 'manager', 'hr'):
        from flask import abort; abort(403)

    employees = Employee.query.filter_by(status='active').order_by(Employee.first_name).all()
    msg = None

    if request.method == 'POST':
        emp_code    = request.form.get('employee_code', '').strip()
        att_date    = request.form.get('attendance_date', '').strip()
        punch_in_s  = request.form.get('punch_in', '').strip()
        punch_out_s = request.form.get('punch_out', '').strip()
        status      = request.form.get('status', 'Present')
        in_device   = request.form.get('in_device', 'MANUAL').strip()

        try:
            att_date_obj = datetime.strptime(att_date, '%Y-%m-%d').date()
            pin  = datetime.strptime(f"{att_date} {punch_in_s}",  '%Y-%m-%d %H:%M') if punch_in_s  else None
            pout = datetime.strptime(f"{att_date} {punch_out_s}", '%Y-%m-%d %H:%M') if punch_out_s else None

            total_hours = None
            if pin and pout and pout > pin:
                total_hours = round((pout - pin).total_seconds() / 3600, 2)

            # Auto status
            if status == 'auto':
                if not pin:                 status = 'Absent'
                elif not pout:              status = 'MIS-PUNCH'
                elif total_hours < 4:       status = 'Half Day'
                else:                       status = 'Present'

            existing = Attendance.query.filter_by(
                employee_code=emp_code, attendance_date=att_date_obj
            ).first()

            if existing:
                existing.punch_in    = pin
                existing.punch_out   = pout
                existing.in_device   = in_device
                existing.out_device  = in_device if pout else None
                existing.total_hours = total_hours
                existing.status      = status
                existing.updated_at  = datetime.now()
                msg = ('success', f'Attendance updated for {emp_code} on {att_date}')
            else:
                att = Attendance(
                    employee_code=emp_code, attendance_date=att_date_obj,
                    punch_in=pin, punch_out=pout,
                    in_device=in_device, out_device=in_device if pout else None,
                    total_hours=total_hours, status=status,
                )
                db.session.add(att)
                msg = ('success', f'Attendance added for {emp_code} on {att_date}')

            db.session.commit()

            # Raw punch log bhi add karo
            if pin:
                if not RawPunchLog.query.filter_by(employee_code=emp_code, log_date=pin).first():
                    db.session.add(RawPunchLog(
                        employee_code=emp_code, log_date=pin,
                        serial_number='MANUAL', punch_direction='IN', synced_at=datetime.now()
                    ))
            if pout:
                if not RawPunchLog.query.filter_by(employee_code=emp_code, log_date=pout).first():
                    db.session.add(RawPunchLog(
                        employee_code=emp_code, log_date=pout,
                        serial_number='MANUAL', punch_direction='OUT', synced_at=datetime.now()
                    ))
            db.session.commit()

        except Exception as e:
            db.session.rollback()
            msg = ('error', f'Error: {str(e)}')

    # Recent entries
    recent = Attendance.query.order_by(
        Attendance.updated_at.desc()
    ).limit(20).all()

    return render_template('hr/attendance/manual.html',
        employees=employees, msg=msg, recent=recent,
        today=date.today().strftime('%Y-%m-%d'),
        active_page='att_manual'
    )


@attendance_bp.route('/hr/attendance/<int:att_id>/edit', methods=['GET', 'POST'])
@login_required
def attendance_edit(att_id):
    from flask_login import current_user
    if current_user.role not in ('admin', 'manager', 'hr'):
        from flask import abort; abort(403)

    att       = Attendance.query.get_or_404(att_id)
    employees = Employee.query.filter_by(status='active').order_by(Employee.first_name).all()
    msg       = None

    if request.method == 'POST':
        try:
            att_date    = request.form.get('attendance_date', '')
            punch_in_s  = request.form.get('punch_in', '').strip()
            punch_out_s = request.form.get('punch_out', '').strip()
            status      = request.form.get('status', att.status)

            att.attendance_date = datetime.strptime(att_date, '%Y-%m-%d').date()
            att.punch_in   = datetime.strptime(f"{att_date} {punch_in_s}",  '%Y-%m-%d %H:%M') if punch_in_s  else None
            att.punch_out  = datetime.strptime(f"{att_date} {punch_out_s}", '%Y-%m-%d %H:%M') if punch_out_s else None
            att.in_device  = request.form.get('in_device', att.in_device or 'MANUAL')
            att.status     = status
            att.updated_at = datetime.now()

            if att.punch_in and att.punch_out and att.punch_out > att.punch_in:
                att.total_hours = round((att.punch_out - att.punch_in).total_seconds() / 3600, 2)
            else:
                att.total_hours = None

            db.session.commit()
            msg = ('success', 'Attendance updated successfully!')
        except Exception as e:
            db.session.rollback()
            msg = ('error', str(e))

    return render_template('hr/attendance/edit.html',
        att=att, employees=employees, msg=msg,
        active_page='my_attendance'
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# LATE COMERS & ABSENT REPORT
# GET /hr/attendance/late-absent
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/late-absent')
@login_required
def attendance_late_absent():
    report_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    report_type = request.args.get('type', 'late')  # late / absent / mispunch
    dept_filter = request.args.get('dept', '')

    try:
        filter_date = datetime.strptime(report_date, '%Y-%m-%d').date()
    except Exception:
        filter_date = date.today()

    SHIFT_START_H, SHIFT_START_M = 9, 0
    GRACE_MINUTES = 15

    # Late comers â€” punch_in after 9:15
    if report_type == 'late':
        shift_start = datetime(filter_date.year, filter_date.month, filter_date.day,
                               SHIFT_START_H, SHIFT_START_M + GRACE_MINUTES)
        records = Attendance.query.filter(
            Attendance.attendance_date == filter_date,
            Attendance.punch_in != None,
            Attendance.punch_in > shift_start,
            Attendance.status.in_(['Present', 'Half Day'])
        ).order_by(Attendance.punch_in.asc()).all()

        # Late minutes calculate
        for rec in records:
            if rec.punch_in:
                base = datetime(filter_date.year, filter_date.month, filter_date.day,
                                SHIFT_START_H, SHIFT_START_M)
                rec._late_min = max(0, int((rec.punch_in - base).total_seconds() / 60))
            else:
                rec._late_min = 0

    elif report_type == 'absent':
        records = Attendance.query.filter(
            Attendance.attendance_date == filter_date,
            Attendance.status == 'Absent'
        ).order_by(Attendance.employee_code).all()
        for rec in records:
            rec._late_min = 0

    else:  # mispunch
        records = Attendance.query.filter(
            Attendance.attendance_date == filter_date,
            Attendance.status == 'MIS-PUNCH'
        ).order_by(Attendance.punch_in.asc()).all()
        for rec in records:
            rec._late_min = 0

    # Dept filter
    if dept_filter:
        records = [r for r in records if r.employee and r.employee.department == dept_filter]

    # All departments for filter dropdown
    departments = db.session.query(Employee.department).filter(
        Employee.department != None, Employee.status == 'active'
    ).distinct().order_by(Employee.department).all()
    departments = [d[0] for d in departments]

    # Summary counts for the date
    summary = dict(
        db.session.query(Attendance.status, db.func.count(Attendance.id))
        .filter(Attendance.attendance_date == filter_date)
        .group_by(Attendance.status).all()
    )
    # Late count
    shift_start_dt = datetime(filter_date.year, filter_date.month, filter_date.day,
                               SHIFT_START_H, SHIFT_START_M + GRACE_MINUTES)
    late_count = Attendance.query.filter(
        Attendance.attendance_date == filter_date,
        Attendance.punch_in > shift_start_dt,
        Attendance.status.in_(['Present', 'Half Day'])
    ).count()

    return render_template('hr/attendance/late_absent.html',
        records=records, report_date=report_date, report_type=report_type,
        filter_date=filter_date, dept_filter=dept_filter,
        departments=departments, summary=summary, late_count=late_count,
        active_page='att_late'
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOLIDAY MASTER
# GET/POST /hr/attendance/holidays
# POST /hr/attendance/holidays/<int:id>/delete
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/holidays', methods=['GET', 'POST'])
@login_required
def holiday_master():
    from flask_login import current_user
    from models.attendance import HolidayMaster
    if current_user.role not in ('admin', 'manager', 'hr'):
        from flask import abort; abort(403)

    msg = None
    if request.method == 'POST':
        title        = request.form.get('title', '').strip()
        holiday_date = request.form.get('holiday_date', '').strip()
        holiday_type = request.form.get('holiday_type', 'National')
        location     = request.form.get('location', 'All').strip() or 'All'
        if location not in ('All', 'HO', 'Plant'):
            location = 'All'
        description  = request.form.get('description', '').strip()

        if not title or not holiday_date:
            msg = ('error', 'Title aur Date required hai.')
        else:
            try:
                hdate = datetime.strptime(holiday_date, '%Y-%m-%d').date()
                # Conflict check: same date par —
                #   same location duplicate, ya 'All' kisi ke saath overlap
                same_day = HolidayMaster.query.filter_by(holiday_date=hdate).all()
                conflict = None
                for ex in same_day:
                    if ex.location == location or ex.location == 'All' or location == 'All':
                        conflict = ex
                        break
                if conflict:
                    msg = ('error', f'{holiday_date} pe pehle se holiday hai '
                                    f'({conflict.location}): {conflict.title}')
                else:
                    h = HolidayMaster(
                        title=title, holiday_date=hdate,
                        holiday_type=holiday_type, location=location,
                        description=description,
                        created_by=current_user.id
                    )
                    db.session.add(h)
                    db.session.commit()
                    msg = ('success', f'Holiday "{title}" added!')
            except Exception as e:
                db.session.rollback()
                msg = ('error', str(e))

    year  = request.args.get('year', date.today().year, type=int)
    holidays = HolidayMaster.query.filter(
        db.extract('year', HolidayMaster.holiday_date) == year
    ).order_by(HolidayMaster.holiday_date).all()

    return render_template('hr/attendance/holidays.html',
        holidays=holidays, msg=msg, year=year,
        today=date.today().strftime('%Y-%m-%d'),
        active_page='att_holidays'
    )


@attendance_bp.route('/hr/attendance/holidays/<int:hid>/delete', methods=['POST'])
@login_required
def holiday_delete(hid):
    from flask_login import current_user
    from models.attendance import HolidayMaster
    if current_user.role not in ('admin',):
        from flask import abort; abort(403)
    h = HolidayMaster.query.get_or_404(hid)
    db.session.delete(h)
    db.session.commit()
    from flask import redirect, url_for
    return redirect(url_for('attendance.holiday_master'))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# EMPLOYEE â€” MY ATTENDANCE
# GET /hr/attendance/my
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/my')
@login_required
def my_attendance():
    from flask_login import current_user
    from flask import abort

    # Employee dhundo current user ke liye
    emp = Employee.query.filter_by(user_id=current_user.id).first()
    if not emp:
        # role admin/hr ho toh redirect to dashboard
        if current_user.role in ('admin', 'manager', 'hr'):
            from flask import redirect, url_for
            return redirect(url_for('attendance.attendance_dashboard'))
        abort(404)

    month_str = request.args.get('month', date.today().strftime('%Y-%m'))
    try:
        month_start = datetime.strptime(month_str + '-01', '%Y-%m-%d').date()
    except Exception:
        month_start = date.today().replace(day=1)

    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year+1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month+1, day=1) - timedelta(days=1)

    # Us mahine ki saari attendance
    records = Attendance.query.filter(
        Attendance.employee_code == emp.employee_code,
        Attendance.attendance_date >= month_start,
        Attendance.attendance_date <= month_end
    ).order_by(Attendance.attendance_date.asc()).all()

    # Month summary
    att_map = {r.attendance_date: r for r in records}
    summary = {'Present':0,'Absent':0,'Half Day':0,'MIS-PUNCH':0,'Holiday':0}
    for r in records:
        summary[r.status] = summary.get(r.status, 0) + 1

    # Today's punches
    today_punches = RawPunchLog.query.filter(
        RawPunchLog.employee_code == (emp.employee_id or emp.employee_code),
        db.func.date(RawPunchLog.log_date) == date.today()
    ).order_by(RawPunchLog.log_date.asc()).all()

    # Calendar data â€” har din ka status
    calendar_days = []
    current_day = month_start
    while current_day <= month_end:
        rec = att_map.get(current_day)
        calendar_days.append({
            'date':    current_day,
            'weekday': current_day.weekday(),
            'status':  rec.status if rec else ('Future' if current_day > date.today() else 'No Data'),
            'punch_in':  rec.punch_in.strftime('%I:%M %p')  if rec and rec.punch_in  else None,
            'punch_out': rec.punch_out.strftime('%I:%M %p') if rec and rec.punch_out else None,
            'hours':     rec.working_hours_display          if rec else None,
        })
        current_day += timedelta(days=1)

    prev_month = (month_start - timedelta(days=1)).strftime('%Y-%m')
    next_month = (month_end + timedelta(days=1)).strftime('%Y-%m')

    return render_template('hr/attendance/my_attendance.html',
        emp=emp, month_str=month_str,
        month_start=month_start, month_end=month_end,
        prev_month=prev_month, next_month=next_month,
        records=records, calendar_days=calendar_days,
        summary=summary, today_punches=today_punches,
        today=date.today(),
        active_page='hr_attendance'
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ATTENDANCE LOG SYNC SCREEN â€” /attendance_log/sync
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# PHP CodeIgniter ke `sync_all_summary_to_attendance()` aur
# `push_to_python()` ka Flask equivalent.
# Source-of-truth Flask schema mein RawPunchLog hai (PHP wala
# tbl_attendance_summary). Attendance table same hi hai.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# â”€â”€ Local biometric device API config â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# PHP code ke hisab se hardcoded; future me settings master me jaane chahiye.
DEVICE_API_URL    = "http://192.168.2.2:82/api/v2/WebAPI/GetDeviceLogs"
DEVICE_API_KEY    = "242511032625"
PUSH_LIVE_URL     = "https://hcperp.in/api/receive_logs"
PUSH_LIVE_KEY     = "HCP_PUSH_2024"

# â”€â”€ Default employees (9001 & 9002) â€” admin / management staff jo
# device pe punch nahi karte par hamesha Present count hone chahiye.
DEFAULT_EMPLOYEES = {
    '9001': {'type': 'HCP OFFICE', 'in': '10:30:00', 'out': '19:00:00', 'hours': 8.50},
    '9002': {'type': 'HCP OFFICE', 'in': '10:30:00', 'out': '19:00:00', 'hours': 8.50},
}

# â”€â”€ Week-off rule: HCP OFFICE = Sunday off; baki sab = Tuesday off â”€â”€
def _is_week_off(emp_type, dt):
    """dt ko us emp_type ka weekly off hai ya nahi."""
    # Python's weekday(): Monday=0..Sunday=6
    # PHP DAYOFWEEK: Sunday=1, Monday=2, ..., Saturday=7
    wd = dt.weekday()
    if (emp_type or '').upper() == 'HCP OFFICE':
        return wd == 6   # Sunday
    return wd == 1       # Tuesday


def _classify_status(emp_type, att_date, punch_in, punch_out):
    """
    Sync ka core classification logic â€” PHP CASE expression ka direct port.
    Order matters:
      1. WOP    = Weekly off pe valid in+out (>0 mins, valid range)
      2. Present= valid in+out, hours >= 7
      3. Half Day= valid in+out, hours >= 6 (lekin <7)
      4. Absent = valid in+out lekin hours <6
      5. MIS-PUNCH = sirf in ya sirf out, ya in==out
    """
    has_in  = punch_in  is not None
    has_out = punch_out is not None

    # MIS-PUNCH cases
    if has_in and not has_out:                  return 'MIS-PUNCH'
    if not has_in and has_out:                  return 'MIS-PUNCH'
    if has_in and has_out and punch_in == punch_out: return 'MIS-PUNCH'

    if has_in and has_out and punch_out > punch_in:
        hours = (punch_out - punch_in).total_seconds() / 3600.0
        if _is_week_off(emp_type, att_date):
            return 'WOP'
        if hours >= 7.0:  return 'Present'
        if hours >= 6.0:  return 'Half Day'
        return 'Absent'
    return 'MIS-PUNCH'


def _sync_one_date(target_date):
    """
    Ek din ke liye saare employees ka attendance recompute karo.
    Returns: {inserted: int, updated: int, skipped_woff: int}

    OPTIMIZED: Pehle har employee per 4 queries chalti thi (employee lookup,
    in_device, out_device, existing attendance). Total ~1000 queries per date.
    Ab sab bulk-fetch ho jata hai â€” total 4 queries per date, regardless of
    employee count. ~250x faster.
    """
    inserted = 0
    updated  = 0
    skipped  = 0

    # â”€â”€ Step 1: Us din ke saare punches ek hi query me lao â”€â”€
    all_punches = RawPunchLog.query.filter(
        db.func.date(RawPunchLog.log_date) == target_date
    ).order_by(RawPunchLog.employee_code, RawPunchLog.log_date.asc()).all()

    if not all_punches:
        # No punches at all â€” sirf default employees handle karo niche
        emp_groups = {}
    else:
        # Group by employee_code in Python (no DB roundtrip)
        emp_groups = {}
        for p in all_punches:
            emp_groups.setdefault(p.employee_code, []).append(p)

    emp_codes_with_punches = list(emp_groups.keys())

    # â”€â”€ Step 2: Saare employees ek hi query me lao â”€â”€
    # employee_code ya employee_id dono se match
    emp_map = {}
    if emp_codes_with_punches:
        emp_rows = Employee.query.filter(
            db.or_(
                Employee.employee_code.in_(emp_codes_with_punches),
                Employee.employee_id.in_(emp_codes_with_punches),
            )
        ).all()
        for e in emp_rows:
            if e.employee_code: emp_map[e.employee_code] = e
            if e.employee_id:   emp_map[e.employee_id]   = e

    # â”€â”€ Step 3: Us date ke saare existing attendance records lao â”€â”€
    all_codes = list(emp_codes_with_punches) + list(DEFAULT_EMPLOYEES.keys())
    existing_atts = {}
    if all_codes:
        att_rows = Attendance.query.filter(
            Attendance.attendance_date == target_date,
            Attendance.employee_code.in_(all_codes),
        ).all()
        for a in att_rows:
            existing_atts[a.employee_code] = a

    # â”€â”€ Step 4: Process each employee with punches â”€â”€
    for emp_code, punches in emp_groups.items():
        first_punch = punches[0]   # already sorted by log_date asc
        last_punch  = punches[-1]

        first_in = first_punch.log_date
        in_dev   = first_punch.serial_number

        actual_out = None
        out_dev    = None
        if last_punch.log_date != first_in:
            actual_out = last_punch.log_date
            out_dev    = last_punch.serial_number

        emp = emp_map.get(emp_code)
        emp_type = (emp.employee_type if emp else '') or ''

        status = _classify_status(emp_type, target_date, first_in, actual_out)
        total_hours = None
        if first_in and actual_out and actual_out > first_in:
            total_hours = round((actual_out - first_in).total_seconds() / 3600.0, 2)

        att = existing_atts.get(emp_code)
        if att is None:
            att = Attendance(employee_code=emp_code, attendance_date=target_date)
            db.session.add(att)
            inserted += 1
        else:
            updated += 1

        att.punch_in    = first_in
        att.punch_out   = actual_out
        att.in_device   = in_dev
        att.out_device  = out_dev
        att.total_hours = total_hours
        att.status      = status
        att.updated_at  = datetime.now()

    # â”€â”€ Step 5: Default employees (9001/9002) â€” existing_atts dict use karo â”€â”€
    for emp_code, conf in DEFAULT_EMPLOYEES.items():
        if _is_week_off(conf['type'], target_date):
            skipped += 1
            continue

        in_time  = datetime.strptime(f"{target_date} {conf['in']}",  "%Y-%m-%d %H:%M:%S")
        out_time = datetime.strptime(f"{target_date} {conf['out']}", "%Y-%m-%d %H:%M:%S")

        existing = existing_atts.get(emp_code)

        # Create-or-fix logic â€” PHP code ke same triggers
        needs_default = False
        if existing is None:
            needs_default = True
        elif existing.status in ('MIS-PUNCH', 'Absent'):
            needs_default = True
        elif (existing.punch_out is None
              or (existing.punch_in and existing.punch_out
                  and existing.punch_out <= existing.punch_in)
              or (existing.punch_in and existing.punch_out
                  and (existing.punch_out - existing.punch_in).total_seconds() < 3600)):
            needs_default = True

        if needs_default:
            if existing is None:
                existing = Attendance(employee_code=emp_code, attendance_date=target_date)
                db.session.add(existing)
                inserted += 1
            else:
                updated += 1
            existing.punch_in    = in_time
            existing.punch_out   = out_time
            existing.in_device   = 'DEFAULT'
            existing.out_device  = 'DEFAULT'
            existing.total_hours = conf['hours']
            existing.status      = 'Present'
            existing.updated_at  = datetime.now()

    return {'inserted': inserted, 'updated': updated, 'skipped_woff': skipped}


# â”€â”€ Sync UI Screen â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@attendance_bp.route('/attendance_log/sync', methods=['GET'])
@login_required
def attendance_log_sync_view():
    if not _require_sub_perm('hr', 'att_sync'):
        flash('Access denied: Sync Data permission nahi hai.', 'error')
        return redirect(url_for('attendance.attendance_dashboard'))
    today = date.today()
    return render_template(
        'hr/attendance/sync.html',
        today_str=today.strftime('%Y-%m-%d'),
        active_page='att_sync',
    )


# â”€â”€ Sync POST endpoint â€” date range process karta hai â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@attendance_bp.route('/attendance_log/sync/run', methods=['POST'])
@login_required
def attendance_log_sync_run():
    if not _require_sub_perm('hr', 'att_sync'):
        return jsonify(success=False, error='Access denied: Sync permission nahi hai.'), 403
    from_date = request.form.get('from_date') or date.today().isoformat()
    to_date   = request.form.get('to_date')   or from_date

    def _parse_flex(s):
        """YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY, MM/DD/YYYY â€” sab try karo."""
        s = (s or '').strip()
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y',
                    '%Y/%m/%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    d_from = _parse_flex(from_date)
    d_to   = _parse_flex(to_date)
    if not d_from or not d_to:
        return jsonify(
            success=False,
            error=f'Invalid date format. Got from={from_date!r}, to={to_date!r}. '
                  f'Use YYYY-MM-DD ya DD-MM-YYYY.'
        ), 400

    if d_to < d_from:
        return jsonify(success=False, error='To Date, From Date se chhoti nahi ho sakti.'), 400

    # Safety cap â€” bahut bada range accidently na chale
    days_count = (d_to - d_from).days + 1
    if days_count > 90:
        return jsonify(success=False, error=f'Max 90 din ek baar me. Tumne {days_count} din maange.'), 400

    log_lines    = []
    grand_ins    = 0
    grand_upds   = 0
    grand_skipw  = 0

    cur = d_from
    while cur <= d_to:
        try:
            res = _sync_one_date(cur)
            # Commit per-date â€” memory clear, progress safe even if next date crashes
            db.session.commit()
            grand_ins   += res['inserted']
            grand_upds  += res['updated']
            grand_skipw += res['skipped_woff']
            log_lines.append(
                f"[{cur.isoformat()}] inserted={res['inserted']}  "
                f"updated={res['updated']}  weekoff_skipped={res['skipped_woff']}"
            )
        except Exception as ex:
            db.session.rollback()
            log_lines.append(f"[{cur.isoformat()}] âŒ ERROR: {ex}")
        cur += timedelta(days=1)

    log_lines.append("")
    log_lines.append(f"âœ… DONE â€” {days_count} day(s) processed")
    log_lines.append(f"   Total inserted : {grand_ins}")
    log_lines.append(f"   Total updated  : {grand_upds}")
    log_lines.append(f"   Weekly off skip: {grand_skipw}")

    return jsonify(
        success  = True,
        from_date= d_from.isoformat(),
        to_date  = d_to.isoformat(),
        inserted = grand_ins,
        updated  = grand_upds,
        skipped  = grand_skipw,
        log      = log_lines,
    )


# â”€â”€ Fetch-from-device API (manual trigger) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# PHP `push_to_python()` ka Flask equivalent. Yeh local LAN biometric
# device API ko call karta hai aur logs ko seedha RawPunchLog me bhar
# deta hai (PHP wala "push to live server" step skip â€” kyunki yahin
# server hai). Agar future me 2 servers chalane hon, to PUSH_LIVE_URL
# par forward karne ka switch turn on kar sakte hain.
@attendance_bp.route('/attendance_log/fetch_device', methods=['POST'])
@login_required
def attendance_log_fetch_device():
    import urllib.request, urllib.error, urllib.parse

    from_date = request.form.get('from_date') or date.today().replace(day=1).isoformat()
    to_date   = request.form.get('to_date')   or date.today().isoformat()
    forward   = request.form.get('forward_to_live') in ('1', 'true', 'on', 'yes')

    qs  = urllib.parse.urlencode({
        'APIKey'  : DEVICE_API_KEY,
        'FromDate': from_date,
        'ToDate'  : to_date,
    })
    url = f"{DEVICE_API_URL}?{qs}"

    log_lines = [f"â†’ GET {url}"]

    # 1. Fetch from device
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'HCP-ERP-Sync/1.0'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode('utf-8', errors='replace')
        log_lines.append(f"â† HTTP {resp.status}, {len(raw)} bytes")
        try:
            payload = json.loads(raw)
        except Exception:
            return jsonify(success=False, error='Device API ne valid JSON nahi diya',
                           log=log_lines), 502
    except urllib.error.URLError as ex:
        log_lines.append(f"âŒ Device API reachable nahi: {ex}")
        return jsonify(success=False, error=f'Device API connect nahi hua: {ex}',
                       log=log_lines), 502
    except Exception as ex:
        log_lines.append(f"âŒ Unexpected error: {ex}")
        return jsonify(success=False, error=f'Fetch failed: {ex}', log=log_lines), 500

    if not payload:
        return jsonify(success=False, error='Device se empty data', log=log_lines), 200

    # 2. Forward to live server (optional, mirrors PHP push_to_python)
    if forward:
        try:
            push_data = json.dumps(payload).encode('utf-8')
            push_req  = urllib.request.Request(
                PUSH_LIVE_URL,
                data=push_data,
                headers={
                    'Content-Type' : 'application/json',
                    'Authorization': PUSH_LIVE_KEY,
                },
                method='POST',
            )
            with urllib.request.urlopen(push_req, timeout=60) as push_resp:
                push_body = push_resp.read().decode('utf-8', errors='replace')
            log_lines.append(f"â†’ Pushed to live server: HTTP {push_resp.status}")
            log_lines.append(f"  â†³ {push_body[:200]}")
            return jsonify(success=True, mode='forwarded',
                           live_response=push_body, log=log_lines)
        except Exception as ex:
            log_lines.append(f"âŒ Forward to live failed: {ex}")
            return jsonify(success=False, error=f'Forward failed: {ex}',
                           log=log_lines), 502

    # 3. Local-mode: store directly into RawPunchLog
    if isinstance(payload, dict):
        logs_list = (payload.get('data') or payload.get('logs') or
                     payload.get('DeviceLogs') or payload.get('Records') or [payload])
    elif isinstance(payload, list):
        logs_list = payload
    else:
        return jsonify(success=False, error='Unexpected payload shape', log=log_lines), 502

    inserted = 0; skipped = 0; errors = []
    for entry in logs_list:
        try:
            emp_code = _get_field(entry,
                'EmployeeCode', 'employee_code', 'CardNo', 'card_no',
                'UserID', 'UserId', 'EmpCode', 'emp_code')
            log_dt = _parse_datetime(_get_field(entry,
                'LogTime', 'log_date', 'PunchTime', 'DateTime', 'datetime', 'Time'))
            if not emp_code or not log_dt:
                skipped += 1; continue

            if RawPunchLog.query.filter_by(employee_code=emp_code, log_date=log_dt).first():
                skipped += 1; continue

            db.session.add(RawPunchLog(
                employee_code=emp_code, log_date=log_dt,
                serial_number=_get_field(entry, 'SerialNumber', 'serial_number',
                                         'DeviceId', 'MachineNo'),
                punch_direction=_get_punch_direction(entry),
                temperature=_get_field(entry, 'Temperature') or 0.00,
                temperature_state=_get_field(entry, 'TemperatureState'),
                synced_at=datetime.now(),
            ))
            inserted += 1
        except Exception as ex:
            errors.append(str(ex))

    try:
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return jsonify(success=False, error=f'DB error: {ex}', log=log_lines), 500

    log_lines.append(f"âœ… Inserted: {inserted}, Skipped (dup): {skipped}, "
                     f"Errors: {len(errors)}")
    return jsonify(success=True, mode='local',
                   inserted=inserted, skipped=skipped,
                   total=len(logs_list), errors=errors[:5], log=log_lines)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DAILY ATTENDANCE VIEW â€” /hr/attendance/daily
# Image 2 ka design â€” date picker, employee type filter, stat chips, table
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@attendance_bp.route('/hr/attendance/daily')
@login_required
def attendance_daily():
    if not _require_sub_perm('hr', 'att_daily'):
        flash('Access denied: Daily Attendance permission nahi hai.', 'error')
        return redirect(url_for('attendance.attendance_dashboard'))
    sel_date_str = request.args.get('date', date.today().isoformat())
    try:
        sel_date = datetime.strptime(sel_date_str, '%Y-%m-%d').date()
    except ValueError:
        sel_date = date.today()

    search_q = (request.args.get('q') or '').strip().lower()
    emp_type = (request.args.get('emp_type') or '').strip()

    # Sab active employees fetch karo + us din ka attendance
    emp_q = Employee.query.filter_by(status='active')
    if emp_type:
        emp_q = emp_q.filter(Employee.employee_type == emp_type)
    employees = emp_q.order_by(Employee.first_name).all()

    att_map = {a.employee_code: a for a in Attendance.query.filter_by(
        attendance_date=sel_date).all()}

    # ── Holidays (location-wise) + approved leaves for this day ──
    from models.attendance import HolidayMaster
    day_hols = HolidayMaster.query.filter(
        HolidayMaster.is_active.is_(True),
        HolidayMaster.holiday_date == sel_date,
    ).all()

    def _emp_zone(emp):
        loc = (emp.location or '').strip().lower()
        if loc:
            return 'HO' if ('office' in loc or loc in ('ho', 'h.o', 'h.o.')) else 'Plant'
        return 'HO' if (emp.employee_type or '').strip().upper() == 'HCP OFFICE' else 'Plant'

    def _holiday_for(emp):
        zone = _emp_zone(emp)
        for h in day_hols:
            if getattr(h, 'location', 'All') in ('All', zone):
                return h
        return None

    from models.hr_rules import HRLeaveApplication
    day_leaves = HRLeaveApplication.query.filter(
        HRLeaveApplication.status == 'approved',
        HRLeaveApplication.from_date <= sel_date,
        HRLeaveApplication.to_date   >= sel_date,
    ).all()
    leave_by_emp = {l.employee_id: l for l in day_leaves}

    rows = []
    for emp in employees:
        # Match by employee_code OR employee_id
        att = att_map.get(emp.employee_code) or att_map.get(emp.employee_id)
        full_name = f"{emp.first_name or ''}{('_' + emp.last_name) if emp.last_name else ''}"
        if search_q:
            hay = ' '.join(filter(None, [emp.first_name, emp.last_name,
                                          emp.employee_code, emp.employee_id,
                                          emp.department or '',
                                          emp.designation or ''])).lower()
            if search_q not in hay:
                continue
        # Status resolve — att row ho to wahi, warna Holiday > Leave > Week Off > Absent
        note = ''
        if att:
            status = att.status
            if status == 'Present':
                h = _holiday_for(emp)
                if h:
                    status = 'HLP'; note = h.title   # Holiday par kaam
            elif status == 'Holiday':
                h = _holiday_for(emp)
                note = h.title if h else ''
            elif status == 'Absent':
                # Absent row par bhi Holiday/Leave/WeekOff override (monthly jaisa)
                h  = _holiday_for(emp)
                lv = leave_by_emp.get(emp.id)
                if h:
                    status = 'Holiday';  note = h.title
                elif lv:
                    status = 'On Leave'; note = lv.type_label
                elif _is_week_off(emp.employee_type or '', sel_date):
                    status = 'Week Off'
        else:
            h  = _holiday_for(emp)
            lv = leave_by_emp.get(emp.id)
            if h:
                status = 'Holiday';  note = h.title
            elif lv:
                status = 'On Leave'; note = lv.type_label
            elif _is_week_off(emp.employee_type or '', sel_date):
                status = 'Week Off'
            else:
                status = 'Absent'
        rows.append({
            'emp':         emp,
            'full_name':   full_name,
            'att':         att,
            'status':      status,
            'note':        note,
        })

    # Stats
    total       = len(rows)
    present_n   = sum(1 for r in rows if r['status'] == 'Present')
    halfday_n   = sum(1 for r in rows if r['status'] == 'Half Day')
    mispunch_n  = sum(1 for r in rows if r['status'] == 'MIS-PUNCH')
    absent_n    = sum(1 for r in rows if r['status'] == 'Absent')
    wop_n       = sum(1 for r in rows if r['status'] == 'WOP')
    holiday_n   = sum(1 for r in rows if r['status'] == 'Holiday')
    hlp_n       = sum(1 for r in rows if r['status'] == 'HLP')
    leave_n     = sum(1 for r in rows if r['status'] == 'On Leave')
    weekoff_n   = sum(1 for r in rows if r['status'] == 'Week Off')

    # Distinct employee types for the filter dropdown
    type_rows = db.session.query(Employee.employee_type).filter(
        Employee.employee_type.isnot(None),
        Employee.employee_type != '',
        Employee.status == 'active',
    ).distinct().order_by(Employee.employee_type).all()
    emp_types = [t[0] for t in type_rows if t[0]]

    return render_template('hr/attendance/daily.html',
        rows=rows, sel_date=sel_date, search_q=search_q, emp_type=emp_type,
        emp_types=emp_types,
        stats=dict(total=total, present=present_n, halfday=halfday_n,
                   mispunch=mispunch_n, absent=absent_n, wop=wop_n,
                   holiday=holiday_n, leave=leave_n, weekoff=weekoff_n,
                   hlp=hlp_n),
        prev_date=(sel_date - timedelta(days=1)).isoformat(),
        next_date=(sel_date + timedelta(days=1)).isoformat(),
        today_str=date.today().isoformat(),
        active_page='att_daily',
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MONTHLY ATTENDANCE VIEW â€” /hr/attendance/monthly
# Image 3 ka design â€” year/month, per-employee day-wise grid
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _build_monthly_data(year, month, emp_type, search_q):
    """Monthly view + Excel export — shared data builder."""
    import calendar as _cal
    days_in_month = _cal.monthrange(year, month)[1]
    month_start   = date(year, month, 1)
    month_end     = date(year, month, days_in_month)

    # Working days = month days minus weekly offs (Sundays for HCP OFFICE
    # logic could be per-employee; here we use a conservative "Mon-Sat = working")
    working_days = sum(1 for i in range(days_in_month)
                       if (month_start + timedelta(days=i)).weekday() != 6)

    emp_q = Employee.query.filter_by(status='active')
    if emp_type:
        emp_q = emp_q.filter(Employee.employee_type == emp_type)
    if search_q:
        like = f'%{search_q}%'
        emp_q = emp_q.filter(db.or_(
            Employee.first_name.ilike(like),
            Employee.last_name.ilike(like),
            Employee.employee_code.ilike(like),
            Employee.employee_id.ilike(like),
            Employee.department.ilike(like),
        ))
    employees = emp_q.order_by(Employee.first_name).all()

    # Bulk attendance fetch for the month
    att_records = Attendance.query.filter(
        Attendance.attendance_date >= month_start,
        Attendance.attendance_date <= month_end,
    ).all()
    # group by emp_code â†’ {date: att}
    att_by_emp = {}
    for a in att_records:
        att_by_emp.setdefault(a.employee_code, {})[a.attendance_date] = a

    # ── Holidays (location-wise: All / HO / Plant) ──
    from models.attendance import HolidayMaster
    hol_rows = HolidayMaster.query.filter(
        HolidayMaster.is_active.is_(True),
        HolidayMaster.holiday_date >= month_start,
        HolidayMaster.holiday_date <= month_end,
    ).all()
    day_hols = {}
    for h in hol_rows:
        day_hols.setdefault(h.holiday_date, []).append(h)

    def _emp_zone(emp):
        """Employee kis zone mein hai — HO ya Plant."""
        loc = (emp.location or '').strip().lower()
        if loc:
            return 'HO' if ('office' in loc or loc in ('ho', 'h.o', 'h.o.')) else 'Plant'
        return 'HO' if (emp.employee_type or '').strip().upper() == 'HCP OFFICE' else 'Plant'

    def _holiday_for(emp, d):
        zone = _emp_zone(emp)
        for h in day_hols.get(d, []):
            if getattr(h, 'location', 'All') in ('All', zone):
                return h
        return None

    # ── Approved leaves (month overlap) ──
    from models.hr_rules import HRLeaveApplication
    leave_rows = HRLeaveApplication.query.filter(
        HRLeaveApplication.status == 'approved',
        HRLeaveApplication.from_date <= month_end,
        HRLeaveApplication.to_date   >= month_start,
    ).all()
    leaves_by_emp = {}
    for l in leave_rows:
        leaves_by_emp.setdefault(l.employee_id, []).append(l)

    def _leave_for(emp, d):
        for l in leaves_by_emp.get(emp.id, []):
            if l.from_date <= d <= l.to_date:
                return l
        return None

    # Build per-employee rows
    emp_rows = []
    day_list = [date(year, month, d) for d in range(1, days_in_month + 1)]
    type_rows = db.session.query(Employee.employee_type).filter(
        Employee.employee_type.isnot(None), Employee.employee_type != '',
    ).distinct().order_by(Employee.employee_type).all()
    emp_types = [t[0] for t in type_rows if t[0]]

    for emp in employees:
        per_emp = att_by_emp.get(emp.employee_code) or att_by_emp.get(emp.employee_id) or {}
        days = []
        cnt = dict(P=0, HD=0, AB=0, MP=0, WO=0, WOP=0, HOL=0, LV=0, HLP=0)
        for d in day_list:
            a    = per_emp.get(d)
            wo   = _is_week_off(emp.employee_type, d)
            hol  = _holiday_for(emp, d)
            lv   = _leave_for(emp, d)
            note = ''
            lv_code = ''
            if a is None:
                # Priority: Holiday > Leave > Week Off > Absent
                if hol:
                    code = 'HOL'; note = hol.title
                elif lv:
                    code = 'LV';  note = lv.type_label; lv_code = lv.leave_type
                elif wo:
                    code = 'WO'
                else:
                    code = 'AB'
                in_t = out_t = ''
                tot = 0.0
            else:
                if a.status == 'Present':
                    if hol:
                        code = 'HLP'; note = hol.title   # Holiday par kaam
                    else:
                        code = 'P'
                elif a.status == 'Half Day':    code = 'HD'
                elif a.status == 'MIS-PUNCH':   code = 'MP'
                elif a.status == 'WOP':         code = 'WOP'
                elif a.status == 'Holiday':     code = 'HOL'; note = hol.title if hol else 'Holiday'
                elif a.status == 'Absent':
                    if hol:  code = 'HOL'; note = hol.title
                    elif lv: code = 'LV';  note = lv.type_label; lv_code = lv.leave_type
                    elif wo: code = 'WO'
                    else:    code = 'AB'
                else:                            code = 'AB'
                in_t  = a.punch_in.strftime('%H:%M')  if a.punch_in  else ''
                out_t = a.punch_out.strftime('%H:%M') if a.punch_out else ''
                tot   = float(a.total_hours or 0)
            cnt[code] = cnt.get(code, 0) + 1
            days.append(dict(date=d, in_t=in_t, out_t=out_t, tot=tot, code=code, note=note, lv_code=lv_code))
        emp_rows.append(dict(emp=emp, days=days, cnt=cnt))
    return dict(emp_rows=emp_rows, day_list=day_list, emp_types=emp_types,
                working_days=working_days, days_in_month=days_in_month,
                month_start=month_start, month_end=month_end,
                employees=employees)


@attendance_bp.route('/hr/attendance/monthly')
@login_required
def attendance_monthly():
    if not _require_sub_perm('hr', 'att_monthly'):
        flash('Access denied: Monthly Attendance permission nahi hai.', 'error')
        return redirect(url_for('attendance.attendance_dashboard'))
    import calendar as _cal
    today = date.today()
    try:
        year  = int(request.args.get('year',  today.year))
        month = int(request.args.get('month', today.month))
    except ValueError:
        year, month = today.year, today.month

    if not (1 <= month <= 12) or year < 2000 or year > 2100:
        year, month = today.year, today.month

    search_q = (request.args.get('q') or '').strip().lower()
    emp_type = (request.args.get('emp_type') or '').strip()

    _md = _build_monthly_data(year, month, emp_type, search_q)
    emp_rows      = _md['emp_rows']
    day_list      = _md['day_list']
    emp_types     = _md['emp_types']
    working_days  = _md['working_days']
    employees     = _md['employees']

    return render_template('hr/attendance/monthly.html',
        year=year, month=month, month_name=_cal.month_name[month],
        emp_rows=emp_rows, working_days=working_days,
        total_emps=len(employees), emp_types=emp_types,
        emp_type=emp_type, search_q=search_q,
        prev_year=(year if month > 1 else year - 1),
        prev_month=(month - 1 if month > 1 else 12),
        next_year=(year if month < 12 else year + 1),
        next_month=(month + 1 if month < 12 else 1),
        active_page='att_monthly',
    )


# ══════════════════════════════════════════════════════════════
# MONTHLY ATTENDANCE — EXCEL EXPORT
# GET /hr/attendance/monthly/export
# ══════════════════════════════════════════════════════════════
@attendance_bp.route('/hr/attendance/monthly/export')
@login_required
def attendance_monthly_export():
    """HCP format Excel — Sheet1: Attendance (In/Out/Shift/Status blocks),
    Sheet2: Late Coming (late times + penalty slabs)."""
    if not _require_sub_perm('hr', 'att_monthly'):
        flash('Access denied: Monthly Attendance permission nahi hai.', 'error')
        return redirect(url_for('attendance.attendance_dashboard'))
    import calendar as _cal
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from models.attendance import LateShiftRule, LatePenaltyRule

    today = date.today()
    try:
        year  = int(request.args.get('year',  today.year))
        month = int(request.args.get('month', today.month))
    except ValueError:
        year, month = today.year, today.month
    if not (1 <= month <= 12) or year < 2000 or year > 2100:
        year, month = today.year, today.month
    search_q = (request.args.get('q') or '').strip().lower()
    emp_type = (request.args.get('emp_type') or '').strip()

    md = _build_monthly_data(year, month, emp_type, search_q)
    emp_rows, day_list = md['emp_rows'], md['day_list']
    n_days     = len(day_list)
    month_name = _cal.month_name[month]

    # ── Styles (sample file se exact) ──
    F  = lambda hexc: PatternFill('solid', start_color=hexc)
    ctr  = Alignment(horizontal='center', vertical='center')
    ctrw = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='B0B0B0')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    TITLE_FILL  = F('F4B084'); HDR_FILL = F('D9E1F2'); SUMM_FILL = F('DEEAF1')
    IN_FILL = F('E2EFDA'); OUT_FILL = F('FCE4D6'); SHIFT_FILL = F('EBF3FB')
    ST_STYLE = {                       # status -> (fill, font color)
        'P':  ('00B050', 'FFFFFF'), 'WO': ('BFBFBF', '000000'),
        'WOP':('FFFF00', '000000'), 'HD': ('00B0F0', 'FFFFFF'),
        'AB': ('FF0000', 'FFFFFF'), 'MP': ('FFC000', '000000'),
        'PH': ('ED7D31', 'FFFFFF'), 'HLP': ('E11D48', 'FFFFFF'),
        'CL': ('0070C0', 'FFFFFF'),
        'SL': ('7030A0', 'FFFFFF'), 'PL': ('0070C0', 'FFFFFF'),
        'LOP':('C00000', 'FFFFFF'),
    }
    SUMMARY_HDRS = ['Total Present', 'Paid Holiday', 'Total Absent', 'Total Half day',
                    'Total Paid leave', 'Total Casual leave', 'Total Sick leave',
                    'Total week off', 'Punch Incomplete', 'Week off Present',
                    'Holiday Present',
                    '8 Hrs', '10 Hrs', '12 Hrs', '16 Hrs', '24 Hrs', 'Half day']
    SUMMARY_CODE = ['P', 'PH', 'AB', 'HD', 'PL', 'CL', 'SL', 'WO', 'MP', 'WOP', 'HLP']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Worksheet'

    FIX = 7                                            # Sr, Punch ID, Period, Name, Desig, Dept, Remarks
    total_cols = FIX + n_days + len(SUMMARY_HDRS)

    # ── Title rows (1-3) ──
    titles = ['HCP Wellness Pvt. Ltd.',
              f'ATTENDANCE SHEET FOR {month_name.upper()} {year}',
              'Address: 403, Maruti Elanza Vertex, Opp. Sterling Hospital, Ahmedabad']
    for r, txt in enumerate(titles, start=1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(bold=True); c.fill = TITLE_FILL; c.alignment = ctr

    # ── Header rows (4-5) ──
    fixed_hdrs = ['Sr.', 'New Punch ID', 'Period', 'Employee Name',
                  'Designation', 'Department', 'Remarks']
    for c, h in enumerate(fixed_hdrs, start=1):
        ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HDR_FILL; cell.alignment = ctr; cell.border = bord
    for i, d in enumerate(day_list):
        c1 = ws.cell(row=4, column=FIX + 1 + i, value=d.day)
        c2 = ws.cell(row=5, column=FIX + 1 + i, value=d.strftime('%a'))
        for cell in (c1, c2):
            cell.fill = HDR_FILL; cell.alignment = ctr; cell.border = bord
    for j, h in enumerate(SUMMARY_HDRS):
        col = FIX + n_days + 1 + j
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = HDR_FILL; cell.alignment = ctrw; cell.border = bord
    ws.row_dimensions[4].height = 30

    # ── Employee blocks (4 rows each: In/Out/Shift/Status) ──
    def daycode(dd):
        if dd['code'] == 'HOL': return 'PH'
        if dd['code'] == 'LV':  return dd.get('lv_code') or 'PL'
        return dd['code']

    r = 6
    for sr, er in enumerate(emp_rows, start=1):
        emp = er['emp']
        rs, re_ = r, r + 3
        fixed_vals = [sr, emp.employee_id or emp.employee_code or '',
                      'Probation' if emp.is_probation else 'Permanent',
                      emp.full_name, emp.designation or '', emp.department or '']
        for c, v in enumerate(fixed_vals, start=1):
            ws.merge_cells(start_row=rs, start_column=c, end_row=re_, end_column=c)
            cell = ws.cell(row=rs, column=c, value=v)
            cell.alignment = ctrw; cell.border = bord
        for rr, (lbl, fill) in zip(range(rs, re_ + 1),
                [('In Time', IN_FILL), ('Out Time', OUT_FILL),
                 ('Shift', SHIFT_FILL), ('Status', None)]):
            g = ws.cell(row=rr, column=7, value=lbl)
            g.border = bord
            if fill: g.fill = fill
        for i, dd in enumerate(er['days']):
            col = FIX + 1 + i
            cin  = ws.cell(row=rs,     column=col, value=dd['in_t'] or None)
            cout = ws.cell(row=rs + 1, column=col, value=dd['out_t'] or None)
            csh  = ws.cell(row=rs + 2, column=col,
                           value=round(dd['tot'], 2) if dd['tot'] else 0)
            code = daycode(dd)
            cst  = ws.cell(row=rs + 3, column=col, value=code)
            cin.fill, cout.fill, csh.fill = IN_FILL, OUT_FILL, SHIFT_FILL
            bg, fg = ST_STYLE.get(code, ('FFFFFF', '000000'))
            cst.fill = F(bg); cst.font = Font(bold=True, color=fg)
            for cell in (cin, cout, csh, cst):
                cell.alignment = ctr; cell.border = bord
            if dd.get('note'):
                from openpyxl.comments import Comment
                cst.comment = Comment(dd['note'], 'HCP ERP')
        # ── Summary (merged 4 rows, COUNTIF formulas) ──
        d1 = get_column_letter(FIX + 1); d2 = get_column_letter(FIX + n_days)
        strow, shrow = rs + 3, rs + 2
        for j, h in enumerate(SUMMARY_HDRS):
            col = FIX + n_days + 1 + j
            ws.merge_cells(start_row=rs, start_column=col, end_row=re_, end_column=col)
            if j < len(SUMMARY_CODE):
                fml = f'=COUNTIF({d1}{strow}:{d2}{strow},"{SUMMARY_CODE[j]}")'
            elif h == '8 Hrs':
                fml = f'=COUNTIFS({d1}{shrow}:{d2}{shrow},">=8",{d1}{shrow}:{d2}{shrow},"<10")'
            elif h == '10 Hrs':
                fml = f'=COUNTIFS({d1}{shrow}:{d2}{shrow},">=10",{d1}{shrow}:{d2}{shrow},"<12")'
            elif h == '12 Hrs':
                fml = f'=COUNTIFS({d1}{shrow}:{d2}{shrow},">=12",{d1}{shrow}:{d2}{shrow},"<16")'
            elif h == '16 Hrs':
                fml = f'=COUNTIFS({d1}{shrow}:{d2}{shrow},">=16",{d1}{shrow}:{d2}{shrow},"<24")'
            elif h == '24 Hrs':
                fml = f'=COUNTIF({d1}{shrow}:{d2}{shrow},">=24")'
            else:   # Half day (hours-wise 4–8)
                fml = f'=COUNTIFS({d1}{shrow}:{d2}{shrow},">=4",{d1}{shrow}:{d2}{shrow},"<8")'
            cell = ws.cell(row=rs, column=col, value=fml)
            cell.fill = SUMM_FILL; cell.alignment = ctr; cell.border = bord
        r += 4

    # widths + freeze
    for col, w in zip('ABCDEFG', [4, 12, 11, 20, 16, 13, 9]):
        ws.column_dimensions[col].width = w
    for i in range(n_days):
        ws.column_dimensions[get_column_letter(FIX + 1 + i)].width = 6
    for j in range(len(SUMMARY_HDRS)):
        ws.column_dimensions[get_column_letter(FIX + n_days + 1 + j)].width = 8
    ws.freeze_panes = 'H6'

    # ════════════════ Sheet 2: LATE COMING ════════════════
    ws2 = wb.create_sheet('Late Coming')
    LFIX = 4
    l_total = LFIX + n_days + 4

    # type-wise late rules + penalty rates
    rules = {r_.employee_type: r_ for r_ in LateShiftRule.query.filter_by(is_active=True).all()}
    def _rates(rule):
        r1 = r2 = r3 = None; band_end = None
        if rule:
            for s in rule.penalty_rules:
                if not s.is_active: continue
                if s.time_to:                      # band slab
                    band_end = s.time_to
                    if s.from_count <= 3 <= (s.to_count or 999): r1 = float(s.penalty_amount)
                    if s.from_count >= 4: r2 = float(s.penalty_amount)
                else:
                    r3 = float(s.penalty_amount)   # after-band slab
        return (r1 or 120, r2 or 250, r3 or 250,
                band_end or '10:59',
                rule.late_after if rule else '10:46')

    for r_, txt in enumerate(['HCP Wellness Pvt. Ltd.',
                              f'LATE COMING FOR THE {month_name.upper()} {year}'], start=1):
        ws2.merge_cells(start_row=r_, start_column=1, end_row=r_, end_column=l_total)
        c = ws2.cell(row=r_, column=1, value=txt)
        c.font = Font(bold=True); c.fill = TITLE_FILL; c.alignment = ctr

    for c, h in enumerate(['Sr.No.', 'Employee Name', 'Designation', 'Department'], start=1):
        ws2.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
        cell = ws2.cell(row=4, column=c, value=h)
        cell.fill = HDR_FILL; cell.alignment = ctr; cell.border = bord
    for i, d in enumerate(day_list):
        c1 = ws2.cell(row=4, column=LFIX + 1 + i, value=d.day)
        c2 = ws2.cell(row=5, column=LFIX + 1 + i, value=d.strftime('%a'))
        for cell in (c1, c2):
            cell.fill = HDR_FILL; cell.alignment = ctr; cell.border = bord
    late_hdrs = ['1st 3 time (band) - slab 1', 'After 3 times (band) - slab 2',
                 'After band time - slab 3', 'Total Amount']
    for j, h in enumerate(late_hdrs):
        col = LFIX + n_days + 1 + j
        ws2.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        cell = ws2.cell(row=4, column=col, value=h)
        cell.fill = SUMM_FILL; cell.alignment = ctrw; cell.border = bord
    ws2.row_dimensions[4].height = 42

    red = Font(color='FF0000')
    for sr, er in enumerate(emp_rows, start=1):
        emp = er['emp']
        rr = 5 + sr
        rule = rules.get(emp.employee_type or '')
        r1, r2, r3, band_end, late_after = _rates(rule)
        band_n = after_n = 0
        for c, v in enumerate([sr, emp.full_name, emp.designation or '',
                               emp.department or ''], start=1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.border = bord
        for i, dd in enumerate(er['days']):
            cell = ws2.cell(row=rr, column=LFIX + 1 + i)
            cell.border = bord
            it = dd['in_t']
            if it and it > late_after:             # late aaya
                cell.value = it; cell.font = red; cell.alignment = ctr
                if it <= band_end: band_n += 1
                else:              after_n += 1
        aj = min(band_n, 3); ak = max(band_n - 3, 0)
        cAJ = get_column_letter(LFIX + n_days + 1)
        cAK = get_column_letter(LFIX + n_days + 2)
        cAL = get_column_letter(LFIX + n_days + 3)
        vals = [aj or None, ak or None, after_n or None]
        for j, v in enumerate(vals):
            cell = ws2.cell(row=rr, column=LFIX + n_days + 1 + j, value=v)
            cell.alignment = ctr; cell.border = bord
        tcell = ws2.cell(row=rr, column=LFIX + n_days + 4,
            value=f'=N({cAJ}{rr})*{r1}+N({cAK}{rr})*{r2}+N({cAL}{rr})*{r3}')
        tcell.alignment = ctr; tcell.border = bord
        tcell.font = Font(bold=True)

    for col, w in zip('ABCD', [6, 20, 16, 13]):
        ws2.column_dimensions[col].width = w
    for i in range(n_days):
        ws2.column_dimensions[get_column_letter(LFIX + 1 + i)].width = 6.5
    for j in range(4):
        ws2.column_dimensions[get_column_letter(LFIX + n_days + 1 + j)].width = 13
    ws2.freeze_panes = 'E6'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
        download_name=f'HCP_Attendance_Report_{month_name}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


